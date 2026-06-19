from __future__ import annotations

import re
import unicodedata
from io import BytesIO
from typing import List, Optional
from urllib.parse import quote

from fastapi import HTTPException
from sqlalchemy.orm import Session

from app.models import Cliente
from app.produtos_models import Marca, PedidoCompra, Produto

PEDIDO_EXPORT_COLUNAS_META = {
    "codigo": {"titulo": "Codigo", "excel_width": 14, "pdf_width_mm": 22},
    "produto": {"titulo": "Descricao", "excel_width": 42, "pdf_width_mm": 68},
    "quantidade": {"titulo": "Quantidade", "excel_width": 14, "pdf_width_mm": 20},
    "preco_unitario": {"titulo": "Custo Unit.", "excel_width": 16, "pdf_width_mm": 24},
    "desconto": {"titulo": "Desconto", "excel_width": 14, "pdf_width_mm": 18},
    "total": {"titulo": "Total", "excel_width": 16, "pdf_width_mm": 22},
}
PEDIDO_EXPORT_COLUNAS_DEFAULT = list(PEDIDO_EXPORT_COLUNAS_META.keys())
PEDIDO_EXPORT_COLUNAS_FINANCEIRAS = {"preco_unitario", "desconto", "total"}


def _buscar_fornecedor_pedido(
    db: Session, tenant_id: int, pedido: PedidoCompra
) -> Optional[Cliente]:
    return (
        db.query(Cliente)
        .filter(Cliente.id == pedido.fornecedor_id, Cliente.tenant_id == tenant_id)
        .first()
    )


def _formatar_token_nome(texto: str) -> str:
    return " ".join(
        token.capitalize() if token.isupper() else token for token in str(texto).split()
    )


def _normalizar_texto_nome_arquivo(texto: Optional[str]) -> str:
    if not texto:
        return ""

    texto_ascii = (
        unicodedata.normalize("NFKD", str(texto))
        .encode("ascii", "ignore")
        .decode("ascii")
    )
    texto_limpo = re.sub(r"[^A-Za-z0-9._ -]+", " ", texto_ascii)
    texto_limpo = re.sub(r"\s+", " ", texto_limpo).strip(" ._-")
    return texto_limpo


def _extrair_nome_curto_fornecedor(fornecedor_nome: Optional[str]) -> str:
    nome_limpo = _normalizar_texto_nome_arquivo(fornecedor_nome)
    if not nome_limpo:
        return "Fornecedor"

    tokens = nome_limpo.split()
    ignorar = {"DE", "DA", "DO", "DAS", "DOS", "E", "LTDA", "ME", "EIRELI", "S/A", "SA"}

    for token in tokens:
        if token.upper() not in ignorar:
            return _formatar_token_nome(token)

    return _formatar_token_nome(tokens[0])


def _buscar_nome_marca_pedido(
    db: Session, tenant_id: int, pedido: PedidoCompra
) -> Optional[str]:
    produto_ids = [item.produto_id for item in pedido.itens if item.produto_id]
    if not produto_ids:
        return None

    marcas = (
        db.query(Marca.nome)
        .join(Produto, Produto.marca_id == Marca.id)
        .filter(
            Produto.id.in_(produto_ids),
            Produto.tenant_id == tenant_id,
            Marca.tenant_id == tenant_id,
        )
        .distinct()
        .all()
    )

    nomes = sorted(
        {
            _normalizar_texto_nome_arquivo(resultado[0])
            for resultado in marcas
            if resultado[0]
        }
    )

    if len(nomes) != 1:
        return None

    return _formatar_token_nome(nomes[0])


def _montar_nome_arquivo_pedido(
    pedido: PedidoCompra,
    fornecedor_nome: Optional[str],
    db: Session,
    tenant_id: int,
    extensao: str,
) -> str:
    numero_pedido = _normalizar_texto_nome_arquivo(
        pedido.numero_pedido or pedido.id
    ) or str(pedido.id)
    fornecedor_curto = _extrair_nome_curto_fornecedor(fornecedor_nome)
    marca_nome = _buscar_nome_marca_pedido(db, tenant_id, pedido)

    partes = ["Pedido", numero_pedido, fornecedor_curto]
    if marca_nome:
        partes.append(marca_nome)

    base = " ".join(parte for parte in partes if parte).strip()
    return f"{base}.{extensao}"


def _montar_content_disposition_attachment(filename: str) -> str:
    fallback = _normalizar_texto_nome_arquivo(filename) or "arquivo"
    return f"attachment; filename=\"{fallback}\"; filename*=UTF-8''{quote(filename)}"


def _normalizar_colunas_exportacao_pedido(colunas: Optional[object]) -> List[str]:
    if colunas is None:
        return PEDIDO_EXPORT_COLUNAS_DEFAULT.copy()

    if isinstance(colunas, str):
        candidatos = colunas.split(",")
    elif isinstance(colunas, (list, tuple, set)):
        candidatos = list(colunas)
    else:
        candidatos = [colunas]

    validas = []
    for coluna in candidatos:
        chave = str(coluna or "").strip().lower()
        if chave in PEDIDO_EXPORT_COLUNAS_META and chave not in validas:
            validas.append(chave)

    return validas or PEDIDO_EXPORT_COLUNAS_DEFAULT.copy()


def _exportacao_pedido_tem_colunas_financeiras(colunas_exportacao: List[str]) -> bool:
    return any(
        coluna in PEDIDO_EXPORT_COLUNAS_FINANCEIRAS for coluna in colunas_exportacao
    )


def _formatar_quantidade_documento(valor: Optional[float]) -> str:
    numero = float(valor or 0)
    texto = f"{numero:.2f}"
    if "." in texto:
        texto = texto.rstrip("0").rstrip(".")
    return texto or "0"


def _carregar_produtos_pedido_exportacao(
    pedido: PedidoCompra,
    db: Session,
    tenant_id: int,
) -> dict[int, Produto]:
    produto_ids = sorted({item.produto_id for item in pedido.itens if item.produto_id})
    if not produto_ids:
        return {}

    produtos = (
        db.query(Produto)
        .filter(
            Produto.id.in_(produto_ids),
            Produto.tenant_id == tenant_id,
        )
        .all()
    )
    return {produto.id: produto for produto in produtos}


def _montar_linhas_exportacao_pedido(
    pedido: PedidoCompra,
    db: Session,
    tenant_id: int,
) -> List[dict]:
    produtos_por_id = _carregar_produtos_pedido_exportacao(pedido, db, tenant_id)
    linhas = []

    for item in pedido.itens:
        produto = produtos_por_id.get(item.produto_id)
        linhas.append(
            {
                "codigo": produto.codigo if produto else "",
                "produto": produto.nome if produto else f"Produto {item.produto_id}",
                "quantidade": item.quantidade_pedida or 0,
                "preco_unitario": item.preco_unitario or 0,
                "desconto": item.desconto_item or 0,
                "total": item.valor_total or 0,
            }
        )

    return linhas


def _montar_resposta_pedido_detalhada(pedido: PedidoCompra) -> dict:
    resposta = {
        "id": pedido.id,
        "numero_pedido": pedido.numero_pedido,
        "fornecedor_id": pedido.fornecedor_id,
        "status": pedido.status,
        "valor_total": pedido.valor_total,
        "valor_frete": pedido.valor_frete,
        "valor_desconto": pedido.valor_desconto,
        "valor_final": pedido.valor_final,
        "data_pedido": pedido.data_pedido,
        "data_prevista_entrega": pedido.data_prevista_entrega,
        "data_recebimento": pedido.data_recebimento,
        "data_envio": pedido.data_envio,
        "data_confirmacao": pedido.data_confirmacao,
        "observacoes": pedido.observacoes,
        "sugestao_ia": pedido.sugestao_ia,
        "confianca_ia": pedido.confianca_ia,
        "updated_at": pedido.updated_at,
        "itens": [],
    }

    for item in pedido.itens:
        resposta["itens"].append(
            {
                "id": item.id,
                "produto_id": item.produto_id,
                "produto_nome": item.produto.nome if item.produto else None,
                "produto_codigo": item.produto.codigo if item.produto else None,
                "quantidade_pedida": item.quantidade_pedida,
                "quantidade_recebida": item.quantidade_recebida,
                "preco_unitario": item.preco_unitario,
                "desconto_item": item.desconto_item,
                "valor_total": item.valor_total,
                "status": item.status,
                "sugestao_ia": item.sugestao_ia,
                "motivo_ia": item.motivo_ia,
            }
        )

    return resposta


def _gerar_excel_pedido_bytes(
    pedido: PedidoCompra,
    fornecedor_nome: str,
    db: Session,
    tenant_id: int,
    colunas_exportacao: Optional[List[str]] = None,
) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Biblioteca openpyxl nao instalada. Execute: pip install openpyxl",
        )

    colunas_exportacao = _normalizar_colunas_exportacao_pedido(colunas_exportacao)
    linhas = _montar_linhas_exportacao_pedido(pedido, db, tenant_id)
    exibir_totais_financeiros = _exportacao_pedido_tem_colunas_financeiras(
        colunas_exportacao
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedido de Compra"

    ws["A1"] = "PEDIDO DE COMPRA"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells(f"A1:{get_column_letter(len(colunas_exportacao))}1")

    row = 3
    ws[f"A{row}"] = "Numero do Pedido:"
    ws[f"B{row}"] = pedido.numero_pedido
    ws[f"B{row}"].font = Font(bold=True)

    row += 1
    ws[f"A{row}"] = "Fornecedor:"
    ws[f"B{row}"] = fornecedor_nome

    row += 1
    ws[f"A{row}"] = "Data do Pedido:"
    ws[f"B{row}"] = pedido.data_pedido.strftime("%d/%m/%Y")

    row += 1
    ws[f"A{row}"] = "Status:"
    ws[f"B{row}"] = pedido.status.replace("_", " ").upper()

    row += 2
    headers = [
        PEDIDO_EXPORT_COLUNAS_META[coluna]["titulo"] for coluna in colunas_exportacao
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )

    row += 1
    for linha in linhas:
        for col, coluna in enumerate(colunas_exportacao, start=1):
            cell = ws.cell(row=row, column=col)
            valor = linha[coluna]
            if coluna in PEDIDO_EXPORT_COLUNAS_FINANCEIRAS:
                cell.value = float(valor or 0)
                cell.number_format = "R$ #,##0.00"
            elif coluna == "quantidade":
                cell.value = float(valor or 0)
                cell.number_format = "#,##0.00"
            else:
                cell.value = valor
        row += 1

    if exibir_totais_financeiros:
        coluna_valor = max(2, len(colunas_exportacao))
        coluna_label = max(1, coluna_valor - 1)

        row += 1
        ws.cell(row=row, column=coluna_label).value = "Frete:"
        ws.cell(row=row, column=coluna_valor).value = float(pedido.valor_frete or 0)
        ws.cell(row=row, column=coluna_valor).number_format = "R$ #,##0.00"

        row += 1
        ws.cell(row=row, column=coluna_label).value = "Desconto:"
        ws.cell(row=row, column=coluna_valor).value = float(pedido.valor_desconto or 0)
        ws.cell(row=row, column=coluna_valor).number_format = "R$ #,##0.00"

        row += 1
        ws.cell(row=row, column=coluna_label).value = "TOTAL:"
        ws.cell(row=row, column=coluna_label).font = Font(bold=True, size=12)
        ws.cell(row=row, column=coluna_valor).value = float(pedido.valor_final or 0)
        ws.cell(row=row, column=coluna_valor).font = Font(bold=True, size=12)
        ws.cell(row=row, column=coluna_valor).number_format = "R$ #,##0.00"

    for indice, coluna in enumerate(colunas_exportacao, start=1):
        letra_coluna = get_column_letter(indice)
        ws.column_dimensions[letra_coluna].width = PEDIDO_EXPORT_COLUNAS_META[coluna][
            "excel_width"
        ]
    if exibir_totais_financeiros and len(colunas_exportacao) == 1:
        ws.column_dimensions[get_column_letter(2)].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _gerar_pdf_pedido_bytes(
    pedido: PedidoCompra,
    fornecedor_nome: str,
    db: Session,
    tenant_id: int,
    colunas_exportacao: Optional[List[str]] = None,
) -> bytes:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Biblioteca reportlab nao instalada. Execute: pip install reportlab",
        )

    colunas_exportacao = _normalizar_colunas_exportacao_pedido(colunas_exportacao)
    linhas = _montar_linhas_exportacao_pedido(pedido, db, tenant_id)
    exibir_totais_financeiros = _exportacao_pedido_tem_colunas_financeiras(
        colunas_exportacao
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4, topMargin=15 * mm, bottomMargin=15 * mm
    )
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=colors.HexColor("#1a56db"),
        spaceAfter=12,
        alignment=TA_CENTER,
    )
    elements.append(Paragraph("PEDIDO DE COMPRA", title_style))
    elements.append(Spacer(1, 10 * mm))

    info_data = [
        ["Numero do Pedido:", pedido.numero_pedido],
        ["Fornecedor:", fornecedor_nome],
        ["Data do Pedido:", pedido.data_pedido.strftime("%d/%m/%Y %H:%M")],
        ["Status:", pedido.status.replace("_", " ").upper()],
    ]

    info_table = Table(info_data, colWidths=[40 * mm, 120 * mm])
    info_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    elements.append(info_table)
    elements.append(Spacer(1, 10 * mm))

    table_data = [
        [PEDIDO_EXPORT_COLUNAS_META[coluna]["titulo"] for coluna in colunas_exportacao]
    ]

    produto_style = ParagraphStyle(
        "ProdutoStyle",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
    )

    for linha in linhas:
        linha_formatada = []
        for coluna in colunas_exportacao:
            valor = linha[coluna]
            if coluna == "produto":
                linha_formatada.append(Paragraph(str(valor or ""), produto_style))
            elif coluna == "quantidade":
                linha_formatada.append(_formatar_quantidade_documento(valor))
            elif coluna in PEDIDO_EXPORT_COLUNAS_FINANCEIRAS:
                linha_formatada.append(f"R$ {float(valor or 0):.2f}")
            else:
                linha_formatada.append(str(valor or ""))
        table_data.append(linha_formatada)

    col_widths = [
        PEDIDO_EXPORT_COLUNAS_META[coluna]["pdf_width_mm"] * mm
        for coluna in colunas_exportacao
    ]
    items_table = Table(table_data, colWidths=col_widths)
    table_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
    ]
    for indice, coluna in enumerate(colunas_exportacao):
        if coluna == "produto":
            table_style.append(("ALIGN", (indice, 1), (indice, -1), "LEFT"))
        elif coluna in PEDIDO_EXPORT_COLUNAS_FINANCEIRAS:
            table_style.append(("ALIGN", (indice, 1), (indice, -1), "RIGHT"))
    items_table.setStyle(TableStyle(table_style))
    elements.append(items_table)
    if exibir_totais_financeiros:
        elements.append(Spacer(1, 5 * mm))
        totals_data = [
            ["Frete:", f"R$ {pedido.valor_frete:.2f}"],
            ["Desconto:", f"R$ {pedido.valor_desconto:.2f}"],
            ["TOTAL:", f"R$ {pedido.valor_final:.2f}"],
        ]

        totals_table = Table(totals_data, colWidths=[35 * mm, 30 * mm], hAlign="RIGHT")
        totals_table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, -1), (-1, -1), 12),
                    ("LINEABOVE", (0, -1), (-1, -1), 2, colors.black),
                ]
            )
        )
        elements.append(totals_table)

    if pedido.observacoes:
        elements.append(Spacer(1, 10 * mm))
        elements.append(Paragraph("<b>Observacoes:</b>", styles["Normal"]))
        elements.append(Paragraph(pedido.observacoes, styles["Normal"]))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def _montar_email_pedido(
    pedido: PedidoCompra,
    fornecedor_nome: str,
    colunas_exportacao: Optional[List[str]] = None,
) -> tuple[str, str, str]:
    assunto = f"Pedido de compra {pedido.numero_pedido} - {fornecedor_nome}"
    data_pedido = pedido.data_pedido.strftime("%d/%m/%Y %H:%M")
    observacoes = pedido.observacoes or "Sem observacoes adicionais."
    total_itens = len(pedido.itens or [])
    colunas_exportacao = _normalizar_colunas_exportacao_pedido(colunas_exportacao)
    exibir_totais_financeiros = _exportacao_pedido_tem_colunas_financeiras(
        colunas_exportacao
    )
    resumo_documento_html = (
        f"<li><strong>Valor final:</strong> R$ {pedido.valor_final:.2f}</li>"
        if exibir_totais_financeiros
        else "<li><strong>Documento anexado:</strong> sem valores de custo e total</li>"
    )
    resumo_documento_texto = (
        f"Valor final: R$ {pedido.valor_final:.2f}\n"
        if exibir_totais_financeiros
        else "Documento anexado sem valores de custo e total\n"
    )

    html_body = f"""
    <html>
      <body style="font-family:Arial,sans-serif;color:#1f2937;max-width:640px;margin:0 auto;">
        <div style="background:#4f46e5;padding:20px;border-radius:12px 12px 0 0;color:#ffffff;">
          <h1 style="margin:0;font-size:22px;">Pedido de Compra {pedido.numero_pedido}</h1>
          <p style="margin:8px 0 0;opacity:0.9;">Fornecedor: {fornecedor_nome}</p>
        </div>
        <div style="border:1px solid #e5e7eb;border-top:none;border-radius:0 0 12px 12px;padding:24px;">
          <p>Ola,</p>
          <p>Segue em anexo o pedido de compra <strong>{pedido.numero_pedido}</strong> gerado em {data_pedido}.</p>
          <ul>
            <li><strong>Itens:</strong> {total_itens}</li>
            {resumo_documento_html}
            <li><strong>Status:</strong> {pedido.status.replace("_", " ").upper()}</li>
          </ul>
          <p><strong>Observacoes:</strong><br />{observacoes}</p>
          <p>Se precisar de qualquer ajuste, responda este e-mail.</p>
        </div>
      </body>
    </html>
    """

    text_body = (
        f"Pedido de compra {pedido.numero_pedido}\n"
        f"Fornecedor: {fornecedor_nome}\n"
        f"Data: {data_pedido}\n"
        f"Itens: {total_itens}\n"
        f"{resumo_documento_texto}\n"
        f"Observacoes: {observacoes}"
    )

    return assunto, html_body, text_body
