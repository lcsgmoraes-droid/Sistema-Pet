"""
ROTAS DE PEDIDOS DE COMPRA - Sistema Pet Shop Pro
Gestão completa de pedidos de compra com estrutura para IA futura

Funcionalidades:
- CRUD completo de pedidos
- Controle de status (rascunho → enviado → confirmado → recebido)
- Recebimento parcial ou total
- Entrada automática no estoque
- Exportação PDF/Excel (TODO)
- Integração WhatsApp/Email (TODO)

TODO - Integração com IA (Fase Futura):
1. Vinculação automática pedido ↔ NF-e XML/PDF
2. Conferência inteligente: pedido vs nota fiscal
   - Detectar divergências de preço, quantidade, produtos
   - Alertar sobre aumentos de preço significativos
3. Ações automáticas sugeridas pela IA:
   - Gerar novo pedido com itens faltantes
   - Criar mensagem questionando fornecedor sobre diferenças
   - Sugerir fornecedores alternativos em caso de problemas
4. Análise de histórico de compras para:
   - Prever necessidade de reposição
   - Identificar padrões de preço
   - Recomendar melhores fornecedores por produto
"""
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import and_, or_, func, desc
from typing import List, Optional
from datetime import datetime, timedelta
from pydantic import BaseModel, Field
import json
import io
from io import BytesIO

from .db import get_session
from .auth import get_current_user
from .auth.dependencies import get_current_user_and_tenant
from .models import User, Cliente
from .produtos_models import (
    Produto, ProdutoLote, EstoqueMovimentacao, 
    PedidoCompra, PedidoCompraItem, ProdutoFornecedor, Marca
)
from .vendas_models import VendaItem, Venda
from .services.email_service import is_email_configured, send_email

import logging
logger = logging.getLogger(__name__)

router = APIRouter(prefix="/pedidos-compra", tags=["Pedidos de Compra"])

# ============================================================================
# SCHEMAS
# ============================================================================

class PedidoCompraItemRequest(BaseModel):
    """Schema para item do pedido"""
    produto_id: int
    quantidade_pedida: float = Field(gt=0)
    preco_unitario: float = Field(ge=0)
    desconto_item: float = Field(default=0, ge=0)
    
    # IA (futuro)
    sugestao_ia: bool = False
    motivo_ia: Optional[str] = None


class PedidoCompraRequest(BaseModel):
    """Schema para criar/editar pedido"""
    fornecedor_id: int
    data_prevista_entrega: Optional[datetime] = None
    valor_frete: float = Field(default=0, ge=0)
    valor_desconto: float = Field(default=0, ge=0)
    observacoes: Optional[str] = None
    itens: List[PedidoCompraItemRequest]
    
    # IA (futuro)
    sugestao_ia: bool = False
    confianca_ia: Optional[float] = None
    dados_ia: Optional[str] = None


class RecebimentoItemRequest(BaseModel):
    """Schema para receber item"""
    item_id: int
    quantidade_recebida: float = Field(gt=0)


class RecebimentoPedidoRequest(BaseModel):
    """Schema para recebimento de pedido"""
    itens: List[RecebimentoItemRequest]
    data_recebimento: Optional[datetime] = None


class PedidoCompraResponse(BaseModel):
    """Schema de resposta do pedido"""
    id: int
    numero_pedido: str
    fornecedor_id: int
    status: str
    valor_total: float
    valor_frete: float
    valor_desconto: float
    valor_final: float
    data_pedido: datetime
    data_prevista_entrega: Optional[datetime]
    observacoes: Optional[str]
    itens_count: int
    
    model_config = {"from_attributes": True}


class PedidoCompraEnvioFormatos(BaseModel):
    pdf: bool = True
    excel: bool = False


class PedidoCompraEnviarRequest(BaseModel):
    email: Optional[str] = None
    whatsapp: Optional[str] = None
    formatos: PedidoCompraEnvioFormatos = Field(default_factory=PedidoCompraEnvioFormatos)
    envio_manual: bool = False


def _buscar_fornecedor_pedido(db: Session, tenant_id: int, pedido: PedidoCompra) -> Optional[Cliente]:
    return db.query(Cliente).filter(
        Cliente.id == pedido.fornecedor_id,
        Cliente.tenant_id == tenant_id
    ).first()


def _gerar_excel_pedido_bytes(
    pedido: PedidoCompra,
    fornecedor_nome: str,
    db: Session,
    tenant_id: int,
) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Biblioteca openpyxl nao instalada. Execute: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedido de Compra"

    ws["A1"] = "PEDIDO DE COMPRA"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells("A1:F1")

    row = 3
    ws[f"A{row}"] = "Numero do Pedido:"
    ws[f"B{row}"] = pedido.numero_pedido
    ws[f"B{row}"].font = Font(bold=True)

    row += 1
    ws[f"A{row}"] = "Fornecedor:"
    ws[f"B{row}"] = fornecedor_nome

    row += 1
    ws[f"A{row}"] = "Data do Pedido:"
    ws[f"B{row}"] = pedido.data_pedido.strftime("%d/%m/%Y")

    row += 1
    ws[f"A{row}"] = "Status:"
    ws[f"B{row}"] = pedido.status.replace("_", " ").upper()

    row += 2
    headers = ["Codigo", "Produto", "Quantidade", "Preco Unit.", "Desconto", "Total"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    row += 1
    for item in pedido.itens:
        produto = db.query(Produto).filter(
            Produto.id == item.produto_id,
            Produto.tenant_id == tenant_id
        ).first()
        ws.cell(row=row, column=1).value = produto.codigo if produto else ""
        ws.cell(row=row, column=2).value = produto.nome if produto else f"Produto {item.produto_id}"
        ws.cell(row=row, column=3).value = item.quantidade_pedida
        ws.cell(row=row, column=4).value = item.preco_unitario
        ws.cell(row=row, column=5).value = item.desconto_item
        ws.cell(row=row, column=6).value = item.valor_total
        row += 1

    row += 1
    ws[f"E{row}"] = "Frete:"
    ws[f"F{row}"] = pedido.valor_frete
    row += 1
    ws[f"E{row}"] = "Desconto:"
    ws[f"F{row}"] = pedido.valor_desconto
    row += 1
    ws[f"E{row}"] = "TOTAL:"
    ws[f"E{row}"].font = Font(bold=True, size=12)
    ws[f"F{row}"] = pedido.valor_final
    ws[f"F{row}"].font = Font(bold=True, size=12)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _gerar_pdf_pedido_bytes(
    pedido: PedidoCompra,
    fornecedor_nome: str,
    db: Session,
    tenant_id: int,
) -> bytes:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Biblioteca reportlab nao instalada. Execute: pip install reportlab"
        )

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=15 * mm, bottomMargin=15 * mm)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=colors.HexColor("#1a56db"),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("PEDIDO DE COMPRA", title_style))
    elements.append(Spacer(1, 10 * mm))

    info_data = [
        ["Numero do Pedido:", pedido.numero_pedido],
        ["Fornecedor:", fornecedor_nome],
        ["Data do Pedido:", pedido.data_pedido.strftime("%d/%m/%Y %H:%M")],
        ["Status:", pedido.status.replace("_", " ").upper()],
    ]

    info_table = Table(info_data, colWidths=[40 * mm, 120 * mm])
    info_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 10 * mm))

    table_data = [["Codigo", "Produto", "Qtd", "Preco Unit.", "Desc.", "Total"]]

    produto_style = ParagraphStyle(
        "ProdutoStyle",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
    )

    for item in pedido.itens:
        produto = db.query(Produto).filter(
            Produto.id == item.produto_id,
            Produto.tenant_id == tenant_id
        ).first()
        nome_produto = produto.nome if produto else f"Produto {item.produto_id}"
        table_data.append([
            produto.codigo if produto else "",
            Paragraph(nome_produto, produto_style),
            f"{item.quantidade_pedida:.0f}",
            f"R$ {item.preco_unitario:.2f}",
            f"R$ {item.desconto_item:.2f}",
            f"R$ {item.valor_total:.2f}",
        ])

    items_table = Table(table_data, colWidths=[20 * mm, 70 * mm, 15 * mm, 23 * mm, 18 * mm, 23 * mm])
    items_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (1, 1), (1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 5 * mm))

    totals_data = [
        ["", "", "", "", "Frete:", f"R$ {pedido.valor_frete:.2f}"],
        ["", "", "", "", "Desconto:", f"R$ {pedido.valor_desconto:.2f}"],
        ["", "", "", "", "TOTAL:", f"R$ {pedido.valor_final:.2f}"],
    ]

    totals_table = Table(totals_data, colWidths=[20 * mm, 70 * mm, 20 * mm, 25 * mm, 20 * mm, 25 * mm])
    totals_table.setStyle(TableStyle([
        ("ALIGN", (4, 0), (-1, -1), "RIGHT"),
        ("FONTNAME", (4, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (4, -1), (-1, -1), 12),
        ("LINEABOVE", (4, -1), (-1, -1), 2, colors.black),
    ]))
    elements.append(totals_table)

    if pedido.observacoes:
        elements.append(Spacer(1, 10 * mm))
        elements.append(Paragraph("<b>Observacoes:</b>", styles["Normal"]))
        elements.append(Paragraph(pedido.observacoes, styles["Normal"]))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def _montar_email_pedido(pedido: PedidoCompra, fornecedor_nome: str) -> tuple[str, str, str]:
    assunto = f"Pedido de compra {pedido.numero_pedido} - {fornecedor_nome}"
    data_pedido = pedido.data_pedido.strftime("%d/%m/%Y %H:%M")
    observacoes = pedido.observacoes or "Sem observacoes adicionais."
    total_itens = len(pedido.itens or [])

    html_body = f"""
    <html>
      <body style="font-family:Arial,sans-serif;color:#1f2937;max-width:640px;margin:0 auto;">
        <div style="background:#4f46e5;padding:20px;border-radius:12px 12px 0 0;color:#ffffff;">
          <h1 style="margin:0;font-size:22px;">Pedido de Compra {pedido.numero_pedido}</h1>
          <p style="margin:8px 0 0;opacity:0.9;">Fornecedor: {fornecedor_nome}</p>
        </div>
        <div style="border:1px solid #e5e7eb;border-top:none;border-radius:0 0 12px 12px;padding:24px;">
          <p>Ola,</p>
          <p>Segue em anexo o pedido de compra <strong>{pedido.numero_pedido}</strong> gerado em {data_pedido}.</p>
          <ul>
            <li><strong>Itens:</strong> {total_itens}</li>
            <li><strong>Valor final:</strong> R$ {pedido.valor_final:.2f}</li>
            <li><strong>Status:</strong> {pedido.status.replace("_", " ").upper()}</li>
          </ul>
          <p><strong>Observacoes:</strong><br />{observacoes}</p>
          <p>Se precisar de qualquer ajuste, responda este e-mail.</p>
        </div>
      </body>
    </html>
    """

    text_body = (
        f"Pedido de compra {pedido.numero_pedido}\n"
        f"Fornecedor: {fornecedor_nome}\n"
        f"Data: {data_pedido}\n"
        f"Itens: {total_itens}\n"
        f"Valor final: R$ {pedido.valor_final:.2f}\n\n"
        f"Observacoes: {observacoes}"
    )

    return assunto, html_body, text_body


# ============================================================================
# LISTAR PEDIDOS
# ============================================================================

@router.get("/", response_model=List[PedidoCompraResponse])
def listar_pedidos(
    status: Optional[str] = Query(None, description="Filtrar por status"),
    fornecedor_id: Optional[int] = Query(None, description="Filtrar por fornecedor"),
    data_inicio: Optional[str] = Query(None, description="Data inicial (YYYY-MM-DD)"),
    data_fim: Optional[str] = Query(None, description="Data final (YYYY-MM-DD)"),
    limit: int = Query(50, le=100),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_session),
    current_user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Lista pedidos de compra com filtros"""
    current_user, tenant_id = current_user_and_tenant
    logger.info(f"📋 Listando pedidos de compra - Usuário: {current_user.nome}")
    
    query = db.query(PedidoCompra).options(joinedload(PedidoCompra.itens)).filter(
        PedidoCompra.tenant_id == tenant_id
    )
    
    # Filtros
    if status:
        query = query.filter(PedidoCompra.status == status)
    if fornecedor_id:
        query = query.filter(PedidoCompra.fornecedor_id == fornecedor_id)
    if data_inicio:
        data = datetime.strptime(data_inicio, "%Y-%m-%d")
        query = query.filter(PedidoCompra.data_pedido >= data)
    if data_fim:
        data = datetime.strptime(data_fim, "%Y-%m-%d")
        query = query.filter(PedidoCompra.data_pedido <= data)
    
    # Ordenar por data decrescente
    query = query.order_by(desc(PedidoCompra.data_pedido))
    
    # Paginação
    total = query.count()
    pedidos = query.offset(offset).limit(limit).all()
    
    # Formatar resposta
    resultado = []
    for pedido in pedidos:
        resultado.append(PedidoCompraResponse(
            id=pedido.id,
            numero_pedido=pedido.numero_pedido,
            fornecedor_id=pedido.fornecedor_id,
            status=pedido.status,
            valor_total=pedido.valor_total,
            valor_frete=pedido.valor_frete,
            valor_desconto=pedido.valor_desconto,
            valor_final=pedido.valor_final,
            data_pedido=pedido.data_pedido,
            data_prevista_entrega=pedido.data_prevista_entrega,
            observacoes=pedido.observacoes,
            itens_count=len(pedido.itens)
        ))
    
    logger.info(f"✅ {len(resultado)} pedidos encontrados (total: {total})")
    return resultado


# ============================================================================
# STATUS DE ENVIO
# ============================================================================

@router.get("/envio/status")
def status_envio_pedidos(
    current_user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Informa se o servidor está apto a enviar pedidos por e-mail."""
    current_user, tenant_id = current_user_and_tenant
    return {
        "email_configurado": is_email_configured()
    }


# ============================================================================
# BUSCAR PEDIDO POR ID
# ============================================================================

@router.get("/{pedido_id}")
def buscar_pedido(
    pedido_id: int,
    db: Session = Depends(get_session),
    current_user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Busca pedido completo com itens"""
    current_user, tenant_id = current_user_and_tenant
    logger.info(f"🔍 Buscando pedido {pedido_id}")
    
    pedido = db.query(PedidoCompra).options(
        joinedload(PedidoCompra.itens).joinedload(PedidoCompraItem.produto)
    ).filter(
        PedidoCompra.id == pedido_id,
        PedidoCompra.tenant_id == tenant_id
    ).first()
    
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")
    
    # Formatar resposta detalhada
    resposta = {
        "id": pedido.id,
        "numero_pedido": pedido.numero_pedido,
        "fornecedor_id": pedido.fornecedor_id,
        "status": pedido.status,
        "valor_total": pedido.valor_total,
        "valor_frete": pedido.valor_frete,
        "valor_desconto": pedido.valor_desconto,
        "valor_final": pedido.valor_final,
        "data_pedido": pedido.data_pedido,
        "data_prevista_entrega": pedido.data_prevista_entrega,
        "data_recebimento": pedido.data_recebimento,
        "observacoes": pedido.observacoes,
        "sugestao_ia": pedido.sugestao_ia,
        "confianca_ia": pedido.confianca_ia,
        "itens": []
    }
    
    for item in pedido.itens:
        resposta["itens"].append({
            "id": item.id,
            "produto_id": item.produto_id,
            "produto_nome": item.produto.nome if item.produto else None,
            "produto_codigo": item.produto.codigo if item.produto else None,
            "quantidade_pedida": item.quantidade_pedida,
            "quantidade_recebida": item.quantidade_recebida,
            "preco_unitario": item.preco_unitario,
            "desconto_item": item.desconto_item,
            "valor_total": item.valor_total,
            "status": item.status,
            "sugestao_ia": item.sugestao_ia,
            "motivo_ia": item.motivo_ia
        })
    
    return resposta


# ============================================================================
# CRIAR PEDIDO
# ============================================================================

@router.post("/")
def criar_pedido(
    request: PedidoCompraRequest,
    db: Session = Depends(get_session),
    current_user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Cria novo pedido de compra"""
    current_user, tenant_id = current_user_and_tenant
    logger.info(f"📝 Criando pedido de compra - Fornecedor: {request.fornecedor_id}")
    
    # Validar fornecedor
    fornecedor = db.query(Cliente).filter(
        Cliente.id == request.fornecedor_id,
        Cliente.tenant_id == tenant_id
    ).first()
    if not fornecedor:
        raise HTTPException(status_code=404, detail="Fornecedor não encontrado")
    
    # Validar produtos
    if not request.itens or len(request.itens) == 0:
        raise HTTPException(status_code=400, detail="Pedido deve ter pelo menos 1 item")
    
    for item_req in request.itens:
        produto = db.query(Produto).filter(
            Produto.id == item_req.produto_id,
            Produto.tenant_id == tenant_id
        ).first()
        if not produto:
            raise HTTPException(
                status_code=404, 
                detail=f"Produto {item_req.produto_id} não encontrado"
            )
    
    # Gerar número do pedido
    ultimo_pedido = db.query(PedidoCompra).order_by(desc(PedidoCompra.id)).first()
    numero = 1 if not ultimo_pedido else ultimo_pedido.id + 1
    numero_pedido = f"PC{datetime.now().year}{numero:05d}"
    
    # Calcular totais
    valor_total = sum(
        (item.preco_unitario - item.desconto_item) * item.quantidade_pedida 
        for item in request.itens
    )
    valor_final = valor_total + request.valor_frete - request.valor_desconto
    
    # Criar pedido
    pedido = PedidoCompra(
        numero_pedido=numero_pedido,
        fornecedor_id=request.fornecedor_id,
        status="rascunho",
        valor_total=valor_total,
        valor_frete=request.valor_frete,
        valor_desconto=request.valor_desconto,
        valor_final=valor_final,
        data_pedido=datetime.utcnow(),
        data_prevista_entrega=request.data_prevista_entrega,
        observacoes=request.observacoes,
        sugestao_ia=request.sugestao_ia,
        confianca_ia=request.confianca_ia,
        dados_ia=request.dados_ia,
        user_id=current_user.id,
        tenant_id=tenant_id
    )
    
    db.add(pedido)
    db.flush()  # Para obter o ID
    
    # Criar itens
    for item_req in request.itens:
        valor_item = (item_req.preco_unitario - item_req.desconto_item) * item_req.quantidade_pedida
        
        item = PedidoCompraItem(
            pedido_compra_id=pedido.id,
            produto_id=item_req.produto_id,
            quantidade_pedida=item_req.quantidade_pedida,
            quantidade_recebida=0,
            preco_unitario=item_req.preco_unitario,
            desconto_item=item_req.desconto_item,
            valor_total=valor_item,
            status="pendente",
            sugestao_ia=item_req.sugestao_ia,
            motivo_ia=item_req.motivo_ia,
            tenant_id=tenant_id
        )
        db.add(item)
    
    db.commit()
    db.refresh(pedido)
    
    logger.info(f"✅ Pedido {pedido.numero_pedido} criado com sucesso")
    
    return {
        "message": "Pedido criado com sucesso",
        "pedido_id": pedido.id,
        "numero_pedido": pedido.numero_pedido,
        "valor_final": pedido.valor_final,
        "itens_count": len(request.itens)
    }


# ============================================================================
# ATUALIZAR PEDIDO
# ============================================================================

@router.put("/{pedido_id}")
def atualizar_pedido(
    pedido_id: int,
    request: PedidoCompraRequest,
    db: Session = Depends(get_session),
    current_user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Atualiza pedido (apenas se status = rascunho)"""
    current_user, tenant_id = current_user_and_tenant
    logger.info(f"✏️ Atualizando pedido {pedido_id}")
    
    pedido = db.query(PedidoCompra).filter(
        PedidoCompra.id == pedido_id,
        PedidoCompra.tenant_id == tenant_id
    ).first()
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")
    
    if pedido.status != "rascunho":
        raise HTTPException(
            status_code=400, 
            detail=f"Pedido não pode ser editado no status '{pedido.status}'"
        )
    
    # Recalcular totais
    valor_total = sum(
        (item.preco_unitario - item.desconto_item) * item.quantidade_pedida 
        for item in request.itens
    )
    valor_final = valor_total + request.valor_frete - request.valor_desconto
    
    # Atualizar pedido
    pedido.fornecedor_id = request.fornecedor_id
    pedido.valor_total = valor_total
    pedido.valor_frete = request.valor_frete
    pedido.valor_desconto = request.valor_desconto
    pedido.valor_final = valor_final
    pedido.data_prevista_entrega = request.data_prevista_entrega
    pedido.observacoes = request.observacoes
    pedido.updated_at = datetime.utcnow()
    
    # Remover itens antigos
    db.query(PedidoCompraItem).filter(
        PedidoCompraItem.pedido_compra_id == pedido_id
    ).delete()
    
    # Adicionar novos itens
    for item_req in request.itens:
        valor_item = (item_req.preco_unitario - item_req.desconto_item) * item_req.quantidade_pedida
        
        item = PedidoCompraItem(
            pedido_compra_id=pedido.id,
            produto_id=item_req.produto_id,
            quantidade_pedida=item_req.quantidade_pedida,
            quantidade_recebida=0,
            preco_unitario=item_req.preco_unitario,
            desconto_item=item_req.desconto_item,
            valor_total=valor_item,
            status="pendente",
            sugestao_ia=item_req.sugestao_ia,
            motivo_ia=item_req.motivo_ia,
            tenant_id=tenant_id
        )
        db.add(item)
    
    db.commit()
    
    logger.info(f"✅ Pedido {pedido.numero_pedido} atualizado")
    
    return {
        "message": "Pedido atualizado com sucesso",
        "pedido_id": pedido.id,
        "numero_pedido": pedido.numero_pedido
    }


# ============================================================================
# ENVIAR PEDIDO
# ============================================================================

@router.post("/{pedido_id}/enviar")
def enviar_pedido(
    pedido_id: int,
    request: PedidoCompraEnviarRequest,
    db: Session = Depends(get_session),
    current_user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Marca pedido como enviado"""
    current_user, tenant_id = current_user_and_tenant
    logger.info(f"📤 Enviando pedido {pedido_id}")
    
    pedido = db.query(PedidoCompra).filter(
        PedidoCompra.id == pedido_id,
        PedidoCompra.tenant_id == tenant_id
    ).first()
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")
    
    if pedido.status != "rascunho":
        raise HTTPException(
            status_code=400, 
            detail=f"Pedido não pode ser enviado no status '{pedido.status}'"
        )
    
    fornecedor = _buscar_fornecedor_pedido(db, tenant_id, pedido)
    fornecedor_nome = fornecedor.nome if fornecedor else f"Fornecedor {pedido.fornecedor_id}"

    if request.envio_manual:
        pedido.status = "enviado"
        pedido.data_envio = datetime.utcnow()
        pedido.updated_at = datetime.utcnow()
        db.commit()

        logger.info(f"Pedido {pedido.numero_pedido} marcado como enviado manualmente")
        return {
            "message": "Pedido marcado como enviado manualmente",
            "pedido_id": pedido.id,
            "numero_pedido": pedido.numero_pedido,
            "status": pedido.status,
            "tipo_envio": "manual"
        }

    email_destino = (request.email or "").strip()
    if not email_destino:
        raise HTTPException(status_code=400, detail="Informe o e-mail do fornecedor")

    formatos = request.formatos or PedidoCompraEnvioFormatos()
    if not formatos.pdf and not formatos.excel:
        raise HTTPException(status_code=400, detail="Selecione pelo menos um formato para envio")

    if not is_email_configured():
        raise HTTPException(
            status_code=503,
            detail="O envio de e-mail nao esta configurado no servidor"
        )

    anexos = []
    if formatos.pdf:
        anexos.append({
            "filename": f"pedido_{pedido.numero_pedido}.pdf",
            "content": _gerar_pdf_pedido_bytes(pedido, fornecedor_nome, db, tenant_id),
            "mime_subtype": "pdf"
        })
    if formatos.excel:
        anexos.append({
            "filename": f"pedido_{pedido.numero_pedido}.xlsx",
            "content": _gerar_excel_pedido_bytes(pedido, fornecedor_nome, db, tenant_id),
            "mime_subtype": "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        })

    assunto, html_body, text_body = _montar_email_pedido(pedido, fornecedor_nome)
    enviado = send_email(
        to=email_destino,
        subject=assunto,
        html_body=html_body,
        text_body=text_body,
        attachments=anexos,
        simulate_if_unconfigured=False
    )

    if not enviado:
        raise HTTPException(
            status_code=502,
            detail="Nao foi possivel enviar o e-mail do pedido. Revise a configuracao SMTP."
        )

    pedido.status = "enviado"
    pedido.data_envio = datetime.utcnow()
    pedido.updated_at = datetime.utcnow()
    
    db.commit()
    logger.info(f"Pedido {pedido.numero_pedido} enviado por e-mail para {email_destino}")
    return {
        "message": "Pedido enviado por e-mail com sucesso",
        "pedido_id": pedido.id,
        "numero_pedido": pedido.numero_pedido,
        "status": pedido.status,
        "tipo_envio": "email",
        "email": email_destino
    }
    
    logger.info(f"✅ Pedido {pedido.numero_pedido} enviado")
    
    return {
        "message": "Pedido enviado com sucesso",
        "pedido_id": pedido.id,
        "numero_pedido": pedido.numero_pedido,
        "status": pedido.status
    }


# ============================================================================
# CONFIRMAR PEDIDO
# ============================================================================

@router.post("/{pedido_id}/confirmar")
def confirmar_pedido(
    pedido_id: int,
    db: Session = Depends(get_session),
    current_user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Confirma pedido (fornecedor confirmou recebimento)"""
    current_user, tenant_id = current_user_and_tenant
    logger.info(f"✅ Confirmando pedido {pedido_id}")
    
    pedido = db.query(PedidoCompra).filter(
        PedidoCompra.id == pedido_id,
        PedidoCompra.tenant_id == tenant_id
    ).first()
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")
    
    if pedido.status not in ["rascunho", "enviado"]:
        raise HTTPException(
            status_code=400, 
            detail=f"Pedido não pode ser confirmado no status '{pedido.status}'"
        )
    
    pedido.status = "confirmado"
    pedido.data_confirmacao = datetime.utcnow()
    pedido.updated_at = datetime.utcnow()
    
    db.commit()
    
    logger.info(f"✅ Pedido {pedido.numero_pedido} confirmado")
    
    return {
        "message": "Pedido confirmado com sucesso",
        "pedido_id": pedido.id,
        "numero_pedido": pedido.numero_pedido,
        "status": pedido.status
    }


# ============================================================================
# CANCELAR PEDIDO
# ============================================================================

@router.post("/{pedido_id}/cancelar")
def cancelar_pedido(
    pedido_id: int,
    motivo: str = Query(..., min_length=10, description="Motivo do cancelamento"),
    db: Session = Depends(get_session),
    current_user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Cancela pedido"""
    current_user, tenant_id = current_user_and_tenant
    logger.info(f"❌ Cancelando pedido {pedido_id}")
    
    pedido = db.query(PedidoCompra).filter(
        PedidoCompra.id == pedido_id,
        PedidoCompra.tenant_id == tenant_id
    ).first()
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")
    
    if pedido.status in ["recebido_total", "cancelado"]:
        raise HTTPException(
            status_code=400, 
            detail=f"Pedido não pode ser cancelado no status '{pedido.status}'"
        )
    
    pedido.status = "cancelado"
    pedido.observacoes = f"{pedido.observacoes or ''}\n\nCANCELADO: {motivo}"
    pedido.updated_at = datetime.utcnow()
    
    # Cancelar todos os itens
    for item in pedido.itens:
        item.status = "cancelado"
    
    db.commit()
    
    logger.info(f"❌ Pedido {pedido.numero_pedido} cancelado")
    
    return {
        "message": "Pedido cancelado com sucesso",
        "pedido_id": pedido.id,
        "numero_pedido": pedido.numero_pedido,
        "status": pedido.status
    }


# ============================================================================
# RECEBER PEDIDO (com entrada automática no estoque)
# ============================================================================

@router.post("/{pedido_id}/receber")
def receber_pedido(
    pedido_id: int,
    request: RecebimentoPedidoRequest,
    db: Session = Depends(get_session),
    current_user_and_tenant = Depends(get_current_user_and_tenant)
):
    """
    Recebe pedido (total ou parcial) e dá entrada automática no estoque
    """
    current_user, tenant_id = current_user_and_tenant
    logger.info(f"📦 Recebendo pedido {pedido_id}")
    
    pedido = db.query(PedidoCompra).options(
        joinedload(PedidoCompra.itens)
    ).filter(
        PedidoCompra.id == pedido_id,
        PedidoCompra.tenant_id == tenant_id
    ).first()
    
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")
    
    if pedido.status not in ["confirmado", "recebido_parcial"]:
        raise HTTPException(
            status_code=400, 
            detail=f"Pedido não pode ser recebido no status '{pedido.status}'"
        )
    
    data_recebimento = request.data_recebimento or datetime.utcnow()
    itens_recebidos = []
    para_sync_bling = []  # (produto_id, estoque_atual)
    
    # Processar cada item
    for receb_item in request.itens:
        # Buscar item
        item = db.query(PedidoCompraItem).filter(
            PedidoCompraItem.id == receb_item.item_id,
            PedidoCompraItem.pedido_compra_id == pedido_id
        ).first()
        
        if not item:
            raise HTTPException(
                status_code=404, 
                detail=f"Item {receb_item.item_id} não encontrado no pedido"
            )
        
        # Validar quantidade
        quantidade_pendente = item.quantidade_pedida - item.quantidade_recebida
        if receb_item.quantidade_recebida > quantidade_pendente:
            raise HTTPException(
                status_code=400,
                detail=f"Item {item.id}: quantidade recebida ({receb_item.quantidade_recebida}) "
                       f"maior que pendente ({quantidade_pendente})"
            )
        
        # Atualizar quantidade recebida
        item.quantidade_recebida += receb_item.quantidade_recebida
        
        # Atualizar status do item
        if item.quantidade_recebida >= item.quantidade_pedida:
            item.status = "recebido_total"
        else:
            item.status = "recebido_parcial"
        
        # DAR ENTRADA NO ESTOQUE
        produto = db.query(Produto).filter(
            Produto.id == item.produto_id,
            Produto.tenant_id == tenant_id
        ).first()
        
        # Criar lote
        numero_lote = f"PC{pedido.id}-{item.id}"
        
        lote = ProdutoLote(
            produto_id=produto.id,
            numero_lote=numero_lote,
            quantidade_inicial=receb_item.quantidade_recebida,
            quantidade_atual=receb_item.quantidade_recebida,
            custo_unitario=item.preco_unitario - item.desconto_item,
            data_fabricacao=None,
            data_validade=None,
            fornecedor=f"Pedido {pedido.numero_pedido}",
            user_id=current_user.id,
            tenant_id=tenant_id
        )
        db.add(lote)
        db.flush()
        
        # Atualizar estoque do produto
        produto.estoque_atual = (produto.estoque_atual or 0) + receb_item.quantidade_recebida
        
        # Recalcular custo médio ponderado
        if produto.custo_medio:
            estoque_anterior = (produto.estoque_atual or 0) - receb_item.quantidade_recebida
            valor_anterior = estoque_anterior * produto.custo_medio
            valor_entrada = receb_item.quantidade_recebida * lote.custo_unitario
            produto.custo_medio = (valor_anterior + valor_entrada) / produto.estoque_atual
        else:
            produto.custo_medio = lote.custo_unitario
        
        # Registrar movimentação
        movimentacao = EstoqueMovimentacao(
            produto_id=produto.id,
            lote_id=lote.id,
            tipo_movimentacao="entrada",
            quantidade=receb_item.quantidade_recebida,
            custo_unitario=lote.custo_unitario,
            motivo=f"Recebimento do pedido {pedido.numero_pedido}",
            documento=pedido.numero_pedido,
            estoque_anterior=(produto.estoque_atual or 0) - receb_item.quantidade_recebida,
            estoque_atual=produto.estoque_atual,
            user_id=current_user.id,
            tenant_id=tenant_id
        )
        db.add(movimentacao)
        
        itens_recebidos.append({
            "item_id": item.id,
            "produto_id": produto.id,
            "produto_nome": produto.nome,
            "quantidade_recebida": receb_item.quantidade_recebida,
            "lote": numero_lote,
            "status": item.status
        })
        para_sync_bling.append((produto.id, produto.estoque_atual))
        
        logger.info(
            f"  ✅ Item {item.id}: {produto.nome} - "
            f"{receb_item.quantidade_recebida} unidades recebidas"
        )
    
    # Atualizar status do pedido
    todos_recebidos = all(item.status == "recebido_total" for item in pedido.itens)
    algum_recebido = any(item.quantidade_recebida > 0 for item in pedido.itens)
    
    if todos_recebidos:
        pedido.status = "recebido_total"
        pedido.data_recebimento = data_recebimento
    elif algum_recebido:
        pedido.status = "recebido_parcial"
        if not pedido.data_recebimento:
            pedido.data_recebimento = data_recebimento
    
    pedido.updated_at = datetime.utcnow()
    
    db.commit()
    
    # SINCRONIZAR ESTOQUE COM BLING para todos os itens recebidos
    try:
        from app.bling_estoque_sync import sincronizar_bling_background
        for produto_id, estoque_novo in para_sync_bling:
            sincronizar_bling_background(produto_id, estoque_novo, "recebimento_pedido_compra")
    except Exception as e_sync:
        logger.warning(f"[BLING-SYNC] Erro ao agendar sync (pedido_compra): {e_sync}")
    
    logger.info(f"✅ Pedido {pedido.numero_pedido} recebido - Status: {pedido.status}")
    
    return {
        "message": "Recebimento processado com sucesso",
        "pedido_id": pedido.id,
        "numero_pedido": pedido.numero_pedido,
        "status": pedido.status,
        "itens_recebidos": len(itens_recebidos),
        "detalhes": itens_recebidos
    }


# ============================================================================
# EXPORTAÇÃO PDF/EXCEL
# ============================================================================

@router.get("/{pedido_id}/export/excel")
def exportar_excel(
    pedido_id: int,
    db: Session = Depends(get_session),
    current_user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Exporta pedido para Excel"""
    current_user, tenant_id = current_user_and_tenant
    pedido = db.query(PedidoCompra).filter(
        PedidoCompra.id == pedido_id,
        PedidoCompra.tenant_id == tenant_id
    ).options(joinedload(PedidoCompra.itens)).first()
    
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=500, 
            detail="Biblioteca openpyxl não instalada. Execute: pip install openpyxl"
        )
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedido de Compra"
    
    # Cabeçalho
    ws['A1'] = "PEDIDO DE COMPRA"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')
    
    # Dados do pedido
    row = 3
    ws[f'A{row}'] = "Número do Pedido:"
    ws[f'B{row}'] = pedido.numero_pedido
    ws[f'B{row}'].font = Font(bold=True)
    
    row += 1
    fornecedor = db.query(Cliente).filter(
        Cliente.id == pedido.fornecedor_id,
        Cliente.tenant_id == tenant_id
    ).first()
    ws[f'A{row}'] = "Fornecedor:"
    ws[f'B{row}'] = fornecedor.nome if fornecedor else f"ID {pedido.fornecedor_id}"
    
    row += 1
    ws[f'A{row}'] = "Data do Pedido:"
    ws[f'B{row}'] = pedido.data_pedido.strftime("%d/%m/%Y")
    
    row += 1
    ws[f'A{row}'] = "Status:"
    ws[f'B{row}'] = pedido.status.replace('_', ' ').upper()
    
    # Itens
    row += 2
    headers = ['Código', 'Produto', 'Quantidade', 'Preço Unit.', 'Desconto', 'Total']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    row += 1
    for item in pedido.itens:
        produto = db.query(Produto).filter(
            Produto.id == item.produto_id,
            Produto.tenant_id == tenant_id
        ).first()
        ws.cell(row=row, column=1).value = produto.codigo if produto else ''
        ws.cell(row=row, column=2).value = produto.nome if produto else f"Produto {item.produto_id}"
        ws.cell(row=row, column=3).value = item.quantidade_pedida
        ws.cell(row=row, column=4).value = item.preco_unitario
        ws.cell(row=row, column=5).value = item.desconto_item
        ws.cell(row=row, column=6).value = item.valor_total
        row += 1
    
    # Totais
    row += 1
    ws[f'E{row}'] = "Frete:"
    ws[f'F{row}'] = pedido.valor_frete
    row += 1
    ws[f'E{row}'] = "Desconto:"
    ws[f'F{row}'] = pedido.valor_desconto
    row += 1
    ws[f'E{row}'] = "TOTAL:"
    ws[f'E{row}'].font = Font(bold=True, size=12)
    ws[f'F{row}'] = pedido.valor_final
    ws[f'F{row}'].font = Font(bold=True, size=12)
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    # Salvar em bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"pedido_{pedido.numero_pedido}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/{pedido_id}/export/pdf")
def exportar_pdf(
    pedido_id: int,
    db: Session = Depends(get_session),
    current_user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Exporta pedido para PDF"""
    current_user, tenant_id = current_user_and_tenant
    pedido = db.query(PedidoCompra).filter(
        PedidoCompra.id == pedido_id,
        PedidoCompra.tenant_id == tenant_id
    ).options(joinedload(PedidoCompra.itens)).first()
    
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")
    
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Biblioteca reportlab não instalada. Execute: pip install reportlab"
        )
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a56db'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("PEDIDO DE COMPRA", title_style))
    elements.append(Spacer(1, 10*mm))
    
    # Info do pedido
    fornecedor = db.query(Cliente).filter(
        Cliente.id == pedido.fornecedor_id,
        Cliente.tenant_id == tenant_id
    ).first()
    fornecedor_nome = fornecedor.nome if fornecedor else f"ID {pedido.fornecedor_id}"
    
    info_data = [
        ['Número do Pedido:', pedido.numero_pedido],
        ['Fornecedor:', fornecedor_nome],
        ['Data do Pedido:', pedido.data_pedido.strftime("%d/%m/%Y %H:%M")],
        ['Status:', pedido.status.replace('_', ' ').upper()],
    ]
    
    info_table = Table(info_data, colWidths=[40*mm, 120*mm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 10*mm))
    
    # Tabela de itens
    table_data = [['Código', 'Produto', 'Qtd', 'Preço Unit.', 'Desc.', 'Total']]
    
    # Estilo para nome do produto (permite quebra de linha)
    produto_style = ParagraphStyle(
        'ProdutoStyle',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
    )
    
    for item in pedido.itens:
        produto = db.query(Produto).filter(
            Produto.id == item.produto_id,
            Produto.tenant_id == tenant_id
        ).first()
        nome_produto = produto.nome if produto else f"Produto {item.produto_id}"
        
        # Usar Paragraph para quebrar linha automaticamente
        nome_produto_paragraph = Paragraph(nome_produto, produto_style)
        
        table_data.append([
            produto.codigo if produto else '',
            nome_produto_paragraph,
            f"{item.quantidade_pedida:.0f}",
            f"R$ {item.preco_unitario:.2f}",
            f"R$ {item.desconto_item:.2f}",
            f"R$ {item.valor_total:.2f}"
        ])
    
    items_table = Table(table_data, colWidths=[20*mm, 70*mm, 15*mm, 23*mm, 18*mm, 23*mm])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 5*mm))
    
    # Totais
    totals_data = [
        ['', '', '', '', 'Frete:', f"R$ {pedido.valor_frete:.2f}"],
        ['', '', '', '', 'Desconto:', f"R$ {pedido.valor_desconto:.2f}"],
        ['', '', '', '', 'TOTAL:', f"R$ {pedido.valor_final:.2f}"],
    ]
    
    totals_table = Table(totals_data, colWidths=[20*mm, 70*mm, 20*mm, 25*mm, 20*mm, 25*mm])
    totals_table.setStyle(TableStyle([
        ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (4, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (4, -1), (-1, -1), 12),
        ('LINEABOVE', (4, -1), (-1, -1), 2, colors.black),
    ]))
    elements.append(totals_table)
    
    # Observações
    if pedido.observacoes:
        elements.append(Spacer(1, 10*mm))
        elements.append(Paragraph("<b>Observações:</b>", styles['Normal']))
        elements.append(Paragraph(pedido.observacoes, styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"pedido_{pedido.numero_pedido}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/{pedido_id}/reverter")
def reverter_status(
    pedido_id: int,
    db: Session = Depends(get_session),
    current_user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Reverte status do pedido (para corrigir cliques acidentais)"""
    current_user, tenant_id = current_user_and_tenant
    pedido = db.query(PedidoCompra).filter(
        PedidoCompra.id == pedido_id,
        PedidoCompra.tenant_id == tenant_id
    ).first()
    
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")
    
    # Lógica de reversão
    status_reverso = {
        'enviado': 'rascunho',
        'confirmado': 'enviado',
        'recebido_parcial': 'confirmado',
        'recebido_total': 'confirmado'
    }
    
    if pedido.status not in status_reverso:
        raise HTTPException(
            status_code=400,
            detail=f"Não é possível reverter pedido com status '{pedido.status}'"
        )
    
    status_anterior = pedido.status
    pedido.status = status_reverso[pedido.status]
    pedido.updated_at = datetime.now()
    
    db.commit()
    logger.info(f"⏪ Pedido {pedido.numero_pedido} revertido: {status_anterior} → {pedido.status}")
    
    return {
        "message": "Status revertido com sucesso",
        "pedido_id": pedido.id,
        "numero_pedido": pedido.numero_pedido,
        "status_anterior": status_anterior,
        "status_atual": pedido.status
    }


# ============================================================================
# 💡 SUGESTÃO INTELIGENTE DE PEDIDO
# ============================================================================

@router.get("/sugestao/{fornecedor_id}")
def sugerir_pedido_inteligente(
    fornecedor_id: int,
    periodo_dias: int = Query(default=90, ge=7, le=365, description="Período de análise (7-365 dias)"),
    dias_cobertura: int = Query(default=30, ge=7, le=180, description="Dias de estoque que o pedido deve cobrir (7-180)"),
    apenas_criticos: bool = Query(default=False, description="Apenas produtos críticos (estoque < 7 dias)"),
    incluir_alerta: bool = Query(default=True, description="Incluir produtos em alerta"),
    marca_ids: Optional[List[int]] = Query(default=None, description="Filtrar por marcas específicas"),
    marca_ids_brackets: Optional[List[int]] = Query(default=None, alias="marca_ids[]"),
    db: Session = Depends(get_session),
    user_and_tenant = Depends(get_current_user_and_tenant)
):
    """
    📊 Sugestão Inteligente de Pedido de Compra
    
    Analisa histórico de vendas, estoque atual e lead time do fornecedor
    para sugerir quantidade ideal a pedir de cada produto.
    
    Algoritmo:
    1. Calcula consumo médio diário baseado em vendas do período
    2. Verifica estoque atual e dias de cobertura restantes
    3. Considera lead time do fornecedor (prazo de entrega)
    4. Sugere quantidade para cobrir: lead_time + dias_cobertura + margem de segurança
    5. Prioriza produtos críticos (estoque baixo) e em alerta
    
    Parâmetros:
    - periodo_dias: Base de cálculo (30/60/90/180 dias)
    - dias_cobertura: Quantos dias de estoque o pedido deve garantir (15/30/45/60/90)
    - apenas_criticos: Filtrar apenas produtos com < 7 dias de estoque
    - incluir_alerta: Incluir produtos que atingiram estoque mínimo
    """
    user, tenant_id = user_and_tenant
    
    logger.info(f"💡 Gerando sugestão de pedido - Fornecedor: {fornecedor_id} | Período: {periodo_dias} dias")
    
    # Validar fornecedor
    fornecedor = db.query(Cliente).filter(
        Cliente.id == fornecedor_id,
        Cliente.tenant_id == tenant_id,
        Cliente.tipo_cadastro == 'fornecedor'
    ).first()
    
    if not fornecedor:
        raise HTTPException(status_code=404, detail="Fornecedor não encontrado")
    
    # Data inicial para análise
    data_inicio = datetime.now() - timedelta(days=periodo_dias)
    
    marcas_filtro = marca_ids or marca_ids_brackets or []

    # Buscar produtos do fornecedor com relacionamento e marca
    produtos_fornecedor_query = db.query(
        Produto,
        ProdutoFornecedor,
        Marca
    ).join(
        ProdutoFornecedor,
        Produto.id == ProdutoFornecedor.produto_id
    ).outerjoin(
        Marca,
        Produto.marca_id == Marca.id
    ).filter(
        Produto.tenant_id == tenant_id,
        Produto.ativo == True,
        ProdutoFornecedor.fornecedor_id == fornecedor_id,
        ProdutoFornecedor.ativo == True
    )

    if marcas_filtro:
        produtos_fornecedor_query = produtos_fornecedor_query.filter(
            Produto.marca_id.in_(marcas_filtro)
        )

    produtos_fornecedor = produtos_fornecedor_query.all()
    
    if not produtos_fornecedor:
        return {
            "fornecedor": {
                "id": fornecedor.id,
                "nome": fornecedor.nome
            },
            "periodo_dias": periodo_dias,
            "sugestoes": [],
            "resumo": {
                "total_produtos": 0,
                "produtos_criticos": 0,
                "produtos_alerta": 0,
                "valor_total_estimado": 0
            },
            "mensagem": "Nenhum produto vinculado a este fornecedor"
        }
    
    sugestoes = []
    total_criticos = 0
    total_alerta = 0
    valor_total = 0

    # Bulk queries de vendas — 2 queries no total em vez de N queries individuais
    ids_produtos = [p.id for p, _pf, _marca in produtos_fornecedor]

    vendas_bulk = dict(
        db.query(VendaItem.produto_id, func.sum(VendaItem.quantidade))
        .join(VendaItem.venda)
        .filter(
            VendaItem.produto_id.in_(ids_produtos),
            VendaItem.tenant_id == tenant_id,
            Venda.data_venda >= data_inicio,
        )
        .group_by(VendaItem.produto_id)
        .all()
    )

    vendas_recentes_bulk: dict = {}
    if periodo_dias >= 60:
        data_recente = datetime.now() - timedelta(days=30)
        vendas_recentes_bulk = dict(
            db.query(VendaItem.produto_id, func.sum(VendaItem.quantidade))
            .join(VendaItem.venda)
            .filter(
                VendaItem.produto_id.in_(ids_produtos),
                VendaItem.tenant_id == tenant_id,
                Venda.data_venda >= data_recente,
            )
            .group_by(VendaItem.produto_id)
            .all()
        )

    for produto, produto_fornecedor, marca in produtos_fornecedor:
        # Consumo no período — lookup em vez de query individual
        vendas_periodo = float(vendas_bulk.get(produto.id) or 0)

        # Consumo médio diário
        consumo_diario = vendas_periodo / periodo_dias if vendas_periodo > 0 else 0

        # Estoque atual
        estoque_atual = float(produto.estoque_atual or 0)
        estoque_minimo = float(produto.estoque_minimo or 0)

        # Dias de estoque restante
        dias_estoque = (estoque_atual / consumo_diario) if consumo_diario > 0 else 999

        # Lead time do fornecedor (padrão 7 dias se não configurado)
        lead_time = produto_fornecedor.prazo_entrega or 7

        # Calcular quantidade sugerida
        # Cobrir: lead_time + dias_cobertura (escolhido pelo usuário) + margem de segurança (7 dias)
        dias_total_cobertura = lead_time + dias_cobertura + 7
        quantidade_ideal = consumo_diario * dias_total_cobertura
        quantidade_sugerida = max(0, quantidade_ideal - estoque_atual)

        # Classificação de prioridade
        if dias_estoque < 7:
            prioridade = "CRÍTICO"
            total_criticos += 1
        elif estoque_atual <= estoque_minimo:
            prioridade = "ALERTA"
            total_alerta += 1
        elif dias_estoque < lead_time + 7:
            prioridade = "ATENÇÃO"
        else:
            prioridade = "NORMAL"

        # Tendência de vendas — lookup em vez de query individual
        if periodo_dias >= 60:
            vendas_recentes = float(vendas_recentes_bulk.get(produto.id) or 0)
            consumo_recente = vendas_recentes / 30

            if consumo_recente > consumo_diario * 1.2:
                tendencia = "CRESCIMENTO"
            elif consumo_recente < consumo_diario * 0.8:
                tendencia = "QUEDA"
            else:
                tendencia = "ESTÁVEL"
        else:
            tendencia = "N/A"

        # Preço unitário
        preco_unitario = float(produto_fornecedor.preco_custo or produto.preco_custo or 0)
        valor_sugestao = quantidade_sugerida * preco_unitario

        # Aplicar filtros
        incluir_produto = True
        
        if apenas_criticos and prioridade != "CRÍTICO":
            incluir_produto = False
        
        if not incluir_alerta and prioridade == "ALERTA":
            incluir_produto = False
        
        # Adicionar à lista (mesmo com qtd 0 para visibilidade se estoque alto)
        if incluir_produto or quantidade_sugerida > 0:
            sugestao = {
                "produto_id": produto.id,
                "produto_nome": produto.nome,
                "produto_sku": produto.codigo,
                "produto_codigo_barras": produto.codigo_barras,
                "marca_id": produto.marca_id,
                "marca_nome": marca.nome if marca else None,
                "estoque_atual": float(estoque_atual),
                "estoque_minimo": float(estoque_minimo),
                "consumo_diario": round(consumo_diario, 2),
                "vendas_periodo": float(vendas_periodo),
                "dias_estoque": round(dias_estoque, 1) if dias_estoque < 999 else None,
                "lead_time": lead_time,
                "quantidade_sugerida": round(quantidade_sugerida, 2),
                "preco_unitario": float(preco_unitario),
                "valor_total": round(valor_sugestao, 2),
                "peso_bruto": float(produto.peso_embalagem or produto.peso_bruto or produto.peso_liquido or 0),
                "prioridade": prioridade,
                "tendencia": tendencia,
                "observacao": _gerar_observacao(prioridade, dias_estoque, tendencia, consumo_diario)
            }
            
            sugestoes.append(sugestao)
            valor_total += valor_sugestao
    
    # Ordenar por prioridade (CRÍTICO > ALERTA > ATENÇÃO > NORMAL)
    ordem_prioridade = {"CRÍTICO": 0, "ALERTA": 1, "ATENÇÃO": 2, "NORMAL": 3}
    sugestoes.sort(key=lambda x: (ordem_prioridade.get(x["prioridade"], 4), -x["valor_total"]))
    
    logger.info(f"✅ Sugestão gerada: {len(sugestoes)} produtos | {total_criticos} críticos | {total_alerta} em alerta")
    
    return {
        "fornecedor": {
            "id": fornecedor.id,
            "nome": fornecedor.nome
        },
        "periodo_dias": periodo_dias,
        "dias_cobertura": dias_cobertura,
        "data_analise_inicio": data_inicio.isoformat(),
        "data_analise_fim": datetime.now().isoformat(),
        "sugestoes": sugestoes,
        "resumo": {
            "total_produtos": len(sugestoes),
            "produtos_criticos": total_criticos,
            "produtos_alerta": total_alerta,
            "produtos_atencao": len([s for s in sugestoes if s["prioridade"] == "ATENÇÃO"]),
            "valor_total_estimado": round(valor_total, 2)
        }
    }


def _gerar_observacao(prioridade: str, dias_estoque: float, tendencia: str, consumo_diario: float) -> str:
    """Gera observação inteligente baseada nos dados"""
    observacoes = []
    
    if prioridade == "CRÍTICO":
        if dias_estoque < 3:
            observacoes.append("🔴 URGENTE - Estoque zerado em menos de 3 dias!")
        else:
            observacoes.append(f"🔴 CRÍTICO - Estoque para apenas {dias_estoque:.1f} dias")
    elif prioridade == "ALERTA":
        observacoes.append("⚠️ Estoque abaixo do mínimo configurado")
    
    if tendencia == "CRESCIMENTO":
        observacoes.append("📈 Vendas em crescimento - considerar pedir mais")
    elif tendencia == "QUEDA":
        observacoes.append("📉 Vendas em queda - avaliar estoque")
    
    if consumo_diario == 0:
        observacoes.append("ℹ️ Sem vendas no período analisado")
    
    return " | ".join(observacoes) if observacoes else "✅ Estoque adequado"


# ============================================================================
# CONFRONTO PEDIDO x NF-e
# ============================================================================

def _realizar_confronto(pedido: PedidoCompra, nota, db: Session, tenant_id: int) -> dict:
    """Gera o confronto completo entre pedido e NF-e."""
    from .produtos_models import NotaEntrada, NotaEntradaItem

    itens_confronto = []
    total_pedido = 0.0
    total_nf = 0.0
    tem_divergencia_qtd = False
    tem_divergencia_preco = False

    for item_pedido in pedido.itens:
        produto = db.query(Produto).filter(
            Produto.id == item_pedido.produto_id,
            Produto.tenant_id == tenant_id
        ).first()

        nome_produto = produto.nome if produto else f"Produto {item_pedido.produto_id}"
        codigo_produto = produto.codigo if produto else None
        ean_produto = produto.codigo_barras if produto else None

        # Tentar matchear item da NF pelo produto_id vinculado, EAN ou código
        item_nf = None
        for it in nota.itens:
            if it.produto_id == item_pedido.produto_id:
                item_nf = it
                break
        if not item_nf and ean_produto:
            for it in nota.itens:
                if it.ean and it.ean == ean_produto:
                    item_nf = it
                    break
        if not item_nf and codigo_produto:
            for it in nota.itens:
                if it.codigo_produto and it.codigo_produto.strip() == str(codigo_produto).strip():
                    item_nf = it
                    break

        qtd_pedida = item_pedido.quantidade_pedida
        preco_pedido = item_pedido.preco_unitario - item_pedido.desconto_item
        valor_pedido = qtd_pedida * preco_pedido
        total_pedido += valor_pedido

        if item_nf:
            qtd_nf = item_nf.quantidade
            preco_nf = item_nf.valor_unitario
            valor_nf = item_nf.valor_total
            total_nf += valor_nf

            dif_qtd = qtd_nf - qtd_pedida
            dif_preco_pct = ((preco_nf - preco_pedido) / preco_pedido * 100) if preco_pedido else 0
            dif_valor = valor_nf - valor_pedido

            if abs(dif_qtd) > 0.001:
                tem_divergencia_qtd = True
            if abs(dif_preco_pct) > 0.5:
                tem_divergencia_preco = True

            status_item = "ok"
            if abs(dif_qtd) > 0.001 and abs(dif_preco_pct) > 0.5:
                status_item = "divergencia_mista"
            elif abs(dif_qtd) > 0.001:
                status_item = "divergencia_quantidade"
            elif abs(dif_preco_pct) > 0.5:
                status_item = "divergencia_preco"

            itens_confronto.append({
                "produto_id": item_pedido.produto_id,
                "produto_nome": nome_produto,
                "produto_codigo": codigo_produto,
                "item_pedido_id": item_pedido.id,
                "item_nf_id": item_nf.id,
                "qtd_pedida": qtd_pedida,
                "qtd_nf": qtd_nf,
                "dif_qtd": round(dif_qtd, 3),
                "preco_pedido": round(preco_pedido, 4),
                "preco_nf": round(preco_nf, 4),
                "dif_preco_pct": round(dif_preco_pct, 2),
                "valor_pedido": round(valor_pedido, 2),
                "valor_nf": round(valor_nf, 2),
                "dif_valor": round(dif_valor, 2),
                "status": status_item,
                "encontrado_na_nf": True,
            })
        else:
            # Item do pedido não encontrado na NF
            tem_divergencia_qtd = True
            itens_confronto.append({
                "produto_id": item_pedido.produto_id,
                "produto_nome": nome_produto,
                "produto_codigo": codigo_produto,
                "item_pedido_id": item_pedido.id,
                "item_nf_id": None,
                "qtd_pedida": qtd_pedida,
                "qtd_nf": 0,
                "dif_qtd": -qtd_pedida,
                "preco_pedido": round(preco_pedido, 4),
                "preco_nf": 0,
                "dif_preco_pct": 0,
                "valor_pedido": round(valor_pedido, 2),
                "valor_nf": 0,
                "dif_valor": -round(valor_pedido, 2),
                "status": "nao_encontrado",
                "encontrado_na_nf": False,
            })

    # Itens na NF que não estavam no pedido
    ids_pedido_produto = {i.produto_id for i in pedido.itens}
    for it in nota.itens:
        if it.produto_id and it.produto_id not in ids_pedido_produto:
            total_nf += it.valor_total
            itens_confronto.append({
                "produto_id": it.produto_id,
                "produto_nome": it.descricao,
                "produto_codigo": it.codigo_produto,
                "item_pedido_id": None,
                "item_nf_id": it.id,
                "qtd_pedida": 0,
                "qtd_nf": it.quantidade,
                "dif_qtd": it.quantidade,
                "preco_pedido": 0,
                "preco_nf": round(it.valor_unitario, 4),
                "dif_preco_pct": 0,
                "valor_pedido": 0,
                "valor_nf": round(it.valor_total, 2),
                "dif_valor": round(it.valor_total, 2),
                "status": "nao_pedido",
                "encontrado_na_nf": True,
            })

    # Status geral do confronto
    if tem_divergencia_qtd and tem_divergencia_preco:
        status_confronto = "divergencia_mista"
    elif tem_divergencia_qtd:
        status_confronto = "divergencia_quantidade"
    elif tem_divergencia_preco:
        status_confronto = "divergencia_preco"
    else:
        status_confronto = "sem_divergencia"

    return {
        "status_confronto": status_confronto,
        "itens": itens_confronto,
        "resumo": {
            "total_pedido": round(total_pedido, 2),
            "total_nf": round(total_nf, 2),
            "dif_total": round(total_nf - total_pedido, 2),
            "frete_pedido": pedido.valor_frete,
            "frete_nf": nota.valor_frete,
            "desconto_pedido": pedido.valor_desconto,
            "desconto_nf": nota.valor_desconto,
            "itens_pedido": len(pedido.itens),
            "itens_nf": len(nota.itens),
        }
    }


@router.get("/{pedido_id}/notas-candidatas")
def listar_notas_candidatas(
    pedido_id: int,
    db: Session = Depends(get_session),
    current_user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Lista NF-e importadas do mesmo fornecedor do pedido, ordenadas pela mais recente."""
    from .produtos_models import NotaEntrada

    current_user, tenant_id = current_user_and_tenant
    pedido = db.query(PedidoCompra).filter(
        PedidoCompra.id == pedido_id,
        PedidoCompra.tenant_id == tenant_id
    ).first()
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")

    # Buscar CNPJ do fornecedor
    fornecedor = db.query(Cliente).filter(
        Cliente.id == pedido.fornecedor_id,
        Cliente.tenant_id == tenant_id
    ).first()

    query = db.query(NotaEntrada).filter(NotaEntrada.tenant_id == tenant_id)

    if fornecedor and fornecedor.cnpj:
        cnpj_limpo = fornecedor.cnpj.replace(".", "").replace("/", "").replace("-", "").strip()
        query = query.filter(
            or_(
                NotaEntrada.fornecedor_id == pedido.fornecedor_id,
                func.replace(func.replace(func.replace(NotaEntrada.fornecedor_cnpj, ".", ""), "/", ""), "-", "") == cnpj_limpo,
            )
        )
    elif fornecedor:
        query = query.filter(NotaEntrada.fornecedor_id == pedido.fornecedor_id)

    notas = query.order_by(desc(NotaEntrada.data_emissao)).limit(20).all()

    result = []
    for n in notas:
        result.append({
            "id": n.id,
            "numero_nota": n.numero_nota,
            "serie": n.serie,
            "chave_acesso": n.chave_acesso,
            "fornecedor_nome": n.fornecedor_nome,
            "data_emissao": n.data_emissao,
            "valor_total": n.valor_total,
            "status": n.status,
            "ja_vinculada": n.id == pedido.nota_entrada_id,
        })

    return {
        "notas": result,
        "nota_vinculada_id": pedido.nota_entrada_id,
    }


@router.post("/{pedido_id}/vincular-nota/{nota_id}")
def vincular_nota_e_confrontar(
    pedido_id: int,
    nota_id: int,
    db: Session = Depends(get_session),
    current_user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Vincula NF-e ao pedido e realiza o confronto completo."""
    from .produtos_models import NotaEntrada

    current_user, tenant_id = current_user_and_tenant
    pedido = db.query(PedidoCompra).options(
        joinedload(PedidoCompra.itens)
    ).filter(
        PedidoCompra.id == pedido_id,
        PedidoCompra.tenant_id == tenant_id
    ).first()
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")

    nota = db.query(NotaEntrada).options(
        joinedload(NotaEntrada.itens)
    ).filter(
        NotaEntrada.id == nota_id,
        NotaEntrada.tenant_id == tenant_id
    ).first()
    if not nota:
        raise HTTPException(status_code=404, detail="Nota fiscal não encontrada")

    confronto = _realizar_confronto(pedido, nota, db, tenant_id)

    # Salvar vínculo e resumo
    pedido.nota_entrada_id = nota_id
    pedido.data_confronto = datetime.utcnow()
    pedido.status_confronto = confronto["status_confronto"]
    pedido.resumo_confronto = json.dumps(confronto, ensure_ascii=False, default=str)
    pedido.updated_at = datetime.utcnow()
    db.commit()

    return {
        "message": "Confronto realizado com sucesso",
        "pedido_id": pedido_id,
        "nota_id": nota_id,
        "confronto": confronto,
    }


@router.get("/{pedido_id}/confronto")
def obter_confronto_salvo(
    pedido_id: int,
    db: Session = Depends(get_session),
    current_user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Retorna o confronto salvo do pedido."""
    current_user, tenant_id = current_user_and_tenant
    pedido = db.query(PedidoCompra).filter(
        PedidoCompra.id == pedido_id,
        PedidoCompra.tenant_id == tenant_id
    ).first()
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")
    if not pedido.nota_entrada_id:
        raise HTTPException(status_code=404, detail="Pedido não possui NF vinculada")

    return {
        "pedido_id": pedido_id,
        "nota_entrada_id": pedido.nota_entrada_id,
        "data_confronto": pedido.data_confronto,
        "status_confronto": pedido.status_confronto,
        "confronto": json.loads(pedido.resumo_confronto) if pedido.resumo_confronto else None,
    }


@router.get("/{pedido_id}/confronto/csv")
def exportar_confronto_csv(
    pedido_id: int,
    db: Session = Depends(get_session),
    current_user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Exporta o confronto do pedido em CSV."""
    current_user, tenant_id = current_user_and_tenant
    pedido = db.query(PedidoCompra).filter(
        PedidoCompra.id == pedido_id,
        PedidoCompra.tenant_id == tenant_id
    ).first()
    if not pedido or not pedido.resumo_confronto:
        raise HTTPException(status_code=404, detail="Confronto não encontrado")

    confronto = json.loads(pedido.resumo_confronto)
    itens = confronto.get("itens", [])
    resumo = confronto.get("resumo", {})

    def fmt(v):
        return str(v).replace(".", ",")

    linhas = [
        "Produto;Código;Qtd Pedida;Qtd NF;Dif. Qtd;Preço Pedido (R$);Preço NF (R$);Dif. Preço (%);Valor Pedido (R$);Valor NF (R$);Dif. Valor (R$);Status"
    ]
    for it in itens:
        linhas.append(
            f"{it.get('produto_nome','')};{it.get('produto_codigo','')};{fmt(it.get('qtd_pedida',0))};{fmt(it.get('qtd_nf',0))};{fmt(it.get('dif_qtd',0))};{fmt(it.get('preco_pedido',0))};{fmt(it.get('preco_nf',0))};{fmt(it.get('dif_preco_pct',0))};{fmt(it.get('valor_pedido',0))};{fmt(it.get('valor_nf',0))};{fmt(it.get('dif_valor',0))};{it.get('status','')}"
        )
    linhas.append("")
    linhas.append(f";;Total Pedido (R$);;{fmt(resumo.get('total_pedido',0))}")
    linhas.append(f";;Total NF (R$);;{fmt(resumo.get('total_nf',0))}")
    linhas.append(f";;Diferença Total (R$);;{fmt(resumo.get('dif_total',0))}")

    csv_content = "\n".join(linhas)
    return StreamingResponse(
        io.StringIO(csv_content),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=confronto_{pedido.numero_pedido}.csv"}
    )


@router.get("/{pedido_id}/confronto/pdf")
def exportar_confronto_pdf(
    pedido_id: int,
    db: Session = Depends(get_session),
    current_user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Exporta o confronto do pedido em PDF."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab não instalado")

    current_user, tenant_id = current_user_and_tenant
    pedido = db.query(PedidoCompra).filter(
        PedidoCompra.id == pedido_id,
        PedidoCompra.tenant_id == tenant_id
    ).first()
    if not pedido or not pedido.resumo_confronto:
        raise HTTPException(status_code=404, detail="Confronto não encontrado")

    confronto = json.loads(pedido.resumo_confronto)
    itens = confronto.get("itens", [])
    resumo = confronto.get("resumo", {})

    fornecedor = db.query(Cliente).filter(
        Cliente.id == pedido.fornecedor_id,
        Cliente.tenant_id == tenant_id
    ).first()
    fornecedor_nome = fornecedor.nome if fornecedor else f"Fornecedor {pedido.fornecedor_id}"

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=12*mm, bottomMargin=12*mm, leftMargin=10*mm, rightMargin=10*mm)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("T", parent=styles["Heading1"], fontSize=16,
                                 textColor=colors.HexColor("#1a56db"), alignment=TA_CENTER, spaceAfter=8)
    sub_style = ParagraphStyle("S", parent=styles["Normal"], fontSize=9, spaceAfter=4)
    small_style = ParagraphStyle("Sm", parent=styles["Normal"], fontSize=7, leading=9)

    elements.append(Paragraph("CONFRONTO PEDIDO x NOTA FISCAL", title_style))
    elements.append(Paragraph(f"Pedido: <b>{pedido.numero_pedido}</b> &nbsp;|&nbsp; Fornecedor: <b>{fornecedor_nome}</b> &nbsp;|&nbsp; Data confronto: <b>{pedido.data_confronto.strftime('%d/%m/%Y %H:%M') if pedido.data_confronto else '-'}</b>", sub_style))
    elements.append(Spacer(1, 4*mm))

    STATUS_LABELS = {
        "ok": "OK",
        "divergencia_quantidade": "Dif. Qtd",
        "divergencia_preco": "Dif. Preço",
        "divergencia_mista": "Dif. Mista",
        "nao_encontrado": "Não Recebido",
        "nao_pedido": "Não Pedido",
    }
    STATUS_COLORS = {
        "ok": colors.HexColor("#d1fae5"),
        "divergencia_quantidade": colors.HexColor("#fef3c7"),
        "divergencia_preco": colors.HexColor("#fef3c7"),
        "divergencia_mista": colors.HexColor("#fee2e2"),
        "nao_encontrado": colors.HexColor("#fee2e2"),
        "nao_pedido": colors.HexColor("#ede9fe"),
    }

    table_data = [["Produto", "Cód.", "Qtd Ped.", "Qtd NF", "Dif.Qtd", "R$ Ped.", "R$ NF", "Dif.%", "Vl.Ped.", "Vl.NF", "Dif.R$", "Status"]]
    row_colors = []

    for idx, it in enumerate(itens):
        st = it.get("status", "ok")
        row_colors.append((idx + 1, STATUS_COLORS.get(st, colors.white)))
        table_data.append([
            Paragraph(it.get("produto_nome", "")[:40], small_style),
            it.get("produto_codigo") or "",
            f"{it.get('qtd_pedida', 0):.2f}".rstrip("0").rstrip("."),
            f"{it.get('qtd_nf', 0):.2f}".rstrip("0").rstrip("."),
            f"{it.get('dif_qtd', 0):+.2f}".rstrip("0").rstrip("."),
            f"R$ {it.get('preco_pedido', 0):.2f}",
            f"R$ {it.get('preco_nf', 0):.2f}",
            f"{it.get('dif_preco_pct', 0):+.1f}%",
            f"R$ {it.get('valor_pedido', 0):.2f}",
            f"R$ {it.get('valor_nf', 0):.2f}",
            f"R$ {it.get('dif_valor', 0):+.2f}",
            STATUS_LABELS.get(st, st),
        ])

    col_widths = [55*mm, 18*mm, 18*mm, 18*mm, 16*mm, 20*mm, 20*mm, 16*mm, 22*mm, 22*mm, 22*mm, 22*mm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a56db")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ALIGN", (2, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
    ]
    for row_idx, color in row_colors:
        style_cmds.append(("BACKGROUND", (0, row_idx), (-1, row_idx), color))
    t.setStyle(TableStyle(style_cmds))
    elements.append(t)
    elements.append(Spacer(1, 5*mm))

    # Resumo financeiro
    resumo_data = [
        ["", "Total Produtos", "Frete", "Desconto"],
        ["Pedido", f"R$ {resumo.get('total_pedido', 0):.2f}", f"R$ {resumo.get('frete_pedido', 0):.2f}", f"R$ {resumo.get('desconto_pedido', 0):.2f}"],
        ["NF",    f"R$ {resumo.get('total_nf', 0):.2f}",    f"R$ {resumo.get('frete_nf', 0):.2f}",    f"R$ {resumo.get('desconto_nf', 0):.2f}"],
        ["Diferença", f"R$ {resumo.get('dif_total', 0):+.2f}", f"R$ {(resumo.get('frete_nf',0)-resumo.get('frete_pedido',0)):+.2f}", "-"],
    ]
    rt = Table(resumo_data, colWidths=[30*mm, 45*mm, 35*mm, 35*mm])
    rt.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(rt)

    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=confronto_{pedido.numero_pedido}.pdf"}
    )


@router.get("/{pedido_id}/confronto/email-texto")
def gerar_email_confronto(
    pedido_id: int,
    db: Session = Depends(get_session),
    current_user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Gera texto de e-mail para enviar ao fornecedor com as divergências."""
    current_user, tenant_id = current_user_and_tenant
    pedido = db.query(PedidoCompra).filter(
        PedidoCompra.id == pedido_id,
        PedidoCompra.tenant_id == tenant_id
    ).first()
    if not pedido or not pedido.resumo_confronto:
        raise HTTPException(status_code=404, detail="Confronto não encontrado")

    confronto = json.loads(pedido.resumo_confronto)
    itens = confronto.get("itens", [])
    resumo = confronto.get("resumo", {})

    fornecedor = db.query(Cliente).filter(Cliente.id == pedido.fornecedor_id, Cliente.tenant_id == tenant_id).first()
    fornecedor_nome = fornecedor.nome if fornecedor else "Fornecedor"

    divergencias = [i for i in itens if i.get("status") != "ok"]

    linhas_divergencia = []
    for d in divergencias:
        partes = []
        st = d.get("status")
        if st in ("divergencia_quantidade", "divergencia_mista", "nao_encontrado"):
            partes.append(f"qtd. pedida: {d['qtd_pedida']}, qtd. recebida: {d['qtd_nf']}")
        if st in ("divergencia_preco", "divergencia_mista"):
            partes.append(f"preço pedido: R$ {d['preco_pedido']:.2f}, preço NF: R$ {d['preco_nf']:.2f} ({d['dif_preco_pct']:+.1f}%)")
        if st == "nao_pedido":
            partes.append("produto não constava no pedido")
        linhas_divergencia.append(f"- {d['produto_nome']}: {'; '.join(partes)}")

    dif_total = resumo.get("dif_total", 0)
    sinal = "a maior" if dif_total > 0 else "a menor"

    corpo = f"""Assunto: Divergências na NF {pedido.nota_entrada_id} referente ao pedido {pedido.numero_pedido}

Prezados {fornecedor_nome},

Ao realizar a conferência do pedido {pedido.numero_pedido} com a nota fiscal recebida, identificamos as seguintes divergências:

{chr(10).join(linhas_divergencia)}

Resumo financeiro:
- Total do pedido: R$ {resumo.get('total_pedido', 0):.2f}
- Total da NF:     R$ {resumo.get('total_nf', 0):.2f}
- Diferença:       R$ {abs(dif_total):.2f} {sinal}

Solicitamos gentilmente que nos informem o motivo das divergências e, se aplicável, o prazo para envio dos itens faltantes ou emissão de nota de crédito pela diferença de valores.

Ficamos à disposição para esclarecimentos.

Atenciosamente,
[Seu nome]
"""

    return {"texto": corpo, "divergencias_count": len(divergencias)}


@router.post("/{pedido_id}/sugerir-pedido-complementar")
def sugerir_pedido_complementar(
    pedido_id: int,
    db: Session = Depends(get_session),
    current_user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Cria pedido rascunho com os itens faltantes após o confronto."""
    current_user, tenant_id = current_user_and_tenant
    pedido = db.query(PedidoCompra).filter(
        PedidoCompra.id == pedido_id,
        PedidoCompra.tenant_id == tenant_id
    ).first()
    if not pedido or not pedido.resumo_confronto:
        raise HTTPException(status_code=404, detail="Confronto não encontrado")

    confronto = json.loads(pedido.resumo_confronto)
    itens_faltantes = [
        i for i in confronto.get("itens", [])
        if i.get("status") in ("nao_encontrado", "divergencia_quantidade")
        and i.get("dif_qtd", 0) < 0
    ]

    if not itens_faltantes:
        raise HTTPException(status_code=400, detail="Não há itens faltantes para criar pedido complementar")

    # Gerar número do pedido
    ultimo = db.query(PedidoCompra).order_by(desc(PedidoCompra.id)).first()
    numero = (ultimo.id + 1) if ultimo else 1
    numero_pedido = f"PC{datetime.now().year}{numero:05d}-C"

    valor_total = 0.0
    itens_novos = []
    for it in itens_faltantes:
        qtd_faltante = abs(it.get("dif_qtd", 0))
        preco = it.get("preco_pedido", 0)
        valor_item = qtd_faltante * preco
        valor_total += valor_item
        itens_novos.append({
            "produto_id": it["produto_id"],
            "qtd": qtd_faltante,
            "preco": preco,
            "valor": valor_item,
        })

    # Criar pedido complementar
    novo_pedido = PedidoCompra(
        numero_pedido=numero_pedido,
        fornecedor_id=pedido.fornecedor_id,
        status="rascunho",
        valor_total=valor_total,
        valor_frete=0,
        valor_desconto=0,
        valor_final=valor_total,
        data_pedido=datetime.utcnow(),
        observacoes=f"Pedido complementar gerado automaticamente após confronto com NF. Pedido original: {pedido.numero_pedido}",
        user_id=current_user.id,
        tenant_id=tenant_id
    )
    db.add(novo_pedido)
    db.flush()

    for it in itens_novos:
        item = PedidoCompraItem(
            pedido_compra_id=novo_pedido.id,
            produto_id=it["produto_id"],
            quantidade_pedida=it["qtd"],
            quantidade_recebida=0,
            preco_unitario=it["preco"],
            desconto_item=0,
            valor_total=it["valor"],
            status="pendente",
            tenant_id=tenant_id
        )
        db.add(item)

    db.commit()

    return {
        "message": "Pedido complementar criado em rascunho",
        "pedido_complementar_id": novo_pedido.id,
        "numero_pedido": numero_pedido,
        "itens_faltantes": len(itens_novos),
        "valor_total": valor_total,
    }

