"""Helpers para extrato financeiro realizado do modulo veterinario."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Iterable, Optional


EXTRATO_COLUNAS_META = {
    "agrupador": {"titulo": "Lancamento", "excel_width": 16, "pdf_width_mm": 24},
    "origem_label": {"titulo": "Origem", "excel_width": 24, "pdf_width_mm": 30},
    "data_hora": {"titulo": "Data", "excel_width": 20, "pdf_width_mm": 25},
    "referencia": {"titulo": "Ref.", "excel_width": 12, "pdf_width_mm": 15},
    "codigo": {"titulo": "Codigo", "excel_width": 16, "pdf_width_mm": 18},
    "nome": {"titulo": "Descricao", "excel_width": 42, "pdf_width_mm": 52},
    "quantidade": {"titulo": "Qtd.", "excel_width": 12, "pdf_width_mm": 14},
    "unidade": {"titulo": "Un.", "excel_width": 10, "pdf_width_mm": 12},
    "custo_unitario": {"titulo": "Custo un.", "excel_width": 14, "pdf_width_mm": 18},
    "custo_total": {"titulo": "Custo", "excel_width": 14, "pdf_width_mm": 18},
    "preco_unitario": {"titulo": "Venda un.", "excel_width": 14, "pdf_width_mm": 18},
    "preco_total": {"titulo": "Venda", "excel_width": 14, "pdf_width_mm": 18},
    "margem_valor": {"titulo": "Margem", "excel_width": 14, "pdf_width_mm": 18},
    "margem_percentual": {"titulo": "Margem %", "excel_width": 12, "pdf_width_mm": 16},
    "contabilizar_label": {"titulo": "Total", "excel_width": 12, "pdf_width_mm": 14},
    "estoque_baixado_label": {"titulo": "Estoque", "excel_width": 14, "pdf_width_mm": 18},
}

EXTRATO_COLUNAS_DEFAULT = [
    "agrupador",
    "origem_label",
    "nome",
    "quantidade",
    "unidade",
    "custo_total",
    "preco_total",
    "margem_valor",
    "contabilizar_label",
]

EXTRATO_COLUNAS_FINANCEIRAS = {
    "custo_unitario",
    "custo_total",
    "preco_unitario",
    "preco_total",
    "margem_valor",
}

ORIGEM_LABELS = {
    "procedimento_consulta": "Procedimento da consulta",
    "insumo_consulta": "Insumo da consulta",
    "procedimento_internacao": "Procedimento da internacao",
    "insumo_internacao": "Insumo da internacao",
}


def _get(obj, key: str, default=None):
    if obj is None:
        return default
    if isinstance(obj, dict):
        return obj.get(key, default)
    return getattr(obj, key, default)


def _as_float(value) -> Optional[float]:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _round_money(value) -> float:
    return round(_as_float(value) or 0.0, 2)


def _round_quantity(value) -> float:
    numero = _as_float(value) or 0.0
    return round(numero, 4)


def _normalizar_data(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _normalizar_texto(value, default: str = "") -> str:
    texto = str(value or "").strip()
    return texto or default


def normalizar_colunas_extrato(colunas: Optional[object]) -> list[str]:
    if colunas is None:
        return EXTRATO_COLUNAS_DEFAULT.copy()
    if isinstance(colunas, str):
        candidatos = colunas.split(",")
    elif isinstance(colunas, (list, tuple, set)):
        candidatos = list(colunas)
    else:
        candidatos = [colunas]

    validas: list[str] = []
    for coluna in candidatos:
        chave = str(coluna or "").strip().lower()
        if chave in EXTRATO_COLUNAS_META and chave not in validas:
            validas.append(chave)
    return validas or EXTRATO_COLUNAS_DEFAULT.copy()


def _produto_payload(produto) -> dict:
    if not produto:
        return {}
    return {
        "id": _get(produto, "id"),
        "codigo": _get(produto, "codigo") or "",
        "nome": _get(produto, "nome") or "",
        "unidade": _get(produto, "unidade") or "",
        "preco_venda": _round_money(_get(produto, "preco_venda")),
        "preco_custo": _round_money(_get(produto, "preco_custo")),
    }


def _normalizar_insumos(insumos: Optional[object]) -> list[dict]:
    if not isinstance(insumos, list):
        return []

    normalizados = []
    for item in insumos:
        if not isinstance(item, dict):
            continue
        produto_id = item.get("produto_id")
        quantidade = _round_quantity(item.get("quantidade"))
        if not produto_id or quantidade <= 0:
            continue
        custo_unitario = _round_money(item.get("custo_unitario"))
        custo_total = _round_money(item.get("custo_total") or (custo_unitario * quantidade))
        normalizados.append({
            **item,
            "produto_id": int(produto_id),
            "quantidade": quantidade,
            "custo_unitario": custo_unitario,
            "custo_total": custo_total,
            "baixar_estoque": bool(item.get("baixar_estoque", True)),
        })
    return normalizados


def coletar_produto_ids_extrato(
    *,
    procedimentos_consulta: Optional[Iterable[object]] = None,
    procedimentos_internacao: Optional[Iterable[dict]] = None,
) -> list[int]:
    ids: set[int] = set()
    for procedimento in list(procedimentos_consulta or []) + list(procedimentos_internacao or []):
        for insumo in _normalizar_insumos(_get(procedimento, "insumos", [])):
            ids.add(int(insumo["produto_id"]))
    return sorted(ids)


def _calcular_margem(preco_total: float, custo_total: float) -> tuple[float, float]:
    margem_valor = _round_money(preco_total - custo_total)
    margem_percentual = round((margem_valor / preco_total) * 100, 2) if preco_total > 0 else 0.0
    return margem_valor, margem_percentual


def _linha_extrato(
    *,
    agrupador: str,
    origem: str,
    nome: str,
    quantidade,
    unidade: Optional[str],
    custo_unitario,
    custo_total,
    preco_unitario,
    preco_total,
    contabilizar_total: bool,
    referencia: str,
    data_hora=None,
    codigo: str = "",
    produto_id: Optional[int] = None,
    consulta_id: Optional[int] = None,
    internacao_id: Optional[int] = None,
    parent_referencia: Optional[str] = None,
    estoque_baixado: bool = False,
    estoque_movimentacao_ids: Optional[list] = None,
    observacoes: Optional[str] = None,
) -> dict:
    quantidade_num = _round_quantity(quantidade)
    custo_unitario_num = _round_money(custo_unitario)
    custo_total_num = _round_money(custo_total)
    preco_unitario_num = _round_money(preco_unitario)
    preco_total_num = _round_money(preco_total)
    margem_valor, margem_percentual = _calcular_margem(preco_total_num, custo_total_num)
    return {
        "agrupador": agrupador,
        "tipo_lancamento": agrupador,
        "origem": origem,
        "origem_label": ORIGEM_LABELS.get(origem, origem.replace("_", " ")),
        "data_hora": _normalizar_data(data_hora),
        "referencia": referencia,
        "parent_referencia": parent_referencia,
        "consulta_id": consulta_id,
        "internacao_id": internacao_id,
        "produto_id": produto_id,
        "codigo": codigo or "",
        "nome": nome,
        "quantidade": quantidade_num,
        "unidade": unidade or "",
        "custo_unitario": custo_unitario_num,
        "custo_total": custo_total_num,
        "preco_unitario": preco_unitario_num,
        "preco_total": preco_total_num,
        "margem_valor": margem_valor,
        "margem_percentual": margem_percentual,
        "contabilizar_total": bool(contabilizar_total),
        "contabilizar_label": "Sim" if contabilizar_total else "Detalhe",
        "estoque_baixado": bool(estoque_baixado),
        "estoque_baixado_label": "Baixado" if estoque_baixado else "Nao",
        "estoque_movimentacao_ids": estoque_movimentacao_ids or [],
        "observacoes": observacoes,
    }


def _linhas_procedimento_consulta(procedimento, produtos_por_id: dict[int, object]) -> list[dict]:
    consulta_id = _get(procedimento, "consulta_id")
    procedimento_id = _get(procedimento, "id")
    insumos = _normalizar_insumos(_get(procedimento, "insumos", []))
    valor = _round_money(_get(procedimento, "valor"))
    custo_total = _round_money(sum(item["custo_total"] for item in insumos))
    tem_valor_cobrado = valor > 0
    referencia = f"CONS-PROC-{procedimento_id or 'novo'}"
    linhas: list[dict] = []

    if tem_valor_cobrado or not insumos:
        linhas.append(_linha_extrato(
            agrupador="consulta",
            origem="procedimento_consulta",
            nome=_normalizar_texto(_get(procedimento, "nome"), "Procedimento da consulta"),
            quantidade=1,
            unidade="servico",
            custo_unitario=custo_total,
            custo_total=custo_total,
            preco_unitario=valor,
            preco_total=valor,
            contabilizar_total=True,
            referencia=referencia,
            data_hora=_get(procedimento, "created_at"),
            consulta_id=consulta_id,
            estoque_baixado=bool(_get(procedimento, "estoque_baixado")),
            estoque_movimentacao_ids=_get(procedimento, "estoque_movimentacao_ids") or [],
            observacoes=_get(procedimento, "observacoes"),
        ))

    for insumo in insumos:
        produto = _produto_payload(produtos_por_id.get(insumo["produto_id"]))
        quantidade = insumo["quantidade"]
        preco_unitario = produto.get("preco_venda") or 0
        preco_total = _round_money(preco_unitario * quantidade)
        linhas.append(_linha_extrato(
            agrupador="consulta",
            origem="insumo_consulta",
            nome=_normalizar_texto(insumo.get("nome") or produto.get("nome"), f"Produto {insumo['produto_id']}"),
            quantidade=quantidade,
            unidade=insumo.get("unidade") or produto.get("unidade"),
            custo_unitario=insumo.get("custo_unitario"),
            custo_total=insumo.get("custo_total"),
            preco_unitario=preco_unitario,
            preco_total=preco_total,
            contabilizar_total=not tem_valor_cobrado,
            referencia=f"CONS-INS-{procedimento_id or 'novo'}-{insumo['produto_id']}",
            parent_referencia=referencia if tem_valor_cobrado else None,
            data_hora=_get(procedimento, "created_at"),
            codigo=produto.get("codigo") or "",
            produto_id=insumo["produto_id"],
            consulta_id=consulta_id,
            estoque_baixado=bool(_get(procedimento, "estoque_baixado")),
            estoque_movimentacao_ids=_get(procedimento, "estoque_movimentacao_ids") or [],
            observacoes=insumo.get("observacoes") or _get(procedimento, "observacoes"),
        ))
    return linhas


def _linhas_procedimento_internacao(procedimento: dict, produtos_por_id: dict[int, object]) -> list[dict]:
    if (procedimento.get("status") or "concluido").strip().lower() != "concluido":
        return []

    internacao_id = procedimento.get("internacao_id")
    procedimento_id = procedimento.get("id")
    insumos = _normalizar_insumos(procedimento.get("insumos") or [])
    valor = _round_money(procedimento.get("valor"))
    custo_total = _round_money(sum(item["custo_total"] for item in insumos))
    tem_valor_cobrado = valor > 0
    referencia = f"INT-PROC-{procedimento_id or 'novo'}"
    data_hora = procedimento.get("horario_execucao") or procedimento.get("data_hora") or procedimento.get("horario_agendado")
    nome = procedimento.get("medicamento") or procedimento.get("nome") or "Procedimento da internacao"
    linhas: list[dict] = []

    if tem_valor_cobrado or not insumos:
        linhas.append(_linha_extrato(
            agrupador="internacao",
            origem="procedimento_internacao",
            nome=_normalizar_texto(nome, "Procedimento da internacao"),
            quantidade=procedimento.get("quantidade_executada") or procedimento.get("quantidade_prevista") or 1,
            unidade=procedimento.get("unidade_quantidade") or "servico",
            custo_unitario=custo_total,
            custo_total=custo_total,
            preco_unitario=valor,
            preco_total=valor,
            contabilizar_total=True,
            referencia=referencia,
            data_hora=data_hora,
            internacao_id=internacao_id,
            estoque_baixado=bool(procedimento.get("estoque_baixado")),
            estoque_movimentacao_ids=procedimento.get("estoque_movimentacao_ids") or [],
            observacoes=procedimento.get("observacao_execucao") or procedimento.get("observacoes_agenda"),
        ))

    for insumo in insumos:
        produto = _produto_payload(produtos_por_id.get(insumo["produto_id"]))
        quantidade = insumo["quantidade"]
        preco_unitario = produto.get("preco_venda") or 0
        preco_total = _round_money(preco_unitario * quantidade)
        linhas.append(_linha_extrato(
            agrupador="internacao",
            origem="insumo_internacao",
            nome=_normalizar_texto(insumo.get("nome") or produto.get("nome"), f"Produto {insumo['produto_id']}"),
            quantidade=quantidade,
            unidade=insumo.get("unidade") or produto.get("unidade"),
            custo_unitario=insumo.get("custo_unitario"),
            custo_total=insumo.get("custo_total"),
            preco_unitario=preco_unitario,
            preco_total=preco_total,
            contabilizar_total=not tem_valor_cobrado,
            referencia=f"INT-INS-{procedimento_id or 'novo'}-{insumo['produto_id']}",
            parent_referencia=referencia if tem_valor_cobrado else None,
            data_hora=data_hora,
            codigo=produto.get("codigo") or "",
            produto_id=insumo["produto_id"],
            internacao_id=internacao_id,
            estoque_baixado=bool(procedimento.get("estoque_baixado")),
            estoque_movimentacao_ids=procedimento.get("estoque_movimentacao_ids") or [],
            observacoes=insumo.get("observacoes") or procedimento.get("observacao_execucao"),
        ))
    return linhas


def _calcular_totais(linhas: list[dict]) -> dict:
    contabilizadas = [linha for linha in linhas if linha.get("contabilizar_total")]
    custo_total = _round_money(sum(_as_float(linha.get("custo_total")) or 0 for linha in contabilizadas))
    preco_total = _round_money(sum(_as_float(linha.get("preco_total")) or 0 for linha in contabilizadas))
    margem_valor, margem_percentual = _calcular_margem(preco_total, custo_total)
    return {
        "custo_total": custo_total,
        "preco_total": preco_total,
        "margem_valor": margem_valor,
        "margem_percentual": margem_percentual,
        "linhas_total": len(linhas),
        "linhas_contabilizadas": len(contabilizadas),
    }


def montar_extrato_atendimento(
    *,
    consulta=None,
    internacoes: Optional[Iterable[object]] = None,
    procedimentos_consulta: Optional[Iterable[object]] = None,
    procedimentos_internacao: Optional[Iterable[dict]] = None,
    produtos_por_id: Optional[dict[int, object]] = None,
    colunas: Optional[object] = None,
) -> dict:
    produtos_por_id = produtos_por_id or {}
    linhas: list[dict] = []

    for procedimento in procedimentos_consulta or []:
        linhas.extend(_linhas_procedimento_consulta(procedimento, produtos_por_id))
    for procedimento in procedimentos_internacao or []:
        linhas.extend(_linhas_procedimento_internacao(procedimento, produtos_por_id))

    linhas.sort(key=lambda item: (
        item.get("data_hora") or "",
        item.get("agrupador") or "",
        0 if item.get("contabilizar_total") else 1,
        item.get("referencia") or "",
    ))
    colunas_normalizadas = normalizar_colunas_extrato(colunas)
    internacoes_lista = list(internacoes or [])
    return {
        "contexto": {
            "consulta_id": _get(consulta, "id"),
            "pet_id": _get(consulta, "pet_id") if consulta is not None else (_get(internacoes_lista[0], "pet_id") if internacoes_lista else None),
            "cliente_id": _get(consulta, "cliente_id"),
            "internacao_ids": [_get(item, "id") for item in internacoes_lista if _get(item, "id") is not None],
        },
        "colunas": colunas_normalizadas,
        "colunas_disponiveis": [
            {"chave": chave, "titulo": meta["titulo"]}
            for chave, meta in EXTRATO_COLUNAS_META.items()
        ],
        "totais": _calcular_totais(linhas),
        "linhas": linhas,
    }


def _valor_documento(coluna: str, valor) -> str:
    if coluna in EXTRATO_COLUNAS_FINANCEIRAS:
        return f"R$ {_round_money(valor):.2f}"
    if coluna == "margem_percentual":
        return f"{_round_money(valor):.2f}%"
    if coluna == "quantidade":
        numero = _round_quantity(valor)
        texto = f"{numero:.4f}".rstrip("0").rstrip(".")
        return texto or "0"
    return str(valor or "")


def gerar_excel_extrato_bytes(extrato: dict, colunas: Optional[object] = None) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("Biblioteca openpyxl nao instalada") from exc

    colunas_exportacao = normalizar_colunas_extrato(colunas or extrato.get("colunas"))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extrato Vet"

    ws["A1"] = "EXTRATO VETERINARIO"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells(f"A1:{get_column_letter(len(colunas_exportacao))}1")

    totais = extrato.get("totais") or {}
    ws["A3"] = "Custo total"
    ws["B3"] = _round_money(totais.get("custo_total"))
    ws["B3"].number_format = 'R$ #,##0.00'
    ws["A4"] = "Preco de venda"
    ws["B4"] = _round_money(totais.get("preco_total"))
    ws["B4"].number_format = 'R$ #,##0.00'
    ws["A5"] = "Margem"
    ws["B5"] = _round_money(totais.get("margem_valor"))
    ws["B5"].number_format = 'R$ #,##0.00'

    row = 7
    for col_idx, coluna in enumerate(colunas_exportacao, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = EXTRATO_COLUNAS_META[coluna]["titulo"]
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    row += 1
    for linha in extrato.get("linhas") or []:
        for col_idx, coluna in enumerate(colunas_exportacao, start=1):
            cell = ws.cell(row=row, column=col_idx)
            valor = linha.get(coluna)
            if coluna in EXTRATO_COLUNAS_FINANCEIRAS:
                cell.value = _round_money(valor)
                cell.number_format = 'R$ #,##0.00'
            elif coluna in {"quantidade", "margem_percentual"}:
                cell.value = _round_quantity(valor)
                cell.number_format = '#,##0.00'
            else:
                cell.value = valor
        row += 1

    for indice, coluna in enumerate(colunas_exportacao, start=1):
        ws.column_dimensions[get_column_letter(indice)].width = EXTRATO_COLUNAS_META[coluna]["excel_width"]

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def gerar_pdf_extrato_bytes(extrato: dict, colunas: Optional[object] = None) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:
        raise RuntimeError("Biblioteca reportlab nao instalada") from exc

    colunas_exportacao = normalizar_colunas_extrato(colunas or extrato.get("colunas"))
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    titulo = ParagraphStyle(
        "ExtratoVetTitulo",
        parent=styles["Heading1"],
        alignment=TA_CENTER,
        fontSize=16,
        textColor=colors.HexColor("#047857"),
    )
    texto_pequeno = ParagraphStyle("ExtratoVetSmall", parent=styles["Normal"], fontSize=7, leading=9)
    elements = [Paragraph("EXTRATO VETERINARIO", titulo), Spacer(1, 5 * mm)]

    totais = extrato.get("totais") or {}
    resumo = [
        ["Custo total", _valor_documento("custo_total", totais.get("custo_total"))],
        ["Preco de venda", _valor_documento("preco_total", totais.get("preco_total"))],
        ["Margem", _valor_documento("margem_valor", totais.get("margem_valor"))],
        ["Linhas", str(totais.get("linhas_total") or 0)],
    ]
    resumo_table = Table(resumo, colWidths=[35 * mm, 35 * mm], hAlign="LEFT")
    resumo_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(resumo_table)
    elements.append(Spacer(1, 6 * mm))

    table_data = [[EXTRATO_COLUNAS_META[coluna]["titulo"] for coluna in colunas_exportacao]]
    for linha in extrato.get("linhas") or []:
        row = []
        for coluna in colunas_exportacao:
            valor = _valor_documento(coluna, linha.get(coluna))
            if coluna == "nome":
                row.append(Paragraph(valor, texto_pequeno))
            else:
                row.append(valor)
        table_data.append(row)

    col_widths = [EXTRATO_COLUNAS_META[coluna]["pdf_width_mm"] * mm for coluna in colunas_exportacao]
    tabela = Table(table_data, colWidths=col_widths, repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#047857")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(tabela)
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
