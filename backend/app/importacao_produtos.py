"""
Sistema de Importação de Produtos via Planilha Excel
Permite criar/atualizar produtos em lote
"""
from fastapi import APIRouter, UploadFile, File, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session as DBSession
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from typing import List, Dict, Any
import logging

from app.db import get_session
from app.auth import get_current_user, get_current_user_and_tenant
from app.models import User
from app.produtos_models import Produto, Categoria, Marca

logger = logging.getLogger(__name__)
router = APIRouter()


def criar_template_excel() -> BytesIO:
    """
    Cria planilha Excel modelo para importação de produtos.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produtos"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalhos (Linha 1)
    headers = [
        'SKU/Código*',
        'Nome*',
        'Descrição',
        'Categoria',
        'Marca',
        'Fornecedor',
        'Código Barras',
        'Preço Custo*',
        'Preço Venda*',
        'Estoque Inicial',
        'Estoque Mínimo',
        'Estoque Máximo',
        'Unidade',
        'Localização',
        'Status',
        'NCM',
        'CEST',
        'CFOP',
        'Origem',
        'ICMS %',
        'PIS %',
        'COFINS %',
        'Observações'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Instruções (Linha 2)
    instructions = ws.cell(row=2, column=1)
    instructions.value = "INSTRUÇÕES: Preencha os dados abaixo. Campos com * são obrigatórios. SKU único identifica o produto para atualização."
    instructions.font = Font(italic=True, color="666666", size=9)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    
    # Exemplos (Linhas 3-5)
    exemplos = [
        ['PROD-001', 'Ração Premium 15kg', 'Ração para cães adultos sabor frango', 'Rações', 'PremiumPet', 'Distribuidora Pet Ltda', '7891234567890', '85.50', '129.90', '50', '10', '100', 'UN', 'A-01', 'ativo', '23099090', '2809300', '5102', '0', '18', '1.65', '7.6', 'Produto em destaque'],
        ['PROD-002', 'Coleira Antipulgas M', 'Coleira antipulgas para cães médios', 'Acessórios', 'SafePet', '', '7891234567891', '15.00', '35.90', '30', '5', '50', 'UN', 'B-15', 'ativo', '42050000', '', '5102', '0', '18', '1.65', '7.6', ''],
        ['PROD-003', 'Shampoo Hipoalergênico 500ml', 'Shampoo para pets com pele sensível', 'Higiene', 'CleanPet', '', '7891234567892', '22.00', '45.90', '25', '5', '40', 'UN', 'C-08', 'ativo', '33051000', '', '5102', '0', '18', '1.65', '7.6', 'Linha premium'],
    ]
    
    for row_num, exemplo in enumerate(exemplos, 3):
        for col_num, value in enumerate(exemplo, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            # Formatação especial para preços
            if col_num in [7, 8]:  # Preço Custo e Preço Venda
                cell.number_format = 'R$ #,##0.00'
    
    # Ajustar largura das colunas
    column_widths = {
        'A': 15,  # SKU
        'B': 30,  # Nome
        'C': 40,  # Descrição
        'D': 15,  # Categoria
        'E': 15,  # Marca
        'F': 25,  # Fornecedor
        'G': 18,  # Código Barras
        'H': 13,  # Preço Custo
        'I': 13,  # Preço Venda
        'J': 15,  # Estoque Inicial
        'K': 15,  # Estoque Mínimo
        'L': 15,  # Estoque Máximo
        'M': 10,  # Unidade
        'N': 15,  # Localização
        'O': 12,  # Status
        'P': 12,  # NCM
        'Q': 12,  # CEST
        'R': 10,  # CFOP
        'S': 10,  # Origem
        'T': 10,  # ICMS %
        'U': 10,  # PIS %
        'V': 10,  # COFINS %
        'W': 30,  # Observações
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Congelar primeira linha
    ws.freeze_panes = 'A3'
    
    # Criar aba de instruções detalhadas
    ws_instrucoes = wb.create_sheet("Instruções")
    instrucoes_texto = [
        ["MANUAL DE IMPORTAÇÃO DE PRODUTOS"],
        [""],
        ["1. CAMPOS OBRIGATÓRIOS (*)"],
        ["   - SKU/Código: Identificador único do produto. Use para atualizar produtos existentes."],
        ["   - Nome: Nome completo do produto"],
        ["   - Preço Custo: Valor de custo do produto (use ponto como separador decimal: 10.50)"],
        ["   - Preço Venda: Valor de venda do produto"],
        [""],
        ["2. CAMPOS OPCIONAIS"],
        ["   - Descrição: Descrição detalhada do produto"],
        ["   - Categoria: Nome da categoria (será criada se não existir)"],
        ["   - Marca: Nome da marca (será criada se não existir)"],
        ["   - Fornecedor: Nome do fornecedor (deve estar CADASTRADO no sistema)"],
        ["   - Código Barras: Código de barras EAN13 ou similar"],
        ["   - Estoque Inicial: Quantidade inicial em estoque (apenas para produtos novos)"],
        ["   - Estoque Mínimo: Quantidade mínima para alerta"],
        ["   - Estoque Máximo: Quantidade máxima recomendada"],
        ["   - Unidade: UN, KG, LT, CX, etc (padrão: UN)"],
        ["   - Localização: Local físico no estoque (ex: A-01, Prateleira 5)"],
        ["   - Status: ativo ou inativo (padrão: ativo)"],
        ["   - NCM: Código NCM - 8 dígitos (ex: 23099090)"],
        ["   - CEST: Código CEST - 7 dígitos (ex: 2809300) - opcional"],
        ["   - CFOP: Código Fiscal de Operações (ex: 5102)"],
        ["   - Origem: 0 a 8 (0=Nacional, 1=Estrangeira importação direta, etc.)"],
        ["   - ICMS %: Alíquota de ICMS em percentual (ex: 18)"],
        ["   - PIS %: Alíquota de PIS em percentual (ex: 1.65)"],
        ["   - COFINS %: Alíquota de COFINS em percentual (ex: 7.6)"],
        ["   - Observações: Informações adicionais"],
        [""],
        ["3. REGRAS DE IMPORTAÇÃO"],
        ["   - Se o SKU já existir, o produto será ATUALIZADO"],
        ["   - Se o SKU não existir, um NOVO produto será criado"],
        ["   - Categorias e Marcas inexistentes serão criadas automaticamente"],
        ["   - Não delete as linhas de exemplo, apenas substitua pelos seus dados"],
        ["   - Mantenha o formato de número com ponto decimal (10.50 e não 10,50)"],
        [""],
        ["4. DICAS"],
        ["   - Use SKUs padronizados (ex: PROD-001, RAC-001, etc)"],
        ["   - Mantenha backup da planilha antes de importar"],
        ["   - Importe em lotes pequenos para facilitar correções"],
        ["   - Verifique os dados após a importação"],
        ["   - Campos tributários são opcionais mas recomendados para emissão de NF-e"],
        ["   - NCM pode ser consultado em: http://www.mdic.gov.br/"],
        ["   - CFOP comum: 5102 (Venda mercadoria adquirida/recebida de terceiros)"],
    ]
    
    for row_num, linha in enumerate(instrucoes_texto, 1):
        cell = ws_instrucoes.cell(row=row_num, column=1)
        cell.value = linha[0]
        if row_num == 1:
            cell.font = Font(bold=True, size=14, color="4472C4")
        elif "CAMPOS" in linha[0] or "REGRAS" in linha[0] or "DICAS" in linha[0]:
            cell.font = Font(bold=True, size=12)
    
    ws_instrucoes.column_dimensions['A'].width = 100
    
    # Salvar em BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


@router.get("/template-importacao")
async def baixar_template_importacao(
    session: DBSession = Depends(get_session),
    user_and_tenant = Depends(get_current_user_and_tenant)
):
    """
    Baixa planilha Excel modelo para importação de produtos.
    """
    current_user, tenant_id = user_and_tenant
    
    try:
        excel_file = criar_template_excel()
        
        filename = f"template_produtos_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    except Exception as e:
        logger.error(f"Erro ao gerar template: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao gerar template: {str(e)}")


@router.post("/importar")
async def importar_produtos(
    file: UploadFile = File(...),
    session: DBSession = Depends(get_session),
    user_and_tenant = Depends(get_current_user_and_tenant)
):
    """
    Importa produtos de planilha Excel.
    Cria novos produtos ou atualiza existentes baseado no SKU.
    """
    current_user, tenant_id = user_and_tenant
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser Excel (.xlsx ou .xls)")
    
    try:
        # Ler arquivo Excel
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
        ws = wb.active
        
        resultados = {
            'sucesso': [],
            'erros': [],
            'atualizados': [],
            'criados': [],
            'total_linhas': 0
        }
        
        # Cache para categorias e marcas
        categorias_cache = {}
        marcas_cache = {}
        
        # Processar linhas (pular cabeçalho e instruções)
        for row_num in range(3, ws.max_row + 1):
            resultados['total_linhas'] += 1
            
            try:
                # Ler valores da linha
                sku = ws.cell(row_num, 1).value
                nome = ws.cell(row_num, 2).value
                descricao = ws.cell(row_num, 3).value
                categoria_nome = ws.cell(row_num, 4).value
                marca_nome = ws.cell(row_num, 5).value
                fornecedor_nome = ws.cell(row_num, 6).value
                codigo_barras = ws.cell(row_num, 7).value
                preco_custo = ws.cell(row_num, 8).value
                preco_venda = ws.cell(row_num, 9).value
                estoque_inicial = ws.cell(row_num, 10).value
                estoque_minimo = ws.cell(row_num, 11).value
                estoque_maximo = ws.cell(row_num, 12).value
                unidade = ws.cell(row_num, 13).value or 'UN'
                localizacao = ws.cell(row_num, 14).value
                status_texto = ws.cell(row_num, 15).value or 'ativo'
                situacao = str(status_texto).lower().strip() == 'ativo'  # Converter para boolean
                
                # Dados fiscais/tributários
                ncm = ws.cell(row_num, 16).value
                cest = ws.cell(row_num, 17).value
                cfop = ws.cell(row_num, 18).value
                origem = ws.cell(row_num, 19).value
                icms_aliquota = ws.cell(row_num, 20).value
                pis_aliquota = ws.cell(row_num, 21).value
                cofins_aliquota = ws.cell(row_num, 22).value
                observacoes = ws.cell(row_num, 23).value
                
                # Validar campos obrigatórios
                if not sku or not nome or preco_venda is None:
                    resultados['erros'].append({
                        'linha': row_num,
                        'sku': sku,
                        'erro': 'Campos obrigatórios faltando (SKU, Nome ou Preço Venda)'
                    })
                    continue
                
                # Converter tipos (suporte para vírgula e ponto como separador decimal)
                try:
                    # Função helper para converter strings com vírgula ou ponto
                    def converter_float(valor):
                        if valor is None or valor == '':
                            return 0
                        # Se for número, retorna direto
                        if isinstance(valor, (int, float)):
                            return float(valor)
                        # Se for string, substitui vírgula por ponto
                        return float(str(valor).replace(',', '.'))
                    
                    preco_custo = converter_float(preco_custo)
                    preco_venda = converter_float(preco_venda)
                    estoque_inicial = int(estoque_inicial) if estoque_inicial else 0
                    estoque_minimo = int(estoque_minimo) if estoque_minimo else 0
                    estoque_maximo = int(estoque_maximo) if estoque_maximo else 0
                    
                    # Converter alíquotas tributárias também
                    icms_aliquota = converter_float(icms_aliquota) if icms_aliquota else None
                    pis_aliquota = converter_float(pis_aliquota) if pis_aliquota else None
                    cofins_aliquota = converter_float(cofins_aliquota) if cofins_aliquota else None
                except ValueError as e:
                    resultados['erros'].append({
                        'linha': row_num,
                        'sku': sku,
                        'erro': f'Erro ao converter números: {str(e)}'
                    })
                    continue
                
                # Buscar ou criar categoria
                categoria_id = None
                if categoria_nome:
                    if categoria_nome in categorias_cache:
                        categoria_id = categorias_cache[categoria_nome]
                    else:
                        categoria = session.query(Categoria).filter(
                            Categoria.nome.ilike(categoria_nome),
                            Categoria.tenant_id == tenant_id
                        ).first()
                        
                        if not categoria:
                            categoria = Categoria(
                                nome=categoria_nome,
                                tenant_id=tenant_id
                            )
                            session.add(categoria)
                            session.flush()
                        
                        categoria_id = categoria.id
                        categorias_cache[categoria_nome] = categoria_id
                
                # Buscar ou criar marca
                marca_id = None
                if marca_nome:
                    if marca_nome in marcas_cache:
                        marca_id = marcas_cache[marca_nome]
                    else:
                        marca = session.query(Marca).filter(
                            Marca.nome.ilike(marca_nome),
                            Marca.tenant_id == tenant_id
                        ).first()
                        
                        if not marca:
                            marca = Marca(
                                nome=marca_nome,
                                tenant_id=tenant_id
                            )
                            session.add(marca)
                            session.flush()
                        
                        marca_id = marca.id
                        marcas_cache[marca_nome] = marca_id
                
                # Buscar fornecedor (obrigatório se fornecido)
                fornecedor_id = None
                if fornecedor_nome:
                    from app.models import Cliente
                    fornecedor = session.query(Cliente).filter(
                        Cliente.nome.ilike(fornecedor_nome),
                        Cliente.tenant_id == tenant_id
                    ).first()
                    
                    if not fornecedor:
                        resultados['erros'].append({
                            'linha': row_num,
                            'sku': sku,
                            'erro': f'Fornecedor "{fornecedor_nome}" não encontrado no sistema'
                        })
                        continue
                    
                    fornecedor_id = fornecedor.id
                
                # Verificar se produto já existe (por SKU ou código)
                produto = session.query(Produto).filter(
                    Produto.tenant_id == tenant_id,
                    Produto.codigo == str(sku)
                ).first()
                
                if produto:
                    # ATUALIZAR produto existente
                    produto.nome = nome
                    produto.descricao_curta = descricao
                    produto.categoria_id = categoria_id
                    produto.marca_id = marca_id
                    produto.fornecedor_id = fornecedor_id
                    produto.codigo_barras = codigo_barras
                    produto.preco_custo = preco_custo
                    produto.preco_venda = preco_venda
                    produto.estoque_minimo = estoque_minimo
                    produto.estoque_maximo = estoque_maximo
                    produto.unidade = unidade
                    produto.localizacao = localizacao
                    produto.situacao = situacao
                    produto.informacoes_adicionais_nf = observacoes
                    
                    # Campos tributários
                    if ncm:
                        produto.ncm = str(ncm).strip()
                    if cest:
                        produto.cest = str(cest).strip()
                    if cfop:
                        produto.cfop = str(cfop).strip()
                    if origem is not None:
                        produto.origem = str(origem).strip()
                    if icms_aliquota is not None:
                        produto.aliquota_icms = icms_aliquota
                    if pis_aliquota is not None:
                        produto.aliquota_pis = pis_aliquota
                    if cofins_aliquota is not None:
                        produto.aliquota_cofins = cofins_aliquota
                    
                    resultados['atualizados'].append({
                        'linha': row_num,
                        'sku': sku,
                        'nome': nome,
                        'id': produto.id
                    })
                else:
                    # CRIAR novo produto
                    produto = Produto(
                        tenant_id=tenant_id,
                        codigo=str(sku),
                        nome=nome,
                        descricao_curta=descricao,
                        categoria_id=categoria_id,
                        marca_id=marca_id,
                        fornecedor_id=fornecedor_id,
                        codigo_barras=codigo_barras,
                        preco_custo=preco_custo,
                        preco_venda=preco_venda,
                        estoque_atual=estoque_inicial,
                        estoque_minimo=estoque_minimo,
                        estoque_maximo=estoque_maximo,
                        unidade=unidade,
                        localizacao=localizacao,
                        situacao=situacao,
                        informacoes_adicionais_nf=observacoes,
                        # Campos tributários
                        ncm=str(ncm).strip() if ncm else None,
                        cest=str(cest).strip() if cest else None,
                        cfop=str(cfop).strip() if cfop else None,
                        origem=str(origem).strip() if origem is not None else None,
                        aliquota_icms=icms_aliquota,
                        aliquota_pis=pis_aliquota,
                        aliquota_cofins=cofins_aliquota
                    )
                    session.add(produto)
                    
                    resultados['criados'].append({
                        'linha': row_num,
                        'sku': sku,
                        'nome': nome
                    })
                
                resultados['sucesso'].append({
                    'linha': row_num,
                    'sku': sku,
                    'nome': nome
                })
                
            except Exception as e:
                logger.error(f"Erro na linha {row_num}: {e}")
                resultados['erros'].append({
                    'linha': row_num,
                    'sku': sku if 'sku' in locals() else 'N/A',
                    'erro': str(e)
                })
        
        # Commit no banco
        session.commit()
        
        # Commit da transação
        session.commit()
        
        # Resumo detalhado
        resumo = {
            'total_processado': resultados['total_linhas'],
            'total_sucesso': len(resultados['sucesso']),
            'total_erros': len(resultados['erros']),
            'total_criados': len(resultados['criados']),
            'total_atualizados': len(resultados['atualizados']),
            'mensagem': f"Importação concluída: {len(resultados['sucesso'])} sucesso, {len(resultados['erros'])} erros",
            'produtos_criados': resultados['criados'],
            'produtos_atualizados': resultados['atualizados'],
            'produtos_com_erro': resultados['erros']
        }
        
        return resumo
        
    except Exception as e:
        session.rollback()
        logger.error(f"Erro ao importar produtos: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao processar planilha: {str(e)}")
