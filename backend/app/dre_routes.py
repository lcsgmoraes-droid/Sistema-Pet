"""
Rotas para DRE (Demonstração do Resultado do Exercício)
Sistema automatizado com categorização inteligente
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, extract
from datetime import datetime, date
from decimal import Decimal
from typing import Optional, List
from pydantic import BaseModel
from io import BytesIO

from .db import get_session
from .auth import get_current_user
from .auth.dependencies import get_current_user_and_tenant
from .models import User
from .vendas_models import Venda, VendaItem, VendaPagamento
from .financeiro_models import ContaPagar, ContaReceber

router = APIRouter(prefix="/financeiro/dre", tags=["DRE"])


# ==================== SCHEMAS ====================

class DREResponse(BaseModel):
    # Período
    periodo: str
    mes: int
    ano: int
    
    # RECEITAS
    receita_bruta: Decimal
    vendas_produtos: Decimal
    vendas_servicos: Decimal
    outras_receitas: Decimal
    
    # DEDUÇÕES
    deducoes_total: Decimal
    descontos: Decimal
    devolucoes: Decimal
    
    # RECEITA LÍQUIDA
    receita_liquida: Decimal
    
    # CUSTOS
    cmv: Decimal  # Custo das Mercadorias Vendidas
    
    # LUCRO BRUTO
    lucro_bruto: Decimal
    margem_bruta: float
    
    # DESPESAS OPERACIONAIS
    despesas_operacionais: Decimal
    despesas_pessoal: Decimal
    despesas_administrativas: Decimal
    taxas_cartao: Decimal
    outras_despesas: Decimal
    
    # RESULTADO OPERACIONAL
    resultado_operacional: Decimal
    margem_operacional: float
    
    # RESULTADO FINANCEIRO
    resultado_financeiro: Decimal
    receitas_financeiras: Decimal
    despesas_financeiras: Decimal
    
    # LUCRO LÍQUIDO
    lucro_liquido: Decimal
    margem_liquida: float
    
    model_config = {"from_attributes": True}


class DREDetalhado(BaseModel):
    """DRE com detalhamento de cada categoria"""
    dre: DREResponse
    detalhes_despesas: List[dict]
    detalhes_receitas: List[dict]
    comparacao_mes_anterior: Optional[dict] = None
    
    model_config = {"from_attributes": True}


# ==================== FUNÇÕES AUXILIARES ====================

def calcular_cmv(db: Session, mes: int, ano: int) -> Decimal:
    """
    Calcula o Custo das Mercadorias Vendidas (CMV)
    CMV = Custo real dos produtos vendidos no período
    """
    # Busca todas as vendas do período
    vendas = db.query(Venda).filter(
        and_(
            extract('month', Venda.data_venda) == mes,
            extract('year', Venda.data_venda) == ano,
            Venda.status.in_(['finalizada', 'pago_nf', 'baixa_parcial'])
        )
    ).all()
    
    cmv_total = Decimal('0')
    
    for venda in vendas:
        # Soma o custo de cada item vendido
        itens = db.query(VendaItem).filter(
            VendaItem.venda_id == venda.id
        ).all()
        
        for item in itens:
            if item.produto and item.produto.preco_custo:
                custo_item = Decimal(str(item.produto.preco_custo)) * item.quantidade
                cmv_total += custo_item
    
    return cmv_total


def obter_despesas_por_categoria(db: Session, mes: int, ano: int) -> dict:
    """
    Agrupa despesas por categoria
    Categorias principais:
    - Despesas com Pessoal (salários, encargos)
    - Despesas Administrativas (água, luz, internet, telefone)
    - Despesas com Ocupação (aluguel, condomínio, IPTU)
    - Despesas com Vendas (marketing, taxas)
    - Outras Despesas
    
    IMPORTANTE: Exclui pagamentos a fornecedores (esses entram no CMV, não em despesas operacionais)
    """
    
    # Busca contas a pagar do período (pagas ou não)
    # EXCLUINDO fornecedores (pagamento de mercadorias já entra no CMV)
    contas_pagar = db.query(ContaPagar).filter(
        and_(
            extract('month', ContaPagar.data_vencimento) == mes,
            extract('year', ContaPagar.data_vencimento) == ano,
            ContaPagar.fornecedor_id.is_(None)  # EXCLUI pagamentos a fornecedores
        )
    ).all()
    
    categorias = {
        'Despesas com Pessoal': Decimal('0'),
        'Despesas Administrativas': Decimal('0'),
        'Despesas com Ocupação': Decimal('0'),
        'Despesas com Vendas': Decimal('0'),
        'Outras Despesas': Decimal('0')
    }
    
    # Palavras-chave para categorização automática
    palavras_pessoal = ['salário', 'salario', 'folha', 'inss', 'fgts', 'vale', 'funcionario', 'funcionário']
    palavras_admin = ['luz', 'água', 'agua', 'internet', 'telefone', 'material', 'limpeza']
    palavras_ocupacao = ['aluguel', 'condomínio', 'condominio', 'iptu']
    palavras_vendas = ['marketing', 'propaganda', 'anúncio', 'anuncio', 'taxa']
    
    for conta in contas_pagar:
        descricao_lower = (conta.descricao or '').lower()
        valor = conta.valor_original
        
        # Categorização inteligente baseada em palavras-chave
        if any(palavra in descricao_lower for palavra in palavras_pessoal):
            categorias['Despesas com Pessoal'] += valor
        elif any(palavra in descricao_lower for palavra in palavras_admin):
            categorias['Despesas Administrativas'] += valor
        elif any(palavra in descricao_lower for palavra in palavras_ocupacao):
            categorias['Despesas com Ocupação'] += valor
        elif any(palavra in descricao_lower for palavra in palavras_vendas):
            categorias['Despesas com Vendas'] += valor
        else:
            categorias['Outras Despesas'] += valor
    
    return categorias


def calcular_taxas_cartao(db: Session, mes: int, ano: int) -> Decimal:
    """
    Calcula o total de taxas de cartão do período
    """
    vendas = db.query(Venda).filter(
        and_(
            extract('month', Venda.data_venda) == mes,
            extract('year', Venda.data_venda) == ano,
            Venda.status.in_(['finalizada', 'pago_nf', 'baixa_parcial'])
        )
    ).all()
    
    taxas_total = Decimal('0')
    
    for venda in vendas:
        pagamentos = db.query(VendaPagamento).filter(
            VendaPagamento.venda_id == venda.id
        ).all()
        
        for pagamento in pagamentos:
            # A tabela venda_pagamentos tem o campo forma_pagamento (texto), não forma_pagamento_id
            # Vamos calcular taxas apenas se houver relacionamento ou pegar da config
            # Por ora, pular cálculo de taxas se não houver relacionamento
            pass
    
    return taxas_total


# ==================== ENDPOINTS ====================

@router.get("", response_model=DREResponse)
def gerar_dre(
    ano: int = Query(..., description="Ano do DRE"),
    mes: int = Query(..., description="Mês do DRE (1-12)"),
    db: Session = Depends(get_session),
    user_and_tenant = Depends(get_current_user_and_tenant)
):
    """
    Gera DRE (Demonstração do Resultado do Exercício) automaticamente
    
    O sistema categoriza automaticamente todas as transações e gera
    um relatório contábil completo seguindo as normas brasileiras.
    """
    
    if mes < 1 or mes > 12:
        raise HTTPException(status_code=400, detail="Mês deve estar entre 1 e 12")
    
    # Nome do mês
    meses = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
             'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    periodo = f"{meses[mes]}/{ano}"
    
    # ========== 1. RECEITAS ==========
    
    # Vendas de produtos e serviços
    vendas = db.query(Venda).filter(
        and_(
            extract('month', Venda.data_venda) == mes,
            extract('year', Venda.data_venda) == ano,
            Venda.status.in_(['finalizada', 'pago_nf', 'baixa_parcial'])
        )
    ).all()
    
    receita_bruta = sum([v.subtotal + (v.taxa_entrega or 0) for v in vendas])
    vendas_produtos = receita_bruta  # Por enquanto, tudo como produtos
    vendas_servicos = Decimal('0')
    outras_receitas = Decimal('0')  # Pode buscar de lançamentos manuais
    
    # ========== 2. DEDUÇÕES ==========
    
    descontos = sum([v.desconto_valor or 0 for v in vendas])
    devolucoes = Decimal('0')  # Implementar quando tiver sistema de devoluções
    deducoes_total = descontos + devolucoes
    
    # ========== 3. RECEITA LÍQUIDA ==========
    
    receita_liquida = receita_bruta - deducoes_total
    
    # ========== 4. CMV ==========
    
    cmv = calcular_cmv(db, mes, ano)
    
    # ========== 5. LUCRO BRUTO ==========
    
    lucro_bruto = receita_liquida - cmv
    margem_bruta = float((lucro_bruto / receita_bruta * 100) if receita_bruta > 0 else 0)
    
    # ========== 6. DESPESAS OPERACIONAIS ==========
    
    categorias_despesas = obter_despesas_por_categoria(db, mes, ano)
    taxas_cartao = calcular_taxas_cartao(db, mes, ano)
    
    despesas_pessoal = categorias_despesas['Despesas com Pessoal']
    despesas_administrativas = (
        categorias_despesas['Despesas Administrativas'] +
        categorias_despesas['Despesas com Ocupação']
    )
    outras_despesas = (
        categorias_despesas['Despesas com Vendas'] +
        categorias_despesas['Outras Despesas']
    )
    
    despesas_operacionais = (
        despesas_pessoal +
        despesas_administrativas +
        taxas_cartao +
        outras_despesas
    )
    
    # ========== 7. RESULTADO OPERACIONAL ==========
    
    resultado_operacional = lucro_bruto - despesas_operacionais
    margem_operacional = float((resultado_operacional / receita_bruta * 100) if receita_bruta > 0 else 0)
    
    # ========== 8. RESULTADO FINANCEIRO ==========
    
    receitas_financeiras = Decimal('0')  # Juros recebidos, etc
    despesas_financeiras = Decimal('0')  # Juros pagos, multas, etc
    resultado_financeiro = receitas_financeiras - despesas_financeiras
    
    # ========== 9. LUCRO LÍQUIDO ==========
    
    lucro_liquido = resultado_operacional + resultado_financeiro
    margem_liquida = float((lucro_liquido / receita_bruta * 100) if receita_bruta > 0 else 0)
    
    # ========== RETORNO ==========
    
    dre = DREResponse(
        periodo=periodo,
        mes=mes,
        ano=ano,
        receita_bruta=receita_bruta,
        vendas_produtos=vendas_produtos,
        vendas_servicos=vendas_servicos,
        outras_receitas=outras_receitas,
        deducoes_total=deducoes_total,
        descontos=descontos,
        devolucoes=devolucoes,
        receita_liquida=receita_liquida,
        cmv=cmv,
        lucro_bruto=lucro_bruto,
        margem_bruta=margem_bruta,
        despesas_operacionais=despesas_operacionais,
        despesas_pessoal=despesas_pessoal,
        despesas_administrativas=despesas_administrativas,
        taxas_cartao=taxas_cartao,
        outras_despesas=outras_despesas,
        resultado_operacional=resultado_operacional,
        margem_operacional=margem_operacional,
        resultado_financeiro=resultado_financeiro,
        receitas_financeiras=receitas_financeiras,
        despesas_financeiras=despesas_financeiras,
        lucro_liquido=lucro_liquido,
        margem_liquida=margem_liquida
    )
    
    return dre


@router.get("/detalhado", response_model=DREDetalhado)
def gerar_dre_detalhado(
    ano: int = Query(..., description="Ano do DRE"),
    mes: int = Query(..., description="Mês do DRE (1-12)"),
    db: Session = Depends(get_session),
    user_and_tenant = Depends(get_current_user_and_tenant)
):
    """
    Gera DRE com detalhamento de cada categoria de despesa e receita
    """
    
    # Gera o DRE básico
    dre = gerar_dre(ano=ano, mes=mes, db=db, current_user=current_user)
    
    # Busca detalhes das despesas (EXCLUINDO fornecedores)
    contas_pagar = db.query(ContaPagar).filter(
        and_(
            extract('month', ContaPagar.data_vencimento) == mes,
            extract('year', ContaPagar.data_vencimento) == ano,
            ContaPagar.fornecedor_id.is_(None)  # EXCLUI pagamentos a fornecedores
        )
    ).all()
    
    detalhes_despesas = [
        {
            'descricao': conta.descricao,
            'valor': float(conta.valor_original),
            'vencimento': conta.data_vencimento.isoformat() if conta.data_vencimento else None,
            'pago': conta.status == 'pago'
        }
        for conta in contas_pagar
    ]
    
    # Busca detalhes das receitas (vendas)
    vendas = db.query(Venda).filter(
        and_(
            extract('month', Venda.data_venda) == mes,
            extract('year', Venda.data_venda) == ano,
            Venda.status.in_(['finalizada', 'pago_nf', 'baixa_parcial'])
        )
    ).all()
    
    detalhes_receitas = [
        {
            'numero_venda': venda.numero_venda,
            'data': venda.data_venda.isoformat(),
            'valor_bruto': float(venda.subtotal + (venda.taxa_entrega or 0)),
            'desconto': float(venda.desconto or 0),
            'valor_liquido': float(venda.total)
        }
        for venda in vendas
    ]
    
    # Comparação com mês anterior (opcional)
    mes_anterior = mes - 1 if mes > 1 else 12
    ano_anterior = ano if mes > 1 else ano - 1
    
    try:
        dre_anterior = gerar_dre(ano=ano_anterior, mes=mes_anterior, db=db, current_user=current_user)
        comparacao = {
            'receita_bruta_variacao': float(dre.receita_bruta - dre_anterior.receita_bruta),
            'lucro_liquido_variacao': float(dre.lucro_liquido - dre_anterior.lucro_liquido),
            'margem_liquida_variacao': dre.margem_liquida - dre_anterior.margem_liquida
        }
    except:
        comparacao = None
    
    return DREDetalhado(
        dre=dre,
        detalhes_despesas=detalhes_despesas,
        detalhes_receitas=detalhes_receitas,
        comparacao_mes_anterior=comparacao
    )


@router.get("/export/pdf")
async def exportar_dre_pdf(
    ano: int = Query(...),
    mes: int = Query(...),
    db: Session = Depends(get_session),
    user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Exporta DRE para PDF"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Biblioteca reportlab não instalada. Execute: pip install reportlab"
        )
    
    # Buscar dados da DRE
    dre = gerar_dre(ano=ano, mes=mes, db=db, current_user=current_user)
    
    # Nomes dos meses
    meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
             'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    mes_nome = meses[mes - 1]
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a56db'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("DEMONSTRAÇÃO DO RESULTADO DO EXERCÍCIO (DRE)", title_style))
    elements.append(Spacer(1, 5*mm))
    
    # Período
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=12,
        alignment=TA_CENTER
    )
    periodo_text = f"Período: {mes_nome}/{ano}"
    elements.append(Paragraph(periodo_text, subtitle_style))
    elements.append(Spacer(1, 10*mm))
    
    # Função para formatar moeda
    def formatar_moeda(valor):
        return f"R$ {float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    
    # Tabela DRE
    dre_data = [
        ['DESCRIÇÃO', 'VALOR', '%'],
        ['', '', ''],
        ['RECEITA BRUTA', formatar_moeda(dre.receita_bruta), '100,00%'],
        ['  Vendas de Produtos', formatar_moeda(dre.vendas_produtos), f"{(float(dre.vendas_produtos)/float(dre.receita_bruta)*100 if float(dre.receita_bruta) > 0 else 0):.2f}%"],
        ['  Vendas de Serviços', formatar_moeda(dre.vendas_servicos), f"{(float(dre.vendas_servicos)/float(dre.receita_bruta)*100 if float(dre.receita_bruta) > 0 else 0):.2f}%"],
        ['  Outras Receitas', formatar_moeda(dre.outras_receitas), f"{(float(dre.outras_receitas)/float(dre.receita_bruta)*100 if float(dre.receita_bruta) > 0 else 0):.2f}%"],
        ['', '', ''],
        ['(-) DEDUÇÕES', formatar_moeda(dre.deducoes_total), f"{dre.margem_bruta - 100:.2f}%"],
        ['  Descontos', formatar_moeda(dre.descontos), ''],
        ['  Devoluções', formatar_moeda(dre.devolucoes), ''],
        ['', '', ''],
        ['(=) RECEITA LÍQUIDA', formatar_moeda(dre.receita_liquida), f"{(float(dre.receita_liquida)/float(dre.receita_bruta)*100 if float(dre.receita_bruta) > 0 else 0):.2f}%"],
        ['', '', ''],
        ['(-) CMV (Custo Mercadorias Vendidas)', formatar_moeda(dre.cmv), ''],
        ['', '', ''],
        ['(=) LUCRO BRUTO', formatar_moeda(dre.lucro_bruto), f"{dre.margem_bruta:.2f}%"],
        ['', '', ''],
        ['(-) DESPESAS OPERACIONAIS', formatar_moeda(dre.despesas_operacionais), ''],
        ['  Despesas de Pessoal', formatar_moeda(dre.despesas_pessoal), ''],
        ['  Despesas Administrativas', formatar_moeda(dre.despesas_administrativas), ''],
        ['  Taxas de Cartão', formatar_moeda(dre.taxas_cartao), ''],
        ['  Outras Despesas', formatar_moeda(dre.outras_despesas), ''],
        ['', '', ''],
        ['(=) RESULTADO OPERACIONAL', formatar_moeda(dre.resultado_operacional), f"{dre.margem_operacional:.2f}%"],
        ['', '', ''],
        ['RESULTADO FINANCEIRO', formatar_moeda(dre.resultado_financeiro), ''],
        ['  (+) Receitas Financeiras', formatar_moeda(dre.receitas_financeiras), ''],
        ['  (-) Despesas Financeiras', formatar_moeda(dre.despesas_financeiras), ''],
        ['', '', ''],
        ['(=) LUCRO LÍQUIDO', formatar_moeda(dre.lucro_liquido), f"{dre.margem_liquida:.2f}%"],
    ]
    
    dre_table = Table(dre_data, colWidths=[100*mm, 40*mm, 30*mm])
    dre_table.setStyle(TableStyle([
        # Cabeçalho
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a56db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        
        # Totais principais (negrito)
        ('FONTNAME', (0, 2), (0, 2), 'Helvetica-Bold'),  # RECEITA BRUTA
        ('FONTNAME', (0, 11), (0, 11), 'Helvetica-Bold'),  # RECEITA LÍQUIDA
        ('FONTNAME', (0, 15), (0, 15), 'Helvetica-Bold'),  # LUCRO BRUTO
        ('FONTNAME', (0, 23), (0, 23), 'Helvetica-Bold'),  # RESULTADO OPERACIONAL
        ('FONTNAME', (0, 29), (0, 29), 'Helvetica-Bold'),  # LUCRO LÍQUIDO
        ('FONTSIZE', (0, 29), (-1, 29), 12),  # LUCRO LÍQUIDO maior
        
        # Background nos totais
        ('BACKGROUND', (0, 2), (-1, 2), colors.lightblue),
        ('BACKGROUND', (0, 11), (-1, 11), colors.lightgreen),
        ('BACKGROUND', (0, 15), (-1, 15), colors.lightyellow),
        ('BACKGROUND', (0, 23), (-1, 23), colors.lightcyan),
        ('BACKGROUND', (0, 29), (-1, 29), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 29), (-1, 29), colors.whitesmoke),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    elements.append(dre_table)
    
    # Rodapé
    elements.append(Spacer(1, 10*mm))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        alignment=TA_CENTER
    )
    elements.append(Paragraph(f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", footer_style))
    
    # Gerar PDF
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=dre_{mes_nome}_{ano}.pdf"}
    )


@router.get("/export/excel")
async def exportar_dre_excel(
    ano: int = Query(...),
    mes: int = Query(...),
    db: Session = Depends(get_session),
    user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Exporta DRE para Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Biblioteca openpyxl não instalada. Execute: pip install openpyxl"
        )
    
    # Buscar dados da DRE
    dre = gerar_dre(ano=ano, mes=mes, db=db, current_user=current_user)
    
    # Nomes dos meses
    meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
             'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    mes_nome = meses[mes - 1]
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DRE"
    
    # Estilos
    title_font = Font(name='Arial', size=14, bold=True, color='1a56db')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1a56db', end_color='1a56db', fill_type='solid')
    total_font = Font(name='Arial', size=10, bold=True)
    total_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    final_fill = PatternFill(start_color='10b981', end_color='10b981', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = 'DEMONSTRAÇÃO DO RESULTADO DO EXERCÍCIO (DRE)'
    ws['A1'].font = title_font
    ws.merge_cells('A1:C1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Período
    ws['A2'] = f'Período: {mes_nome}/{ano}'
    ws.merge_cells('A2:C2')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Cabeçalho
    row = 4
    ws[f'A{row}'] = 'DESCRIÇÃO'
    ws[f'B{row}'] = 'VALOR'
    ws[f'C{row}'] = '%'
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].fill = header_fill
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
    
    # Função para adicionar linha
    def add_row(descricao, valor, percentual='', is_total=False, is_final=False):
        nonlocal row
        row += 1
        ws[f'A{row}'] = descricao
        ws[f'B{row}'] = float(valor) if valor else ''
        ws[f'C{row}'] = percentual
        
        # Aplicar estilos
        if is_final:
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
                ws[f'{col}{row}'].fill = final_fill
        elif is_total:
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = total_font
                ws[f'{col}{row}'].fill = total_fill
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = border
        
        ws[f'B{row}'].number_format = 'R$ #,##0.00'
        ws[f'B{row}'].alignment = Alignment(horizontal='right')
        ws[f'C{row}'].alignment = Alignment(horizontal='right')
    
    # Dados da DRE
    add_row('RECEITA BRUTA', dre.receita_bruta, '100,00%', is_total=True)
    add_row('  Vendas de Produtos', dre.vendas_produtos, f"{(float(dre.vendas_produtos)/float(dre.receita_bruta)*100 if float(dre.receita_bruta) > 0 else 0):.2f}%")
    add_row('  Vendas de Serviços', dre.vendas_servicos, f"{(float(dre.vendas_servicos)/float(dre.receita_bruta)*100 if float(dre.receita_bruta) > 0 else 0):.2f}%")
    add_row('  Outras Receitas', dre.outras_receitas, f"{(float(dre.outras_receitas)/float(dre.receita_bruta)*100 if float(dre.receita_bruta) > 0 else 0):.2f}%")
    row += 1  # Linha em branco
    add_row('(-) DEDUÇÕES', dre.deducoes_total, f"{dre.margem_bruta - 100:.2f}%")
    add_row('  Descontos', dre.descontos)
    add_row('  Devoluções', dre.devolucoes)
    row += 1
    add_row('(=) RECEITA LÍQUIDA', dre.receita_liquida, f"{(float(dre.receita_liquida)/float(dre.receita_bruta)*100 if float(dre.receita_bruta) > 0 else 0):.2f}%", is_total=True)
    row += 1
    add_row('(-) CMV (Custo Mercadorias Vendidas)', dre.cmv)
    row += 1
    add_row('(=) LUCRO BRUTO', dre.lucro_bruto, f"{dre.margem_bruta:.2f}%", is_total=True)
    row += 1
    add_row('(-) DESPESAS OPERACIONAIS', dre.despesas_operacionais)
    add_row('  Despesas de Pessoal', dre.despesas_pessoal)
    add_row('  Despesas Administrativas', dre.despesas_administrativas)
    add_row('  Taxas de Cartão', dre.taxas_cartao)
    add_row('  Outras Despesas', dre.outras_despesas)
    row += 1
    add_row('(=) RESULTADO OPERACIONAL', dre.resultado_operacional, f"{dre.margem_operacional:.2f}%", is_total=True)
    row += 1
    add_row('RESULTADO FINANCEIRO', dre.resultado_financeiro)
    add_row('  (+) Receitas Financeiras', dre.receitas_financeiras)
    add_row('  (-) Despesas Financeiras', dre.despesas_financeiras)
    row += 1
    add_row('(=) LUCRO LÍQUIDO', dre.lucro_liquido, f"{dre.margem_liquida:.2f}%", is_final=True)
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    
    # Salvar em buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=dre_{mes_nome}_{ano}.xlsx"}
    )
