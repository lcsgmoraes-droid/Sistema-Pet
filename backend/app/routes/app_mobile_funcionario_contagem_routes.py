"""Rotas de contagem avulsa do funcionario no app mobile."""

import base64
import html
from datetime import datetime
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import or_
from sqlalchemy.orm import Session, selectinload

from app.db import get_session
from app.models import Cliente, User
from app.produtos_models import FuncionarioContagem, FuncionarioContagemItem, Produto
from app.routes.app_mobile_funcionario_pdv_routes import (
    _get_funcionario_operacional_or_403,
)
from app.routes.ecommerce_auth import _get_current_ecommerce_user

router = APIRouter()


class FuncionarioContagemFornecedorResponse(BaseModel):
    id: int
    nome: str
    documento: Optional[str] = None


class FuncionarioContagemItemRequest(BaseModel):
    produto_id: int
    quantidade: float = Field(gt=0)
    observacao: Optional[str] = None


class FuncionarioContagemRequest(BaseModel):
    titulo: Optional[str] = None
    fornecedor_id: Optional[int] = None
    observacao: Optional[str] = None
    itens: list[FuncionarioContagemItemRequest] = Field(default_factory=list)


class FuncionarioContagemItemResponse(BaseModel):
    id: int
    produto_id: int
    codigo: Optional[str] = None
    codigo_barras: Optional[str] = None
    gtin_ean: Optional[str] = None
    nome: str
    unidade: str
    quantidade: float
    preco_custo: float
    preco_venda: float
    observacao: Optional[str] = None


class FuncionarioContagemResumoResponse(BaseModel):
    id: int
    titulo: str
    status: str
    fornecedor_id: Optional[int] = None
    fornecedor_nome: Optional[str] = None
    observacao: Optional[str] = None
    total_itens: int
    quantidade_total: float
    created_at: datetime


class FuncionarioContagemResponse(FuncionarioContagemResumoResponse):
    itens: list[FuncionarioContagemItemResponse]


class FuncionarioContagemArquivoResponse(BaseModel):
    filename: str
    mime_type: str
    base64: str


def _fornecedor_nome_contagem(fornecedor: Cliente | None) -> Optional[str]:
    if not fornecedor:
        return None
    return (
        getattr(fornecedor, "nome_fantasia", None)
        or getattr(fornecedor, "razao_social", None)
        or fornecedor.nome
    )


def _fornecedor_documento_contagem(fornecedor: Cliente | None) -> Optional[str]:
    if not fornecedor:
        return None
    return getattr(fornecedor, "cnpj", None) or getattr(fornecedor, "cpf", None)


def _serialize_contagem_item(item: FuncionarioContagemItem) -> dict:
    return {
        "id": item.id,
        "produto_id": item.produto_id,
        "codigo": item.codigo,
        "codigo_barras": item.codigo_barras,
        "gtin_ean": item.gtin_ean,
        "nome": item.nome,
        "unidade": item.unidade or "UN",
        "quantidade": float(item.quantidade or 0),
        "preco_custo": float(item.preco_custo_snapshot or 0),
        "preco_venda": float(item.preco_venda_snapshot or 0),
        "observacao": item.observacao,
    }


def _serialize_contagem_resumo(contagem: FuncionarioContagem) -> dict:
    itens = list(getattr(contagem, "itens", []) or [])
    return {
        "id": contagem.id,
        "titulo": contagem.titulo or "Contagem",
        "status": contagem.status or "salva",
        "fornecedor_id": contagem.fornecedor_id,
        "fornecedor_nome": contagem.fornecedor_nome_snapshot,
        "observacao": contagem.observacao,
        "total_itens": len(itens),
        "quantidade_total": sum(float(item.quantidade or 0) for item in itens),
        "created_at": contagem.created_at,
    }


def _serialize_contagem(contagem: FuncionarioContagem) -> dict:
    resumo = _serialize_contagem_resumo(contagem)
    resumo["itens"] = [_serialize_contagem_item(item) for item in contagem.itens]
    return resumo


def _get_contagem_funcionario_or_404(
    db: Session,
    contagem_id: int,
    funcionario_id: int,
    tenant_id,
) -> FuncionarioContagem:
    contagem = (
        db.query(FuncionarioContagem)
        .options(selectinload(FuncionarioContagem.itens))
        .filter(
            FuncionarioContagem.id == contagem_id,
            FuncionarioContagem.tenant_id == tenant_id,
            FuncionarioContagem.funcionario_id == funcionario_id,
        )
        .first()
    )
    if not contagem:
        raise HTTPException(status_code=404, detail="Contagem nao encontrada.")
    return contagem


def _colunas_exportacao_contagem(
    mostrar_custo: bool = False,
    mostrar_venda: bool = False,
) -> list[tuple[str, str]]:
    colunas = [
        ("SKU", "codigo"),
        ("Codigo barras", "codigo_barras"),
        ("Produto", "nome"),
        ("Un.", "unidade"),
        ("Qtd.", "quantidade"),
        ("Obs.", "observacao"),
    ]
    if mostrar_custo:
        colunas.extend(
            [
                ("Custo unitario", "preco_custo_snapshot"),
                ("Total custo", "total_custo"),
            ]
        )
    if mostrar_venda:
        colunas.extend(
            [
                ("Venda unitaria", "preco_venda_snapshot"),
                ("Total venda", "total_venda"),
            ]
        )
    return colunas


def _formatar_quantidade_contagem(valor: float | int | None) -> str:
    numero = float(valor or 0)
    texto = f"{numero:,.3f}".rstrip("0").rstrip(".")
    return texto.replace(",", "X").replace(".", ",").replace("X", ".")


def _formatar_moeda_contagem(valor: float | int | None) -> str:
    numero = float(valor or 0)
    texto = f"{numero:,.2f}"
    return f"R$ {texto}".replace(",", "X").replace(".", ",").replace("X", ".")


def _valor_item_exportacao(
    item: FuncionarioContagemItem,
    chave: str,
    formatado: bool,
):
    quantidade = float(item.quantidade or 0)
    custo = float(item.preco_custo_snapshot or 0)
    venda = float(item.preco_venda_snapshot or 0)
    if chave == "quantidade":
        return _formatar_quantidade_contagem(quantidade) if formatado else quantidade
    if chave == "preco_custo_snapshot":
        return _formatar_moeda_contagem(custo) if formatado else custo
    if chave == "preco_venda_snapshot":
        return _formatar_moeda_contagem(venda) if formatado else venda
    if chave == "total_custo":
        total = quantidade * custo
        return _formatar_moeda_contagem(total) if formatado else total
    if chave == "total_venda":
        total = quantidade * venda
        return _formatar_moeda_contagem(total) if formatado else total
    return getattr(item, chave, None) or ""


def _filename_contagem(contagem: FuncionarioContagem, formato: str) -> str:
    data = datetime.now().strftime("%Y%m%d-%H%M%S")
    extensao = "xlsx" if formato in {"xlsx", "excel"} else "pdf"
    return f"contagem-{contagem.id}-{data}.{extensao}"


def _gerar_pdf_contagem(
    contagem: FuncionarioContagem,
    mostrar_custo: bool,
    mostrar_venda: bool,
) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError as exc:
        raise HTTPException(
            status_code=500, detail="Biblioteca reportlab nao instalada."
        ) from exc

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle(
        "TituloContagem",
        parent=styles["Heading1"],
        alignment=TA_CENTER,
        fontSize=16,
        spaceAfter=8,
    )
    texto_style = styles["Normal"]
    pequeno_style = ParagraphStyle(
        "PequenoContagem", parent=styles["Normal"], fontSize=8
    )

    fornecedor = contagem.fornecedor_nome_snapshot or "Nao informado"
    criado = (
        contagem.created_at.strftime("%d/%m/%Y %H:%M") if contagem.created_at else "-"
    )
    story = [
        Paragraph(html.escape(contagem.titulo or "Contagem"), titulo_style),
        Paragraph(f"<b>Fornecedor:</b> {html.escape(fornecedor)}", texto_style),
        Paragraph(f"<b>Data:</b> {criado}", texto_style),
    ]
    if contagem.observacao:
        story.append(
            Paragraph(
                f"<b>Observacao:</b> {html.escape(contagem.observacao)}", texto_style
            )
        )
    story.append(Spacer(1, 8))

    colunas = _colunas_exportacao_contagem(mostrar_custo, mostrar_venda)
    dados = [[label for label, _chave in colunas]]
    for item in contagem.itens:
        linha = []
        for _label, chave in colunas:
            valor = _valor_item_exportacao(item, chave, formatado=True)
            linha.append(Paragraph(html.escape(str(valor)), pequeno_style))
        dados.append(linha)

    tabela = Table(dados, repeatRows=1)
    tabela.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#F9FAFB")],
                ),
            ]
        )
    )
    story.append(tabela)

    total_quantidade = sum(float(item.quantidade or 0) for item in contagem.itens)
    totais = [f"Quantidade total: {_formatar_quantidade_contagem(total_quantidade)}"]
    if mostrar_custo:
        total_custo = sum(
            float(item.quantidade or 0) * float(item.preco_custo_snapshot or 0)
            for item in contagem.itens
        )
        totais.append(f"Total custo: {_formatar_moeda_contagem(total_custo)}")
    if mostrar_venda:
        total_venda = sum(
            float(item.quantidade or 0) * float(item.preco_venda_snapshot or 0)
            for item in contagem.itens
        )
        totais.append(f"Total venda: {_formatar_moeda_contagem(total_venda)}")
    story.append(Spacer(1, 8))
    story.append(Paragraph(" | ".join(totais), texto_style))
    doc.build(story)
    return buffer.getvalue()


def _gerar_excel_contagem(
    contagem: FuncionarioContagem,
    mostrar_custo: bool,
    mostrar_venda: bool,
) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(
            status_code=500, detail="Biblioteca openpyxl nao instalada."
        ) from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contagem"
    colunas = _colunas_exportacao_contagem(mostrar_custo, mostrar_venda)

    ws.append([contagem.titulo or "Contagem"])
    ws.append(["Fornecedor", contagem.fornecedor_nome_snapshot or "Nao informado"])
    ws.append(
        [
            "Data",
            contagem.created_at.strftime("%d/%m/%Y %H:%M")
            if contagem.created_at
            else "-",
        ]
    )
    if contagem.observacao:
        ws.append(["Observacao", contagem.observacao])
    ws.append([])
    cabecalho_linha = ws.max_row + 1
    ws.append([label for label, _chave in colunas])
    for item in contagem.itens:
        ws.append(
            [
                _valor_item_exportacao(item, chave, formatado=False)
                for _label, chave in colunas
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in ws[cabecalho_linha]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    money_columns = {
        "preco_custo_snapshot",
        "preco_venda_snapshot",
        "total_custo",
        "total_venda",
    }
    for column_index, (_label, chave) in enumerate(colunas, start=1):
        letter = get_column_letter(column_index)
        largura = 16
        if chave == "nome":
            largura = 42
        elif chave in {"observacao", "codigo_barras"}:
            largura = 24
        ws.column_dimensions[letter].width = largura
        if chave == "quantidade":
            for row in range(cabecalho_linha + 1, ws.max_row + 1):
                ws.cell(row=row, column=column_index).number_format = "#,##0.000"
        if chave in money_columns:
            for row in range(cabecalho_linha + 1, ws.max_row + 1):
                ws.cell(row=row, column=column_index).number_format = '"R$" #,##0.00'

    total_row = ws.max_row + 2
    total_quantidade = sum(float(item.quantidade or 0) for item in contagem.itens)
    ws.cell(total_row, 1, "Quantidade total")
    ws.cell(total_row, 2, total_quantidade)
    ws.cell(total_row, 2).number_format = "#,##0.000"
    total_row += 1
    if mostrar_custo:
        ws.cell(total_row, 1, "Total custo")
        ws.cell(
            total_row,
            2,
            sum(
                float(item.quantidade or 0) * float(item.preco_custo_snapshot or 0)
                for item in contagem.itens
            ),
        )
        ws.cell(total_row, 2).number_format = '"R$" #,##0.00'
        total_row += 1
    if mostrar_venda:
        ws.cell(total_row, 1, "Total venda")
        ws.cell(
            total_row,
            2,
            sum(
                float(item.quantidade or 0) * float(item.preco_venda_snapshot or 0)
                for item in contagem.itens
            ),
        )
        ws.cell(total_row, 2).number_format = '"R$" #,##0.00'

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _gerar_arquivo_contagem(
    contagem: FuncionarioContagem,
    formato: str,
    mostrar_custo: bool,
    mostrar_venda: bool,
) -> tuple[bytes, str, str]:
    formato_normalizado = (formato or "").strip().lower()
    if formato_normalizado in {"xlsx", "excel"}:
        arquivo = _gerar_excel_contagem(contagem, mostrar_custo, mostrar_venda)
        return (
            arquivo,
            _filename_contagem(contagem, "xlsx"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    if formato_normalizado == "pdf":
        arquivo = _gerar_pdf_contagem(contagem, mostrar_custo, mostrar_venda)
        return arquivo, _filename_contagem(contagem, "pdf"), "application/pdf"
    raise HTTPException(status_code=400, detail="Formato invalido. Use pdf ou xlsx.")


@router.get(
    "/funcionario/contagens/fornecedores/buscar",
    response_model=list[FuncionarioContagemFornecedorResponse],
)
def buscar_fornecedores_contagem_funcionario(
    q: str = "",
    current_user: User = Depends(_get_current_ecommerce_user),
    db: Session = Depends(get_session),
):
    _funcionario, tenant_id = _get_funcionario_operacional_or_403(db, current_user)
    termo = (q or "").strip()
    if len(termo) < 2:
        return []

    like = f"%{termo}%"
    fornecedores = (
        db.query(Cliente)
        .filter(
            Cliente.tenant_id == tenant_id,
            Cliente.tipo_cadastro == "fornecedor",
            Cliente.ativo.is_(True),
            or_(
                Cliente.nome.ilike(like),
                Cliente.nome_fantasia.ilike(like),
                Cliente.razao_social.ilike(like),
                Cliente.cnpj.ilike(like),
                Cliente.cpf.ilike(like),
                Cliente.codigo.ilike(like),
            ),
        )
        .order_by(Cliente.nome.asc())
        .limit(20)
        .all()
    )
    return [
        {
            "id": fornecedor.id,
            "nome": _fornecedor_nome_contagem(fornecedor) or fornecedor.nome,
            "documento": _fornecedor_documento_contagem(fornecedor),
        }
        for fornecedor in fornecedores
    ]


@router.post("/funcionario/contagens", response_model=FuncionarioContagemResponse)
def criar_contagem_funcionario(
    payload: FuncionarioContagemRequest,
    current_user: User = Depends(_get_current_ecommerce_user),
    db: Session = Depends(get_session),
):
    funcionario, tenant_id = _get_funcionario_operacional_or_403(db, current_user)
    if not payload.itens:
        raise HTTPException(status_code=400, detail="Informe ao menos um item contado.")

    fornecedor = None
    if payload.fornecedor_id:
        fornecedor = (
            db.query(Cliente)
            .filter(
                Cliente.id == payload.fornecedor_id,
                Cliente.tenant_id == tenant_id,
                Cliente.tipo_cadastro == "fornecedor",
                Cliente.ativo.is_(True),
            )
            .first()
        )
        if not fornecedor:
            raise HTTPException(status_code=404, detail="Fornecedor nao encontrado.")

    produto_ids = {item.produto_id for item in payload.itens}
    produtos = (
        db.query(Produto)
        .filter(
            Produto.tenant_id == tenant_id,
            Produto.ativo.is_(True),
            Produto.situacao.is_not(False),
            Produto.id.in_(produto_ids),
        )
        .all()
    )
    produtos_por_id = {produto.id: produto for produto in produtos}
    faltantes = sorted(produto_ids - set(produtos_por_id))
    if faltantes:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Produto nao encontrado: {faltantes[0]}",
        )

    contagem = FuncionarioContagem(
        tenant_id=tenant_id,
        funcionario_id=funcionario.id,
        user_id=current_user.id,
        fornecedor_id=fornecedor.id if fornecedor else None,
        fornecedor_nome_snapshot=_fornecedor_nome_contagem(fornecedor),
        titulo=(payload.titulo or "Contagem para devolucao").strip()[:160],
        observacao=(payload.observacao or "").strip() or None,
        status="salva",
    )
    db.add(contagem)
    db.flush()

    for ordem, item_payload in enumerate(payload.itens, start=1):
        produto = produtos_por_id[item_payload.produto_id]
        item = FuncionarioContagemItem(
            tenant_id=tenant_id,
            contagem_id=contagem.id,
            produto_id=produto.id,
            ordem=ordem,
            codigo=produto.codigo,
            codigo_barras=produto.codigo_barras,
            gtin_ean=produto.gtin_ean,
            nome=produto.nome,
            unidade=produto.unidade or "UN",
            quantidade=float(item_payload.quantidade),
            preco_custo_snapshot=float(produto.preco_custo or 0),
            preco_venda_snapshot=float(produto.preco_venda or 0),
            observacao=(item_payload.observacao or "").strip() or None,
        )
        db.add(item)

    db.commit()
    contagem = _get_contagem_funcionario_or_404(
        db, contagem.id, funcionario.id, tenant_id
    )
    return _serialize_contagem(contagem)


@router.get(
    "/funcionario/contagens", response_model=list[FuncionarioContagemResumoResponse]
)
def listar_contagens_funcionario(
    limit: int = Query(default=20, ge=1, le=100),
    current_user: User = Depends(_get_current_ecommerce_user),
    db: Session = Depends(get_session),
):
    funcionario, tenant_id = _get_funcionario_operacional_or_403(db, current_user)
    contagens = (
        db.query(FuncionarioContagem)
        .options(selectinload(FuncionarioContagem.itens))
        .filter(
            FuncionarioContagem.tenant_id == tenant_id,
            FuncionarioContagem.funcionario_id == funcionario.id,
        )
        .order_by(FuncionarioContagem.created_at.desc(), FuncionarioContagem.id.desc())
        .limit(limit)
        .all()
    )
    return [_serialize_contagem_resumo(contagem) for contagem in contagens]


@router.get(
    "/funcionario/contagens/{contagem_id}", response_model=FuncionarioContagemResponse
)
def obter_contagem_funcionario(
    contagem_id: int,
    current_user: User = Depends(_get_current_ecommerce_user),
    db: Session = Depends(get_session),
):
    funcionario, tenant_id = _get_funcionario_operacional_or_403(db, current_user)
    contagem = _get_contagem_funcionario_or_404(
        db, contagem_id, funcionario.id, tenant_id
    )
    return _serialize_contagem(contagem)


@router.get("/funcionario/contagens/{contagem_id}/export/{formato}")
def exportar_contagem_funcionario(
    contagem_id: int,
    formato: str,
    mostrar_custo: bool = False,
    mostrar_venda: bool = False,
    current_user: User = Depends(_get_current_ecommerce_user),
    db: Session = Depends(get_session),
):
    funcionario, tenant_id = _get_funcionario_operacional_or_403(db, current_user)
    contagem = _get_contagem_funcionario_or_404(
        db, contagem_id, funcionario.id, tenant_id
    )
    arquivo, filename, mime_type = _gerar_arquivo_contagem(
        contagem,
        formato,
        mostrar_custo=mostrar_custo,
        mostrar_venda=mostrar_venda,
    )
    return StreamingResponse(
        BytesIO(arquivo),
        media_type=mime_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/funcionario/contagens/{contagem_id}/export/{formato}/mobile",
    response_model=FuncionarioContagemArquivoResponse,
)
def exportar_contagem_funcionario_mobile(
    contagem_id: int,
    formato: str,
    mostrar_custo: bool = False,
    mostrar_venda: bool = False,
    current_user: User = Depends(_get_current_ecommerce_user),
    db: Session = Depends(get_session),
):
    funcionario, tenant_id = _get_funcionario_operacional_or_403(db, current_user)
    contagem = _get_contagem_funcionario_or_404(
        db, contagem_id, funcionario.id, tenant_id
    )
    arquivo, filename, mime_type = _gerar_arquivo_contagem(
        contagem,
        formato,
        mostrar_custo=mostrar_custo,
        mostrar_venda=mostrar_venda,
    )
    return {
        "filename": filename,
        "mime_type": mime_type,
        "base64": base64.b64encode(arquivo).decode("ascii"),
    }
