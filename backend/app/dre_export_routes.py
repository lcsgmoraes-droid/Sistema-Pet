"""Exportacoes PDF e Excel da DRE."""

from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from .auth.dependencies import get_current_user_and_tenant
from .db import get_session
from .dre_base_routes import gerar_dre

router = APIRouter(prefix="/financeiro/dre", tags=["DRE"])


@router.get("/export/pdf")
async def exportar_dre_pdf(
    ano: int = Query(...),
    mes: int = Query(...),
    db: Session = Depends(get_session),
    user_and_tenant=Depends(get_current_user_and_tenant),
):
    """Exporta DRE para PDF"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Biblioteca reportlab não instalada. Execute: pip install reportlab",
        )

    # Buscar dados da DRE
    dre = gerar_dre(ano=ano, mes=mes, db=db, user_and_tenant=user_and_tenant)

    # Nomes dos meses
    meses = [
        "Janeiro",
        "Fevereiro",
        "Março",
        "Abril",
        "Maio",
        "Junho",
        "Julho",
        "Agosto",
        "Setembro",
        "Outubro",
        "Novembro",
        "Dezembro",
    ]
    mes_nome = meses[mes - 1]

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4, topMargin=15 * mm, bottomMargin=15 * mm
    )
    elements = []
    styles = getSampleStyleSheet()

    # Título
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=colors.HexColor("#1a56db"),
        spaceAfter=12,
        alignment=TA_CENTER,
    )
    elements.append(
        Paragraph("DEMONSTRAÇÃO DO RESULTADO DO EXERCÍCIO (DRE)", title_style)
    )
    elements.append(Spacer(1, 5 * mm))

    # Período
    subtitle_style = ParagraphStyle(
        "Subtitle", parent=styles["Normal"], fontSize=12, alignment=TA_CENTER
    )
    periodo_text = f"Período: {mes_nome}/{ano}"
    elements.append(Paragraph(periodo_text, subtitle_style))
    elements.append(Spacer(1, 10 * mm))

    # Função para formatar moeda
    def formatar_moeda(valor):
        return (
            f"R$ {float(valor):,.2f}".replace(",", "X")
            .replace(".", ",")
            .replace("X", ".")
        )

    # Tabela DRE
    dre_data = [
        ["DESCRIÇÃO", "VALOR", "%"],
        ["", "", ""],
        ["RECEITA BRUTA", formatar_moeda(dre.receita_bruta), "100,00%"],
        [
            "  Vendas de Produtos",
            formatar_moeda(dre.vendas_produtos),
            f"{(float(dre.vendas_produtos) / float(dre.receita_bruta) * 100 if float(dre.receita_bruta) > 0 else 0):.2f}%",
        ],
        [
            "  Vendas de Serviços",
            formatar_moeda(dre.vendas_servicos),
            f"{(float(dre.vendas_servicos) / float(dre.receita_bruta) * 100 if float(dre.receita_bruta) > 0 else 0):.2f}%",
        ],
        [
            "  Outras Receitas",
            formatar_moeda(dre.outras_receitas),
            f"{(float(dre.outras_receitas) / float(dre.receita_bruta) * 100 if float(dre.receita_bruta) > 0 else 0):.2f}%",
        ],
        ["", "", ""],
        [
            "(-) DEDUÇÕES",
            formatar_moeda(dre.deducoes_total),
            f"{dre.margem_bruta - 100:.2f}%",
        ],
        ["  Descontos", formatar_moeda(dre.descontos), ""],
        ["  Devoluções", formatar_moeda(dre.devolucoes), ""],
        ["", "", ""],
        [
            "(=) RECEITA LÍQUIDA",
            formatar_moeda(dre.receita_liquida),
            f"{(float(dre.receita_liquida) / float(dre.receita_bruta) * 100 if float(dre.receita_bruta) > 0 else 0):.2f}%",
        ],
        ["", "", ""],
        ["(-) CMV (Custo Mercadorias Vendidas)", formatar_moeda(dre.cmv), ""],
        ["", "", ""],
        [
            "(=) LUCRO BRUTO",
            formatar_moeda(dre.lucro_bruto),
            f"{dre.margem_bruta:.2f}%",
        ],
        ["", "", ""],
        ["(-) DESPESAS OPERACIONAIS", formatar_moeda(dre.despesas_operacionais), ""],
        ["  Despesas de Pessoal", formatar_moeda(dre.despesas_pessoal), ""],
        [
            "  Despesas Administrativas",
            formatar_moeda(dre.despesas_administrativas),
            "",
        ],
        ["  Taxas de Cartão", formatar_moeda(dre.taxas_cartao), ""],
        ["  Outras Despesas", formatar_moeda(dre.outras_despesas), ""],
        ["", "", ""],
        [
            "(=) RESULTADO OPERACIONAL",
            formatar_moeda(dre.resultado_operacional),
            f"{dre.margem_operacional:.2f}%",
        ],
        ["", "", ""],
        ["RESULTADO FINANCEIRO", formatar_moeda(dre.resultado_financeiro), ""],
        ["  (+) Receitas Financeiras", formatar_moeda(dre.receitas_financeiras), ""],
        ["  (-) Despesas Financeiras", formatar_moeda(dre.despesas_financeiras), ""],
        ["", "", ""],
        [
            "(=) LUCRO LÍQUIDO",
            formatar_moeda(dre.lucro_liquido),
            f"{dre.margem_liquida:.2f}%",
        ],
    ]

    dre_table = Table(dre_data, colWidths=[100 * mm, 40 * mm, 30 * mm])
    dre_table.setStyle(
        TableStyle(
            [
                # Cabeçalho
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a56db")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                # Totais principais (negrito)
                ("FONTNAME", (0, 2), (0, 2), "Helvetica-Bold"),  # RECEITA BRUTA
                ("FONTNAME", (0, 11), (0, 11), "Helvetica-Bold"),  # RECEITA LÍQUIDA
                ("FONTNAME", (0, 15), (0, 15), "Helvetica-Bold"),  # LUCRO BRUTO
                (
                    "FONTNAME",
                    (0, 23),
                    (0, 23),
                    "Helvetica-Bold",
                ),  # RESULTADO OPERACIONAL
                ("FONTNAME", (0, 29), (0, 29), "Helvetica-Bold"),  # LUCRO LÍQUIDO
                ("FONTSIZE", (0, 29), (-1, 29), 12),  # LUCRO LÍQUIDO maior
                # Background nos totais
                ("BACKGROUND", (0, 2), (-1, 2), colors.lightblue),
                ("BACKGROUND", (0, 11), (-1, 11), colors.lightgreen),
                ("BACKGROUND", (0, 15), (-1, 15), colors.lightyellow),
                ("BACKGROUND", (0, 23), (-1, 23), colors.lightcyan),
                ("BACKGROUND", (0, 29), (-1, 29), colors.HexColor("#10b981")),
                ("TEXTCOLOR", (0, 29), (-1, 29), colors.whitesmoke),
                # Grid
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
            ]
        )
    )
    elements.append(dre_table)

    # Rodapé
    elements.append(Spacer(1, 10 * mm))
    footer_style = ParagraphStyle(
        "Footer", parent=styles["Normal"], fontSize=8, alignment=TA_CENTER
    )
    elements.append(
        Paragraph(
            f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", footer_style
        )
    )

    # Gerar PDF
    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=dre_{mes_nome}_{ano}.pdf"
        },
    )


@router.get("/export/excel")
async def exportar_dre_excel(
    ano: int = Query(...),
    mes: int = Query(...),
    db: Session = Depends(get_session),
    user_and_tenant=Depends(get_current_user_and_tenant),
):
    """Exporta DRE para Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Biblioteca openpyxl não instalada. Execute: pip install openpyxl",
        )

    # Buscar dados da DRE
    dre = gerar_dre(ano=ano, mes=mes, db=db, user_and_tenant=user_and_tenant)

    # Nomes dos meses
    meses = [
        "Janeiro",
        "Fevereiro",
        "Março",
        "Abril",
        "Maio",
        "Junho",
        "Julho",
        "Agosto",
        "Setembro",
        "Outubro",
        "Novembro",
        "Dezembro",
    ]
    mes_nome = meses[mes - 1]

    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DRE"

    # Estilos
    title_font = Font(name="Arial", size=14, bold=True, color="1a56db")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1a56db", end_color="1a56db", fill_type="solid"
    )
    total_font = Font(name="Arial", size=10, bold=True)
    total_fill = PatternFill(
        start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
    )
    final_fill = PatternFill(
        start_color="10b981", end_color="10b981", fill_type="solid"
    )
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Título
    ws["A1"] = "DEMONSTRAÇÃO DO RESULTADO DO EXERCÍCIO (DRE)"
    ws["A1"].font = title_font
    ws.merge_cells("A1:C1")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Período
    ws["A2"] = f"Período: {mes_nome}/{ano}"
    ws.merge_cells("A2:C2")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Cabeçalho
    row = 4
    ws[f"A{row}"] = "DESCRIÇÃO"
    ws[f"B{row}"] = "VALOR"
    ws[f"C{row}"] = "%"
    for col in ["A", "B", "C"]:
        ws[f"{col}{row}"].font = header_font
        ws[f"{col}{row}"].fill = header_fill
        ws[f"{col}{row}"].border = border
        ws[f"{col}{row}"].alignment = Alignment(horizontal="center")

    # Função para adicionar linha
    def add_row(descricao, valor, percentual="", is_total=False, is_final=False):
        nonlocal row
        row += 1
        ws[f"A{row}"] = descricao
        ws[f"B{row}"] = float(valor) if valor else ""
        ws[f"C{row}"] = percentual

        # Aplicar estilos
        if is_final:
            for col in ["A", "B", "C"]:
                ws[f"{col}{row}"].font = Font(bold=True, color="FFFFFF")
                ws[f"{col}{row}"].fill = final_fill
        elif is_total:
            for col in ["A", "B", "C"]:
                ws[f"{col}{row}"].font = total_font
                ws[f"{col}{row}"].fill = total_fill

        for col in ["A", "B", "C"]:
            ws[f"{col}{row}"].border = border

        ws[f"B{row}"].number_format = "R$ #,##0.00"
        ws[f"B{row}"].alignment = Alignment(horizontal="right")
        ws[f"C{row}"].alignment = Alignment(horizontal="right")

    # Dados da DRE
    add_row("RECEITA BRUTA", dre.receita_bruta, "100,00%", is_total=True)
    add_row(
        "  Vendas de Produtos",
        dre.vendas_produtos,
        f"{(float(dre.vendas_produtos) / float(dre.receita_bruta) * 100 if float(dre.receita_bruta) > 0 else 0):.2f}%",
    )
    add_row(
        "  Vendas de Serviços",
        dre.vendas_servicos,
        f"{(float(dre.vendas_servicos) / float(dre.receita_bruta) * 100 if float(dre.receita_bruta) > 0 else 0):.2f}%",
    )
    add_row(
        "  Outras Receitas",
        dre.outras_receitas,
        f"{(float(dre.outras_receitas) / float(dre.receita_bruta) * 100 if float(dre.receita_bruta) > 0 else 0):.2f}%",
    )
    row += 1  # Linha em branco
    add_row("(-) DEDUÇÕES", dre.deducoes_total, f"{dre.margem_bruta - 100:.2f}%")
    add_row("  Descontos", dre.descontos)
    add_row("  Devoluções", dre.devolucoes)
    row += 1
    add_row(
        "(=) RECEITA LÍQUIDA",
        dre.receita_liquida,
        f"{(float(dre.receita_liquida) / float(dre.receita_bruta) * 100 if float(dre.receita_bruta) > 0 else 0):.2f}%",
        is_total=True,
    )
    row += 1
    add_row("(-) CMV (Custo Mercadorias Vendidas)", dre.cmv)
    row += 1
    add_row(
        "(=) LUCRO BRUTO", dre.lucro_bruto, f"{dre.margem_bruta:.2f}%", is_total=True
    )
    row += 1
    add_row("(-) DESPESAS OPERACIONAIS", dre.despesas_operacionais)
    add_row("  Despesas de Pessoal", dre.despesas_pessoal)
    add_row("  Despesas Administrativas", dre.despesas_administrativas)
    add_row("  Taxas de Cartão", dre.taxas_cartao)
    add_row("  Outras Despesas", dre.outras_despesas)
    row += 1
    add_row(
        "(=) RESULTADO OPERACIONAL",
        dre.resultado_operacional,
        f"{dre.margem_operacional:.2f}%",
        is_total=True,
    )
    row += 1
    add_row("RESULTADO FINANCEIRO", dre.resultado_financeiro)
    add_row("  (+) Receitas Financeiras", dre.receitas_financeiras)
    add_row("  (-) Despesas Financeiras", dre.despesas_financeiras)
    row += 1
    add_row(
        "(=) LUCRO LÍQUIDO",
        dre.lucro_liquido,
        f"{dre.margem_liquida:.2f}%",
        is_final=True,
    )

    # Ajustar largura das colunas
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15

    # Salvar em buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=dre_{mes_nome}_{ano}.xlsx"
        },
    )
