"""
Sistema de importação de Pessoas (Clientes, Fornecedores, Veterinários)
"""
from fastapi import APIRouter, UploadFile, File, Depends
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from decimal import Decimal
from datetime import datetime
import io

from .models import Cliente
from .db import get_session as get_db
from .auth import get_current_user, get_current_user_and_tenant

router = APIRouter(prefix="/pessoas", tags=["pessoas"])

def converter_float(valor):
    """Converte string para float, aceitando tanto '.' quanto ',' como separador decimal"""
    if not valor or valor == '':
        return None
    
    # Se já é número, retorna
    if isinstance(valor, (int, float)):
        return float(valor)
    
    # Converte string
    valor_str = str(valor).strip()
    
    # Substitui vírgula por ponto se necessário
    valor_str = valor_str.replace(',', '.')
    
    try:
        return float(valor_str)
    except:
        return None


@router.get("/template-importacao")
async def criar_template_pessoas(current_user = Depends(get_current_user)):
    """Gera template Excel para importação de Pessoas"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Pessoas"
    
    # Largura das colunas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 20
    ws.column_dimensions['Q'].width = 15
    ws.column_dimensions['R'].width = 30
    
    # Estilo de header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Instrução
    ws['A1'] = "INSTRUÇÕES: Preencha os dados abaixo para importar Pessoas (Clientes, Fornecedores, Veterinários)"
    ws['A1'].font = Font(bold=True, size=11, color="FF0000")
    ws.merge_cells('A1:R1')
    
    # Headers (linha 2)
    headers = [
        "Tipo Cadastro",      # A - cliente, fornecedor, veterinario
        "Tipo Pessoa",        # B - PF ou PJ
        "Código",             # C - Código único
        "Nome",               # D - Nome (PF) ou Nome Fantasia (PJ)
        "CPF",                # E - Para PF
        "CNPJ",               # F - Para PJ
        "Razão Social",       # G - Para PJ
        "Responsável",        # H - Contato responsável
        "CRMV",               # I - Para veterinário
        "Inscrição Estadual", # J - Para PJ
        "Telefone",           # K
        "Celular",            # L
        "Email",              # M
        "CEP",                # N
        "Endereço",           # O
        "Número",             # P
        "Complemento",        # Q
        "Bairro"              # R
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Exemplos de dados
    exemplos = [
        # Cliente PF
        ["cliente", "PF", "1001", "João Silva", "123.456.789-00", "", "", "João Silva", "", "", "1133334444", "11999998888", "joao@email.com", "01234-567", "Rua A", "100", "Apto 10", "Vila A"],
        # Fornecedor PJ
        ["fornecedor", "PJ", "2001", "Fornecedor XYZ", "", "12.345.678/0001-90", "Fornecedor XYZ LTDA", "Maria Souza", "", "123.456.789.012", "1133334444", "11999997777", "fornecedor@email.com", "02345-678", "Rua B", "200", "", "Vila B"],
        # Veterinário PF
        ["veterinario", "PF", "3001", "Dr. Carlos", "987.654.321-00", "", "", "Dr. Carlos Oliveira", "12345-SP", "", "1133335555", "11999996666", "carlos@email.com", "03456-789", "Rua C", "300", "sala 5", "Vila C"],
    ]
    
    for row_num, dados in enumerate(exemplos, 3):
        for col_num, valor in enumerate(dados, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = valor
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border
    
    # Folha de orientações
    ws_info = wb.create_sheet("Orientações")
    ws_info.column_dimensions['A'].width = 50
    ws_info.column_dimensions['B'].width = 50
    
    orientacoes = [
        ("CAMPO", "DESCRIÇÃO"),
        ("", ""),
        ("Tipo Cadastro", "Obrigatório. Valores: cliente, fornecedor, veterinario"),
        ("Tipo Pessoa", "Obrigatório. Valores: PF (Pessoa Física) ou PJ (Pessoa Jurídica)"),
        ("Código", "Opcional. Código único para identificação rápida"),
        ("Nome", "Obrigatório. Nome completo (PF) ou Nome Fantasia (PJ)"),
        ("CPF", "Para PF. Formato: XXX.XXX.XXX-XX"),
        ("CNPJ", "Para PJ. Formato: XX.XXX.XXX/0001-XX"),
        ("Razão Social", "Para PJ. Razão social da empresa"),
        ("Responsável", "Opcional. Nome da pessoa de contato"),
        ("CRMV", "Para veterinário. Número do registro no CRMV"),
        ("Inscrição Estadual", "Para PJ. IE da empresa"),
        ("Telefone", "Opcional. Telefone de contato"),
        ("Celular", "Opcional. Celular de contato"),
        ("Email", "Opcional. Email para contato"),
        ("CEP", "Opcional. CEP do endereço"),
        ("Endereço", "Opcional. Rua/Avenida do endereço"),
        ("Número", "Opcional. Número do endereço"),
        ("Complemento", "Opcional. Complemento do endereço (apto, etc)"),
        ("Bairro", "Opcional. Bairro do endereço"),
    ]
    
    for row_num, (campo, descricao) in enumerate(orientacoes, 1):
        cell_a = ws_info.cell(row=row_num, column=1)
        cell_b = ws_info.cell(row=row_num, column=2)
        cell_a.value = campo
        cell_b.value = descricao
        
        if row_num == 1:
            cell_a.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell_b.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell_a.font = Font(bold=True, color="FFFFFF")
            cell_b.font = Font(bold=True, color="FFFFFF")
    
    # Retornar arquivo
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return {
        "filename": "template_importacao_pessoas.xlsx",
        "content": output.getvalue()
    }


@router.post("/importar")
async def importar_pessoas(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user_and_tenant = Depends(get_current_user_and_tenant)
):
    """Importa Pessoas do arquivo Excel"""
    current_user, tenant_id = user_and_tenant
    
    try:
        # Ler arquivo
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        criados = []
        atualizados = []
        erros = []
        
        # Processar linhas (começar da linha 3, pois linha 1 é instrução e 2 é header)
        for row_num in range(3, ws.max_row + 1):
            try:
                # Ler dados da linha
                tipo_cadastro = ws[f'A{row_num}'].value
                tipo_pessoa = ws[f'B{row_num}'].value
                codigo = ws[f'C{row_num}'].value
                nome = ws[f'D{row_num}'].value
                cpf = ws[f'E{row_num}'].value
                cnpj = ws[f'F{row_num}'].value
                razao_social = ws[f'G{row_num}'].value
                responsavel = ws[f'H{row_num}'].value
                crmv = ws[f'I{row_num}'].value
                inscricao_estadual = ws[f'J{row_num}'].value
                telefone = ws[f'K{row_num}'].value
                celular = ws[f'L{row_num}'].value
                email = ws[f'M{row_num}'].value
                cep = ws[f'N{row_num}'].value
                endereco = ws[f'O{row_num}'].value
                numero = ws[f'P{row_num}'].value
                complemento = ws[f'Q{row_num}'].value
                bairro = ws[f'R{row_num}'].value
                
                # Pular linhas vazias
                if not nome or not tipo_cadastro:
                    continue
                
                # Validações
                tipo_cadastro = str(tipo_cadastro).strip().lower()
                if tipo_cadastro not in ['cliente', 'fornecedor', 'veterinario']:
                    raise ValueError(f"Tipo cadastro inválido: {tipo_cadastro}")
                
                tipo_pessoa = str(tipo_pessoa).strip().upper()
                if tipo_pessoa not in ['PF', 'PJ']:
                    raise ValueError(f"Tipo pessoa inválido: {tipo_pessoa}")
                
                # Validar dados obrigatórios por tipo
                if tipo_pessoa == 'PF' and not cpf:
                    raise ValueError("CPF obrigatório para Pessoa Física")
                
                if tipo_pessoa == 'PJ' and not cnpj:
                    raise ValueError("CNPJ obrigatório para Pessoa Jurídica")
                
                # Procurar pessoa existente por CPF/CNPJ ou código
                pessoa_existente = None
                
                if cpf:
                    pessoa_existente = db.query(Cliente).filter(
                        Cliente.cpf == cpf.strip(),
                        Cliente.tenant_id == tenant_id
                    ).first()
                elif cnpj:
                    pessoa_existente = db.query(Cliente).filter(
                        Cliente.cnpj == cnpj.strip(),
                        Cliente.tenant_id == tenant_id
                    ).first()
                elif codigo:
                    pessoa_existente = db.query(Cliente).filter(
                        Cliente.codigo == str(codigo).strip(),
                        Cliente.tenant_id == tenant_id
                    ).first()
                
                # Criar ou atualizar
                if pessoa_existente:
                    # Atualizar
                    pessoa_existente.tipo_cadastro = tipo_cadastro
                    pessoa_existente.tipo_pessoa = tipo_pessoa
                    pessoa_existente.nome = str(nome).strip()
                    
                    if cpf:
                        pessoa_existente.cpf = cpf
                    if cnpj:
                        pessoa_existente.cnpj = cnpj
                    if razao_social:
                        pessoa_existente.razao_social = razao_social
                    if responsavel:
                        pessoa_existente.responsavel = responsavel
                    if crmv:
                        pessoa_existente.crmv = crmv
                    if inscricao_estadual:
                        pessoa_existente.inscricao_estadual = inscricao_estadual
                    if telefone:
                        pessoa_existente.telefone = telefone
                    if celular:
                        pessoa_existente.celular = celular
                    if email:
                        pessoa_existente.email = email
                    if cep:
                        pessoa_existente.cep = cep
                    if endereco:
                        pessoa_existente.endereco = endereco
                    if numero:
                        pessoa_existente.numero = numero
                    if complemento:
                        pessoa_existente.complemento = complemento
                    if bairro:
                        pessoa_existente.bairro = bairro
                    
                    db.add(pessoa_existente)
                    atualizados.append({
                        "linha": row_num,
                        "nome": str(nome).strip(),
                        "tipo": tipo_cadastro
                    })
                else:
                    # Criar novo
                    nova_pessoa = Cliente(
                        tenant_id=tenant_id,
                        codigo=str(codigo).strip() if codigo else None,
                        tipo_cadastro=tipo_cadastro,
                        tipo_pessoa=tipo_pessoa,
                        nome=str(nome).strip(),
                        cpf=cpf.strip() if cpf else None,
                        cnpj=cnpj.strip() if cnpj else None,
                        razao_social=razao_social,
                        responsavel=responsavel,
                        crmv=crmv,
                        inscricao_estadual=inscricao_estadual,
                        telefone=telefone,
                        celular=celular,
                        email=email,
                        cep=cep,
                        endereco=endereco,
                        numero=numero,
                        complemento=complemento,
                        bairro=bairro,
                        ativo=True
                    )
                    db.add(nova_pessoa)
                    criados.append({
                        "linha": row_num,
                        "nome": str(nome).strip(),
                        "tipo": tipo_cadastro
                    })
            
            except Exception as e:
                erros.append({
                    "linha": row_num,
                    "nome": str(ws[f'D{row_num}'].value or "?"),
                    "tipo": str(ws[f'A{row_num}'].value or "?"),
                    "erro": str(e)
                })
        
        # Salvar todas as mudanças
        db.commit()
        
        return {
            "sucesso": True,
            "mensagem": f"Importação concluída: {len(criados)} criados, {len(atualizados)} atualizados",
            "criados": criados,
            "atualizados": atualizados,
            "erros": erros,
            "total_erros": len(erros),
            "total_processados": len(criados) + len(atualizados) + len(erros)
        }
    
    except Exception as e:
        return {
            "sucesso": False,
            "erro": str(e)
        }
