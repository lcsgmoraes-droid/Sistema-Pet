"""
ABA 7: Exportador de DRE para PDF e Excel
Gera relatórios profissionais do DRE
"""

from typing import Optional
from sqlalchemy.orm import Session
from datetime import date
from io import BytesIO
import json

# PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.ia.aba7_models import DREPeriodo, DREProduto, DRECategoria


class ExportadorDRE:
    """Exporta DRE para PDF e Excel com formatação profissional"""
    
    def __init__(self, db: Session):
        self.db = db
    
    def exportar_pdf(
        self,
        dre_periodo_id: int,
        usuario_id: int,
        incluir_produtos: bool = True,
        incluir_categorias: bool = True
    ) -> BytesIO:
        """
        Gera PDF do DRE com formatação profissional
        
        Returns:
            BytesIO buffer com o PDF
        """
        # 1. Buscar dados
        dre = (
            self.db.query(DREPeriodo)
            .filter(
                DREPeriodo.id == dre_periodo_id,
                DREPeriodo.usuario_id == usuario_id
            )
            .first()
        )
        
        if not dre:
            raise ValueError("DRE não encontrado")
        
        # 2. Criar PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=30, leftMargin=30,
                                topMargin=30, bottomMargin=30)
        
        elementos = []
        styles = getSampleStyleSheet()
        
        # Estilo customizado
        titulo_style = ParagraphStyle(
            'Titulo',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e3a8a'),
            alignment=TA_CENTER,
            spaceAfter=12
        )
        
        subtitulo_style = ParagraphStyle(
            'Subtitulo',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#64748b'),
            alignment=TA_CENTER,
            spaceAfter=20
        )
        
        # Título
        elementos.append(Paragraph("Demonstração do Resultado do Exercício (DRE)", titulo_style))
        periodo_texto = f"{dre.data_inicio.strftime('%d/%m/%Y')} a {dre.data_fim.strftime('%d/%m/%Y')}"
        elementos.append(Paragraph(periodo_texto, subtitulo_style))
        elementos.append(Spacer(1, 0.2*inch))
        
        # Tabela principal do DRE
        dados_dre = [
            ['DEMONSTRAÇÃO DO RESULTADO', 'Valor (R$)', '%'],
            ['', '', ''],
            ['RECEITAS', '', ''],
            ['Receita Bruta', f'{dre.receita_bruta:,.2f}', '100,0%'],
            ['(-) Deduções da Receita', f'({dre.deducoes_receita:,.2f})', f'({dre.deducoes_receita/dre.receita_bruta*100:.1f}%)' if dre.receita_bruta > 0 else '0%'],
            ['(=) Receita Líquida', f'{dre.receita_liquida:,.2f}', f'{dre.receita_liquida/dre.receita_bruta*100:.1f}%' if dre.receita_bruta > 0 else '100%'],
            ['', '', ''],
            ['CUSTOS', '', ''],
            ['(-) Custo dos Produtos Vendidos (CMV)', f'({dre.custo_produtos_vendidos:,.2f})', f'({dre.custo_produtos_vendidos/dre.receita_bruta*100:.1f}%)' if dre.receita_bruta > 0 else '0%'],
            ['(=) LUCRO BRUTO', f'{dre.lucro_bruto:,.2f}', f'{dre.margem_bruta_percent:.1f}%'],
            ['', '', ''],
            ['DESPESAS OPERACIONAIS', '', ''],
            ['Despesas com Vendas', f'({dre.despesas_vendas:,.2f})', f'({dre.despesas_vendas/dre.receita_bruta*100:.1f}%)' if dre.receita_bruta > 0 else '0%'],
            ['Despesas Administrativas', f'({dre.despesas_administrativas:,.2f})', f'({dre.despesas_administrativas/dre.receita_bruta*100:.1f}%)' if dre.receita_bruta > 0 else '0%'],
            ['Despesas Financeiras', f'({dre.despesas_financeiras:,.2f})', f'({dre.despesas_financeiras/dre.receita_bruta*100:.1f}%)' if dre.receita_bruta > 0 else '0%'],
            ['Outras Despesas', f'({dre.outras_despesas:,.2f})', f'({dre.outras_despesas/dre.receita_bruta*100:.1f}%)' if dre.receita_bruta > 0 else '0%'],
            ['(=) Total Despesas Operacionais', f'({dre.total_despesas_operacionais:,.2f})', f'({dre.total_despesas_operacionais/dre.receita_bruta*100:.1f}%)' if dre.receita_bruta > 0 else '0%'],
            ['', '', ''],
            ['(=) LUCRO OPERACIONAL', f'{dre.lucro_operacional:,.2f}', f'{dre.margem_operacional_percent:.1f}%'],
        ]
        
        # Adicionar impostos se existir
        if dre.impostos and dre.impostos > 0:
            dados_dre.extend([
                ['', '', ''],
                ['IMPOSTOS', '', ''],
                [f'(-) Impostos ({dre.regime_tributario or ""})', f'({dre.impostos:,.2f})', f'({dre.aliquota_efetiva_percent:.1f}%)'],
            ])
        
        dados_dre.extend([
            ['', '', ''],
            ['(=) LUCRO LÍQUIDO', f'{dre.lucro_liquido:,.2f}', f'{dre.margem_liquida_percent:.1f}%'],
        ])
        
        tabela_dre = Table(dados_dre, colWidths=[4*inch, 1.5*inch, 0.8*inch])
        
        estilo_tabela = TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Conteúdo
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            
            # Linhas de título de seção
            ('FONTNAME', (0, 2), (0, 2), 'Helvetica-Bold'),
            ('FONTNAME', (0, 7), (0, 7), 'Helvetica-Bold'),
            ('FONTNAME', (0, 11), (0, 11), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#f1f5f9')),
            ('BACKGROUND', (0, 7), (-1, 7), colors.HexColor('#f1f5f9')),
            ('BACKGROUND', (0, 11), (-1, 11), colors.HexColor('#f1f5f9')),
            
            # Totais importantes (negrito)
            ('FONTNAME', (0, 5), (-1, 5), 'Helvetica-Bold'),  # Receita Líquida
            ('FONTNAME', (0, 9), (-1, 9), 'Helvetica-Bold'),  # Lucro Bruto
            ('FONTNAME', (0, len(dados_dre)-1), (-1, len(dados_dre)-1), 'Helvetica-Bold'),  # Lucro Líquido
            ('BACKGROUND', (0, len(dados_dre)-1), (-1, len(dados_dre)-1), colors.HexColor('#dcfce7') if dre.lucro_liquido > 0 else colors.HexColor('#fee2e2')),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ])
        
        tabela_dre.setStyle(estilo_tabela)
        elementos.append(tabela_dre)
        
        # Indicadores de performance
        elementos.append(Spacer(1, 0.3*inch))
        elementos.append(Paragraph("<b>Indicadores de Performance</b>", styles['Heading2']))
        elementos.append(Spacer(1, 0.1*inch))
        
        status_color = colors.green if dre.status == "lucro" else colors.red if dre.status == "prejuizo" else colors.orange
        status_texto = dre.status.upper()
        
        dados_indicadores = [
            ['Indicador', 'Valor'],
            ['Status Financeiro', status_texto],
            ['Score de Saúde', f'{dre.score_saude}/100 pontos'],
            ['Margem Bruta', f'{dre.margem_bruta_percent:.2f}%'],
            ['Margem Operacional', f'{dre.margem_operacional_percent:.2f}%'],
            ['Margem Líquida', f'{dre.margem_liquida_percent:.2f}%'],
        ]
        
        tabela_indicadores = Table(dados_indicadores, colWidths=[3*inch, 2*inch])
        tabela_indicadores.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 11),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ]))
        
        elementos.append(tabela_indicadores)
        
        # Produtos (se solicitado)
        if incluir_produtos:
            produtos = (
                self.db.query(DREProduto)
                .filter(DREProduto.dre_periodo_id == dre_periodo_id)
                .order_by(DREProduto.ranking_rentabilidade)
                .limit(10)
                .all()
            )
            
            if produtos:
                elementos.append(PageBreak())
                elementos.append(Paragraph("<b>Top 10 Produtos Mais Rentáveis</b>", styles['Heading2']))
                elementos.append(Spacer(1, 0.1*inch))
                
                dados_produtos = [['#', 'Produto', 'Qtd', 'Receita', 'Margem %']]
                for prod in produtos:
                    dados_produtos.append([
                        str(prod.ranking_rentabilidade),
                        prod.produto_nome[:40],
                        str(prod.quantidade_vendida),
                        f'R$ {prod.receita_total:,.2f}',
                        f'{prod.margem_percent:.1f}%'
                    ])
                
                tabela_produtos = Table(dados_produtos, colWidths=[0.4*inch, 2.8*inch, 0.7*inch, 1.2*inch, 0.9*inch])
                tabela_produtos.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                ]))
                
                elementos.append(tabela_produtos)
        
        # Gerar PDF
        doc.build(elementos)
        buffer.seek(0)
        return buffer
    
    def exportar_excel(
        self,
        dre_periodo_id: int,
        usuario_id: int
    ) -> BytesIO:
        """
        Gera planilha Excel do DRE
        
        Returns:
            BytesIO buffer com o Excel
        """
        # 1. Buscar dados
        dre = (
            self.db.query(DREPeriodo)
            .filter(
                DREPeriodo.id == dre_periodo_id,
                DREPeriodo.usuario_id == usuario_id
            )
            .first()
        )
        
        if not dre:
            raise ValueError("DRE não encontrado")
        
        # 2. Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "DRE"
        
        # Estilos
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        bold_font = Font(bold=True)
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws['A1'] = "DEMONSTRAÇÃO DO RESULTADO DO EXERCÍCIO (DRE)"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        periodo = f"{dre.data_inicio.strftime('%d/%m/%Y')} a {dre.data_fim.strftime('%d/%m/%Y')}"
        ws['A2'] = periodo
        ws.merge_cells('A2:D2')
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Header da tabela
        row = 4
        ws[f'A{row}'] = "DESCRIÇÃO"
        ws[f'B{row}'] = "VALOR (R$)"
        ws[f'C{row}'] = "%"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
        
        # Dados
        row += 1
        dados = [
            ("", "", ""),
            ("RECEITAS", "", ""),
            ("Receita Bruta", dre.receita_bruta, "100,0%"),
            ("(-) Deduções", -dre.deducoes_receita, f"-{dre.deducoes_receita/dre.receita_bruta*100:.1f}%"),
            ("(=) Receita Líquida", dre.receita_liquida, f"{dre.receita_liquida/dre.receita_bruta*100:.1f}%"),
            ("", "", ""),
            ("CUSTOS", "", ""),
            ("(-) CMV", -dre.custo_produtos_vendidos, f"-{dre.custo_produtos_vendidos/dre.receita_bruta*100:.1f}%"),
            ("(=) LUCRO BRUTO", dre.lucro_bruto, f"{dre.margem_bruta_percent:.1f}%"),
            ("", "", ""),
            ("DESPESAS OPERACIONAIS", "", ""),
            ("Despesas com Vendas", -dre.despesas_vendas, f"-{dre.despesas_vendas/dre.receita_bruta*100:.1f}%"),
            ("Despesas Administrativas", -dre.despesas_administrativas, f"-{dre.despesas_administrativas/dre.receita_bruta*100:.1f}%"),
            ("Despesas Financeiras", -dre.despesas_financeiras, f"-{dre.despesas_financeiras/dre.receita_bruta*100:.1f}%"),
            ("Outras Despesas", -dre.outras_despesas, f"-{dre.outras_despesas/dre.receita_bruta*100:.1f}%"),
            ("(=) Total Despesas", -dre.total_despesas_operacionais, f"-{dre.total_despesas_operacionais/dre.receita_bruta*100:.1f}%"),
            ("", "", ""),
            ("(=) LUCRO OPERACIONAL", dre.lucro_operacional, f"{dre.margem_operacional_percent:.1f}%"),
        ]
        
        if dre.impostos and dre.impostos > 0:
            dados.extend([
                ("", "", ""),
                ("IMPOSTOS", "", ""),
                (f"(-) Impostos ({dre.regime_tributario or ''})", -dre.impostos, f"-{dre.aliquota_efetiva_percent:.1f}%"),
            ])
        
        dados.append(("", "", ""))
        dados.append(("(=) LUCRO LÍQUIDO", dre.lucro_liquido, f"{dre.margem_liquida_percent:.1f}%"))
        
        for desc, valor, perc in dados:
            ws[f'A{row}'] = desc
            ws[f'B{row}'] = valor if isinstance(valor, (int, float)) else ""
            ws[f'C{row}'] = perc
            
            # Formatação
            if "=" in desc:  # Linhas de total
                ws[f'A{row}'].font = bold_font
                ws[f'B{row}'].font = bold_font
                ws[f'C{row}'].font = bold_font
            
            if isinstance(valor, (int, float)):
                ws[f'B{row}'].number_format = '#,##0.00'
            
            row += 1
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        
        # Gerar buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
